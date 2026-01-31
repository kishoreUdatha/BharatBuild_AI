from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

file_path = r'C:\Users\Lekhana\Downloads\SmartGrow_Infotech_Operating_Costs_FILLED.xlsx'

wb = load_workbook(file_path)

# Correct pricing
project_price = 9999
cost_per_project = 4500
profit_per_project = project_price - cost_per_project  # 5499
monthly_operating_cost = 3280000

# Break-even calculation
breakeven_projects = int(monthly_operating_cost / profit_per_project) + 1
margin_pct = round((profit_per_project/project_price)*100, 1)

print("=" * 60)
print("CORRECTED PRICING")
print("=" * 60)
print(f"Project Price: Rs {project_price}")
print(f"Cost per Project: Rs {cost_per_project}")
print(f"Profit per Project: Rs {profit_per_project}")
print(f"Gross Margin: {margin_pct}%")
print(f"Break-even: {breakeven_projects} projects/month")

# ============ Update Cost_Per_User_Calculator Sheet ============
ws3 = wb['Cost_Per_User_Calculator']

ws3.cell(row=2, column=1, value="Project Price (INR)")
ws3.cell(row=2, column=2, value=project_price)
ws3.cell(row=2, column=3, value="One-time per project")

ws3.cell(row=3, column=1, value="Agents per Project")
ws3.cell(row=3, column=2, value=3)
ws3.cell(row=3, column=3, value="AI agents working on each project")

ws3.cell(row=4, column=1, value="LLM Cost (INR)")
ws3.cell(row=4, column=2, value=2500)
ws3.cell(row=4, column=3, value="Claude/GPT API tokens")

ws3.cell(row=5, column=1, value="Cloud Cost (INR)")
ws3.cell(row=5, column=2, value=1500)
ws3.cell(row=5, column=3, value="Compute, storage, CDN")

ws3.cell(row=6, column=1, value="Support/Overhead (INR)")
ws3.cell(row=6, column=2, value=500)
ws3.cell(row=6, column=3, value="Tools, support, misc")

ws3.cell(row=7, column=1, value="Total Cost per Project (INR)")
ws3.cell(row=7, column=2, value=cost_per_project)
ws3.cell(row=7, column=3, value="Sum of all costs")

ws3.cell(row=8, column=1, value="Profit per Project (INR)")
ws3.cell(row=8, column=2, value=profit_per_project)
ws3.cell(row=8, column=3, value="Price - Cost")

ws3.cell(row=9, column=1, value="Gross Margin (%)")
ws3.cell(row=9, column=2, value=margin_pct)
ws3.cell(row=9, column=3, value="Profit / Price x 100")

# ============ Update Pricing_Tiers Sheet ============
ws4 = wb['Pricing_Tiers']

tiers = [
    ("Student", "Students", "1", 2500, 50, 4999, "Basic project, limited features"),
    ("Basic", "Individuals", "1", 3500, 53, 7499, "Standard project, 3 revisions"),
    ("Standard", "Professionals", "1", 4500, 55, 9999, "Full project, 3 agents, unlimited revisions"),
    ("Premium", "Businesses", "3", 4200, 58, 27999, "3 projects, priority support"),
    ("Enterprise", "Large Orgs", "10", 4000, 60, 89999, "10 projects, dedicated support, SLA"),
]

for i, (name, segment, projects, cost, margin, price, inclusions) in enumerate(tiers, start=2):
    ws4.cell(row=i, column=1, value=name)
    ws4.cell(row=i, column=2, value=segment)
    ws4.cell(row=i, column=3, value=projects)
    ws4.cell(row=i, column=4, value=cost)
    ws4.cell(row=i, column=5, value=margin)
    ws4.cell(row=i, column=6, value=price)
    ws4.cell(row=i, column=7, value=inclusions)

# ============ Recreate User_Analysis Sheet ============
if 'User_Analysis' in wb.sheetnames:
    del wb['User_Analysis']
ws5 = wb.create_sheet('User_Analysis')

# Styles
header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
section_font = Font(bold=True, color="A5B4FC", size=14)
profit_font = Font(bold=True, color="22C55E", size=11)
loss_font = Font(bold=True, color="EF4444", size=11)

row = 1

# Section 1: Key Metrics
ws5.cell(row=row, column=1, value="KEY METRICS").font = section_font
row += 1

headers = ["Metric", "Value", "Unit", "Notes"]
for col, header in enumerate(headers, 1):
    cell = ws5.cell(row=row, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
row += 1

metrics = [
    ("Monthly Operating Cost", monthly_operating_cost, "INR", "Fixed + Variable costs"),
    ("Project Price", project_price, "INR", "Customer pays"),
    ("Cost per Project", cost_per_project, "INR", "LLM + Cloud + Support"),
    ("Profit per Project", profit_per_project, "INR", "Price - Cost"),
    ("Gross Margin", margin_pct, "%", "Profit / Price"),
    ("Break-even Projects", breakeven_projects, "projects/month", "Minimum to cover costs"),
    ("Break-even Revenue", breakeven_projects * project_price, "INR/month", f"{breakeven_projects} x {project_price}"),
]

for metric, value, unit, notes in metrics:
    ws5.cell(row=row, column=1, value=metric)
    ws5.cell(row=row, column=2, value=value)
    ws5.cell(row=row, column=3, value=unit)
    ws5.cell(row=row, column=4, value=notes)
    row += 1

row += 1

# Section 2: Profitability
ws5.cell(row=row, column=1, value="PROFITABILITY BY PROJECT VOLUME").font = section_font
row += 1

headers = ["Scenario", "Projects/Month", "Revenue (INR)", "Profit (INR)", "Status"]
for col, header in enumerate(headers, 1):
    cell = ws5.cell(row=row, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
row += 1

scenarios = [
    ("Break-even", 597),
    ("Small Scale", 750),
    ("Medium Scale", 1000),
    ("Growth Phase", 1500),
    ("Scaling", 2000),
    ("Large Scale", 3000),
    ("Enterprise", 5000),
]

for name, projects in scenarios:
    revenue = projects * project_price
    variable_cost = projects * cost_per_project
    profit = revenue - monthly_operating_cost - variable_cost
    status = "Profit" if profit > 0 else "Break-even" if abs(profit) < 50000 else "Loss"

    ws5.cell(row=row, column=1, value=name)
    ws5.cell(row=row, column=2, value=projects)
    ws5.cell(row=row, column=3, value=revenue)
    cell = ws5.cell(row=row, column=4, value=profit)
    cell.font = profit_font if profit > 0 else loss_font
    ws5.cell(row=row, column=5, value=status)
    row += 1

row += 1

# Section 3: Users
ws5.cell(row=row, column=1, value="USERS TO PROJECTS (10% Conversion Rate)").font = section_font
row += 1

headers = ["Total Users", "Paying Users (10%)", "Projects/Month", "Monthly Profit (INR)", "Status"]
for col, header in enumerate(headers, 1):
    cell = ws5.cell(row=row, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
row += 1

user_list = [6000, 7500, 10000, 15000, 20000, 30000, 50000]

for users in user_list:
    paying = int(users * 0.10)
    revenue = paying * project_price
    variable_cost = paying * cost_per_project
    profit = revenue - monthly_operating_cost - variable_cost
    status = "Profit" if profit > 0 else "Loss"

    ws5.cell(row=row, column=1, value=users)
    ws5.cell(row=row, column=2, value=paying)
    ws5.cell(row=row, column=3, value=paying)
    cell = ws5.cell(row=row, column=4, value=profit)
    cell.font = profit_font if profit > 0 else loss_font
    ws5.cell(row=row, column=5, value=status)
    row += 1

row += 1

# Section 4: Summary
ws5.cell(row=row, column=1, value="SUMMARY & TARGETS").font = section_font
row += 1

headers = ["Target", "Value", "Notes", ""]
for col, header in enumerate(headers, 1):
    cell = ws5.cell(row=row, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
row += 1

targets = [
    ("Minimum Users (Break-even)", "6,000", "At 10% conversion = 600 projects"),
    ("Target Users (Profitable)", "7,500+", "750 projects = Rs 8.4L profit"),
    ("Ideal Users (Growth)", "10,000+", "1000 projects = Rs 22L profit"),
    ("Minimum Projects/Month", str(breakeven_projects), "Break-even point"),
    ("Target Projects/Month", "750+", "Healthy profit margin"),
    ("Revenue Target (Monthly)", "Rs 75 Lakhs+", "750 projects x 9999"),
]

for target, value, notes in targets:
    ws5.cell(row=row, column=1, value=target)
    ws5.cell(row=row, column=2, value=value)
    ws5.cell(row=row, column=3, value=notes)
    row += 1

# Column widths
ws5.column_dimensions['A'].width = 35
ws5.column_dimensions['B'].width = 20
ws5.column_dimensions['C'].width = 20
ws5.column_dimensions['D'].width = 25
ws5.column_dimensions['E'].width = 15

wb.save(file_path)

print("\n--- Profitability ---")
for name, projects in scenarios:
    revenue = projects * project_price
    variable_cost = projects * cost_per_project
    profit = revenue - monthly_operating_cost - variable_cost
    print(f"{name:<15} {projects:>5} projects -> Profit: Rs {profit:>12,}")

print("\n--- Users Needed ---")
for users in user_list:
    paying = int(users * 0.10)
    profit = (paying * project_price) - monthly_operating_cost - (paying * cost_per_project)
    status = "Profit" if profit > 0 else "Loss"
    print(f"{users:>6,} users -> {paying:>4} projects -> Rs {profit:>12,} ({status})")

print(f"\nFile saved: {file_path}")
