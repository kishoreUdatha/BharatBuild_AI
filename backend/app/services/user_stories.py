"""
User Stories - a batch's product backlog, once AI planning has let it through.

The scope is the backlog rather than every drafted story: a story appears here
after a trainer approved it and moved it across. That is the same gate the AI
Story Approval screen enforces, seen from the other side, and it is why a
batch still in review shows an empty board with a pointer back to planning
instead of work nobody has agreed to yet.

Two rules the numbers on the screen depend on:

* Every count is taken from the rows the screen is showing. Filter the list
  and the tiles follow, so the header, the tiles and the per-student cards can
  never disagree with the table.
* An assignee must be an active member of the batch. The dropdown is built
  from the roster, and so is the check - a story cannot be handed to somebody
  who is not on the team by posting a different id.
"""

import csv
import io
import re
from collections import defaultdict
from datetime import date, datetime
from math import ceil
from typing import Dict, List, Optional, Tuple

from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.logging_config import logger
from app.models.ai_planning import (
    CriterionKind,
    ProjectEpic,
    ProjectUserStory,
    StoryCriterion,
    StoryPriority,
    StoryReviewStatus,
)
from app.models.backlog import (
    ProjectSprint,
    SprintState,
    StoryAttachment,
    StoryComment,
    StoryEvent,
    StoryEventKind,
    StoryType,
    StoryWorkflowStatus,
)
from app.models.faculty import ProjectBatch, ProjectBatchMember
from app.services import file_store
from app.models.project_tracking import ProjectTask, TaskStatus
from app.models.story_commit import StoryCommit
from app.models.user import User

# Declaration order is board order and sort order, so it reads left to right
# as the work actually moves. Blocked comes last on purpose: it is a holding
# pen beside the flow, not a step within it.
WORKFLOW_LABELS = {
    StoryWorkflowStatus.TO_DO: "To Do",
    StoryWorkflowStatus.IN_PROGRESS: "In Progress",
    StoryWorkflowStatus.TESTING: "Testing",
    StoryWorkflowStatus.IN_REVIEW: "In Review",
    StoryWorkflowStatus.DONE: "Done",
    StoryWorkflowStatus.BLOCKED: "Blocked",
}

TYPE_LABELS = {
    StoryType.STORY: "Story",
    StoryType.TASK: "Task",
    StoryType.BUG: "Bug",
    StoryType.SPIKE: "Spike",
}

SPRINT_STATE_LABELS = {
    SprintState.PLANNED: "Planned",
    SprintState.ACTIVE: "Active",
    SprintState.COMPLETED: "Completed",
}

REVIEW_LABELS = {
    StoryReviewStatus.NEEDS_REVIEW: "Needs Review",
    StoryReviewStatus.REVIEWED: "Reviewed",
    StoryReviewStatus.APPROVED: "Approved",
    StoryReviewStatus.REJECTED: "Rejected",
    StoryReviewStatus.REVISION_REQUESTED: "Revision Requested",
}

assert set(WORKFLOW_LABELS) == set(StoryWorkflowStatus)
assert set(TYPE_LABELS) == set(StoryType)

# Sorting the table. The default matches the screen: newest first.
SORTS = {
    "created_desc": ("Created Date (Newest)",
                     lambda r: (r["created_at"] or datetime.min, r["key"]), True),
    "created_asc": ("Created Date (Oldest)",
                    lambda r: (r["created_at"] or datetime.min, r["key"]), False),
    "key": ("Story ID", lambda r: r["key"], False),
    "points": ("Story Points", lambda r: r["story_points"] or 0, True),
    "priority": ("Priority", lambda r: {"high": 0, "medium": 1, "low": 2}[r["priority"]], False),
    "status": ("Status",
               lambda r: list(WORKFLOW_LABELS).index(StoryWorkflowStatus(r["status"])), False),
}

# The trainer's import template, as issued: one row per story, twelve columns.
# The tuple is (heading, other spellings we accept), so a hand-made CSV with
# plainer names still imports and the template stays the documented shape.
IMPORT_FIELDS = {
    "story_id": ("Story ID", ("story id", "story_id", "key", "id")),
    "work_type": ("Work Type", ("work type", "work_type", "type")),
    "epic": ("Epic", ("epic", "epic name", "feature")),
    "summary": ("Summary", ("summary", "title", "user story title")),
    "description": ("Description", ("description", "narrative", "user story")),
    "acceptance": ("Acceptance Criteria",
                   ("acceptance criteria", "acceptance_criteria", "acceptance")),
    "priority": ("Priority", ("priority",)),
    "points": ("Story Points", ("story points", "story_points", "points", "estimate")),
    "roll": ("Assignee Roll No",
             ("assignee roll no", "assignee_roll_no", "roll no", "roll number", "roll")),
    "assignee": ("Assignee Name", ("assignee name", "assignee_name", "assignee")),
    "sprint": ("Sprint", ("sprint",)),
    "status": ("Status", ("status",)),
    "labels": ("Labels", ("labels", "tags")),
}

IMPORT_COLUMNS = [heading for heading, _ in IMPORT_FIELDS.values()]

# The template offers five priorities against our three. Highest and Lowest
# fold into High and Low rather than being refused - a sheet written to the
# template must import, and the extra step of the scale is not one this
# product tracks.
PRIORITY_WORDS = {
    "highest": StoryPriority.HIGH, "high": StoryPriority.HIGH,
    "medium": StoryPriority.MEDIUM, "normal": StoryPriority.MEDIUM,
    "low": StoryPriority.LOW, "lowest": StoryPriority.LOW,
}

TYPE_WORDS = {
    "story": StoryType.STORY, "user story": StoryType.STORY,
    "task": StoryType.TASK, "spike": StoryType.SPIKE,
    "bug": StoryType.BUG, "defect": StoryType.BUG,
}

STATUS_WORDS = {
    "to do": StoryWorkflowStatus.TO_DO, "todo": StoryWorkflowStatus.TO_DO,
    "to_do": StoryWorkflowStatus.TO_DO, "backlog": StoryWorkflowStatus.TO_DO,
    "in progress": StoryWorkflowStatus.IN_PROGRESS,
    "in_progress": StoryWorkflowStatus.IN_PROGRESS,
    "in review": StoryWorkflowStatus.IN_REVIEW,
    "in_review": StoryWorkflowStatus.IN_REVIEW,
    "done": StoryWorkflowStatus.DONE, "completed": StoryWorkflowStatus.DONE,
}

# "1. Valid credentials allow login" -> "Valid credentials allow login". The
# template asks for numbered lines in one cell, and the number is formatting.
_BULLET = re.compile(r"^\s*(?:\(?\d+[.)]|[-*\u2022])\s*")


def _criteria_lines(cell: str) -> List[str]:
    """One criterion per line, or per pipe or semicolon when it is all one line."""
    text = (cell or "").replace("\r", "\n")
    parts = [p for p in text.split("\n") if p.strip()]
    if len(parts) <= 1:
        parts = re.split(r"[|;]", text)
    return [_BULLET.sub("", p).strip() for p in parts if p.strip()]



class StoryError(Exception):
    """A refusal the caller can show the trainer as-is."""


def _pct(part: int, whole: int) -> float:
    return round(part / whole * 100, 1) if whole else 0.0


def _initials(name: Optional[str]) -> str:
    parts = [p for p in (name or "").split() if p]
    return "".join(p[0] for p in parts[:2]).upper() or "?"


def _person(user: Optional[User]) -> Optional[dict]:
    if user is None:
        return None
    name = user.full_name or (user.email or "").split("@")[0]
    return {
        "id": str(user.id),
        "name": name,
        "roll": user.roll_number,
        "initials": _initials(name),
    }


def _window(sprint: Optional[ProjectSprint]) -> Optional[str]:
    """'20 May - 02 Jun 2025', or None when the dates were never set."""
    if sprint is None or not sprint.start_date or not sprint.end_date:
        return None
    start, end = sprint.start_date, sprint.end_date
    left = start.strftime("%d %b") if start.year == end.year else start.strftime("%d %b %Y")
    return f"{left} - {end.strftime('%d %b %Y')}"


def _sprint(sprint: Optional[ProjectSprint]) -> Optional[dict]:
    if sprint is None:
        return None
    return {
        "id": str(sprint.id),
        "key": sprint.key,
        "name": sprint.name,
        "goal": sprint.goal,
        "state": sprint.state.value,
        "state_label": SPRINT_STATE_LABELS[sprint.state],
        "window": _window(sprint),
        "start_date": sprint.start_date,
        "end_date": sprint.end_date,
    }


class UserStoryService:
    def __init__(self, db: AsyncSession):
        self.db = db

    # -------------------------------------------------------------- loading

    async def load_batch(self, identifier: str) -> Optional[ProjectBatch]:
        # The roster is what the assignee dropdown and the per-student cards
        # are built from, and lazy-loading it under async raises MissingGreenlet.
        key = (identifier or "").strip()
        return (await self.db.execute(
            select(ProjectBatch)
            .where((ProjectBatch.batch_code == key) | (ProjectBatch.join_code == key.upper()))
            .options(
                selectinload(ProjectBatch.guide),
                selectinload(ProjectBatch.members).selectinload(ProjectBatchMember.student),
            )
        )).scalars().first()

    async def _backlog(self, batch_id) -> List[ProjectUserStory]:
        """Only what a trainer approved and moved across. That is the backlog."""
        return (await self.db.execute(
            select(ProjectUserStory)
            .where(ProjectUserStory.batch_id == batch_id)
            .where(ProjectUserStory.moved_to_backlog_at.isnot(None))
            .options(
                selectinload(ProjectUserStory.epic),
                selectinload(ProjectUserStory.sprint),
                selectinload(ProjectUserStory.assignee),
                selectinload(ProjectUserStory.created_by),
            )
            .order_by(ProjectUserStory.key)
        )).scalars().all()

    async def _planning_state(self, batch_id) -> dict:
        """What is still upstream in AI planning, so the empty state can say so."""
        rows = (await self.db.execute(
            select(ProjectUserStory.review_status, ProjectUserStory.moved_to_backlog_at)
            .where(ProjectUserStory.batch_id == batch_id)
        )).all()
        return {
            "drafted": len(rows),
            "needs_review": sum(1 for s, _ in rows if s == StoryReviewStatus.NEEDS_REVIEW),
            "awaiting_move": sum(1 for s, moved in rows
                                 if moved is None and s == StoryReviewStatus.APPROVED),
        }

    async def sprints(self, batch_id) -> List[ProjectSprint]:
        return (await self.db.execute(
            select(ProjectSprint)
            .where(ProjectSprint.batch_id == batch_id)
            .order_by(ProjectSprint.position, ProjectSprint.key)
        )).scalars().all()

    async def _epics(self, batch_id) -> List[ProjectEpic]:
        return (await self.db.execute(
            select(ProjectEpic)
            .where(ProjectEpic.batch_id == batch_id)
            .order_by(ProjectEpic.position)
        )).scalars().all()

    @staticmethod
    def _roster(batch: ProjectBatch) -> List[ProjectBatchMember]:
        return [m for m in batch.members if m.is_active and m.student is not None]

    # ------------------------------------------------------------ fragments

    def _row(self, story: ProjectUserStory) -> dict:
        return {
            "id": str(story.id),
            "key": story.key,
            "title": story.title,
            "type": story.story_type.value,
            "type_label": TYPE_LABELS[story.story_type],
            "epic_key": story.epic.key if story.epic else None,
            "epic_title": story.epic.title if story.epic else None,
            "assignee": _person(story.assignee),
            "sprint": _sprint(story.sprint),
            "priority": story.priority.value,
            "priority_label": story.priority.value.title(),
            "story_points": story.story_points or 0,
            "position": story.position or 0,
            "status": story.workflow_status.value,
            "status_label": WORKFLOW_LABELS[story.workflow_status],
            "review_status": story.review_status.value,
            "review_status_label": REVIEW_LABELS[story.review_status],
            "labels": ([l for l in (story.labels or "").split(",") if l]),
            "due_date": story.due_date,
            # Derived on every read: a stored "late" flag is wrong from the
            # next morning, and a story that is done is not late any more.
            "overdue": bool(story.due_date and story.due_date < date.today()
                            and story.workflow_status != StoryWorkflowStatus.DONE),
            "created_at": story.created_at,
            # The AI-drafted set has no creator, and saying so is more honest
            # than crediting whoever happened to approve it.
            "created_by": (story.created_by.full_name if story.created_by else "AI Planning"),
        }

    async def _detail(self, story: ProjectUserStory) -> dict:
        criteria = (await self.db.execute(
            select(StoryCriterion)
            .where(StoryCriterion.story_id == story.id)
            .order_by(StoryCriterion.position)
        )).scalars().all()
        tasks = (await self.db.execute(
            select(ProjectTask)
            .where(ProjectTask.story_id == story.id)
            .options(selectinload(ProjectTask.assignee))
            .order_by(ProjectTask.created_at)
        )).scalars().all()
        comments = (await self.db.execute(
            select(StoryComment)
            .where(StoryComment.story_id == story.id)
            .order_by(StoryComment.created_at.desc())
        )).scalars().all()
        events = (await self.db.execute(
            select(StoryEvent)
            .where(StoryEvent.story_id == story.id)
            .order_by(StoryEvent.occurred_at.desc())
        )).scalars().all()
        attachments = (await self.db.execute(
            select(StoryAttachment)
            .where(StoryAttachment.story_id == story.id)
            .options(selectinload(StoryAttachment.file),
                     selectinload(StoryAttachment.uploaded_by))
            .order_by(StoryAttachment.uploaded_at.desc())
        )).scalars().all()

        def of_kind(kind: CriterionKind) -> List[dict]:
            return [{"id": str(c.id), "text": c.text, "met": c.met}
                    for c in criteria if c.kind == kind]

        return {
            **self._row(story),
            "narrative": story.narrative,
            "dependencies": story.dependencies,
            "trainer_comment": story.trainer_comment,
            "ai_confidence": int(story.ai_confidence) if story.ai_confidence is not None else None,
            "started_at": story.started_at,
            "completed_at": story.completed_at,
            "updated_at": story.updated_at,
            "acceptance_criteria": of_kind(CriterionKind.ACCEPTANCE),
            "definition_of_done": of_kind(CriterionKind.DEFINITION_OF_DONE),
            "tasks": [{
                "id": str(t.id),
                "title": t.title,
                "status": t.status.value,
                "status_label": t.status.value.replace("_", " ").title(),
                "done": t.status == TaskStatus.DONE,
                "assignee": _person(t.assignee),
                "due_date": t.due_date,
            } for t in tasks],
            "comments": [{
                "id": str(c.id),
                "author": c.author_name or "Unknown",
                "body": c.body,
                "created_at": c.created_at,
            } for c in comments],
            "activity": [{
                "id": str(e.id),
                "kind": e.kind.value,
                "actor": e.actor_name or "System",
                "summary": e.summary,
                "from_value": e.from_value,
                "to_value": e.to_value,
                "occurred_at": e.occurred_at,
            } for e in events],
            "attachments": [{
                "id": str(a.id),
                "name": a.name,
                "mime": a.file.mime_type if a.file else None,
                "size": a.file.byte_size if a.file else 0,
                "size_label": (file_store.human_size(a.file.byte_size)
                               if a.file else "—"),
                "uploaded_by": (a.uploaded_by.full_name if a.uploaded_by else None),
                "uploaded_by_id": str(a.uploaded_by_id) if a.uploaded_by_id else None,
                "uploaded_at": a.uploaded_at,
            } for a in attachments],
            "counts": {
                "tasks": len(tasks),
                "tasks_done": sum(1 for t in tasks if t.status == TaskStatus.DONE),
                "comments": len(comments),
                "activity": len(events),
                "attachments": len(attachments),
            },
        }

    # ----------------------------------------------------------------- view

    async def board(
        self,
        batch: ProjectBatch,
        *,
        search: Optional[str] = None,
        status: Optional[str] = None,
        epic: Optional[str] = None,
        assignee: Optional[str] = None,
        sprint: Optional[str] = None,
        priority: Optional[str] = None,
        points: Optional[str] = None,
        story_type: Optional[str] = None,
        created_by: Optional[str] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        sort: str = "created_desc",
        page: int = 1,
        per_page: int = 10,
        selected: Optional[str] = None,
    ) -> dict:
        stories = await self._backlog(batch.id)
        by_id = {str(s.id): s for s in stories}
        rows = [self._row(s) for s in stories]

        def keep(r: dict) -> bool:
            if status and r["status"] != status:
                return False
            if epic and r["epic_key"] != epic:
                return False
            if assignee:
                # "Unassigned" is a real answer to "who is doing this", and the
                # one a trainer most often wants to see.
                if assignee == "unassigned":
                    if r["assignee"] is not None:
                        return False
                elif not r["assignee"] or r["assignee"]["id"] != assignee:
                    return False
            if sprint:
                if sprint == "unscheduled":
                    if r["sprint"] is not None:
                        return False
                elif not r["sprint"] or r["sprint"]["id"] != sprint:
                    return False
            if priority and r["priority"] != priority:
                return False
            if points and str(r["story_points"]) != str(points):
                return False
            if story_type and r["type"] != story_type:
                return False
            if created_by and r["created_by"] != created_by:
                return False
            created = r["created_at"]
            if date_from and created and created.date().isoformat() < date_from:
                return False
            if date_to and created and created.date().isoformat() > date_to:
                return False
            if search:
                needle = search.lower()
                story = by_id[r["id"]]
                blob = " ".join(filter(None, [
                    r["key"], r["title"], story.narrative, r["epic_title"],
                    story.labels,
                    r["assignee"]["name"] if r["assignee"] else None,
                ])).lower()
                if needle not in blob:
                    return False
            return True

        matched = [r for r in rows if keep(r)]

        _, key_of, reverse = SORTS.get(sort, SORTS["created_desc"])
        matched.sort(key=key_of, reverse=reverse)

        total = len(matched)
        per_page = max(1, min(per_page, 200))
        pages = max(1, ceil(total / per_page)) if total else 1
        page = max(1, min(page, pages))
        window = matched[(page - 1) * per_page: page * per_page]

        # The panel follows the table: an explicit selection wins, otherwise the
        # first row on this page, so the screen never opens with a blank panel.
        chosen = by_id.get(selected or "")
        if chosen is None or not any(r["id"] == str(chosen.id) for r in matched):
            chosen = by_id.get(window[0]["id"]) if window else None

        counted = {s: sum(1 for r in matched if r["status"] == s.value)
                   for s in StoryWorkflowStatus}
        story_points = sum(r["story_points"] for r in matched)
        epics = await self._epics(batch.id)
        sprint_rows = await self.sprints(batch.id)

        return {
            "header": {
                "batch_id": str(batch.id),
                "batch_code": batch.batch_code,
                "project_title": batch.title,
                "department": batch.department,
                "section": batch.section,
                "guide": batch.guide.full_name if batch.guide else None,
                "members": len(self._roster(batch)),
            },
            "kpis": [
                {"id": "total", "value": total, "label": "Total Stories", "percent": None},
                {"id": "to_do", "value": counted[StoryWorkflowStatus.TO_DO],
                 "label": "To Do", "percent": _pct(counted[StoryWorkflowStatus.TO_DO], total)},
                {"id": "in_progress", "value": counted[StoryWorkflowStatus.IN_PROGRESS],
                 "label": "In Progress",
                 "percent": _pct(counted[StoryWorkflowStatus.IN_PROGRESS], total)},
                {"id": "in_review", "value": counted[StoryWorkflowStatus.IN_REVIEW],
                 "label": "In Review",
                 "percent": _pct(counted[StoryWorkflowStatus.IN_REVIEW], total)},
                {"id": "done", "value": counted[StoryWorkflowStatus.DONE],
                 "label": "Done", "percent": _pct(counted[StoryWorkflowStatus.DONE], total)},
                {"id": "points", "value": story_points, "label": "Total Story Points",
                 "percent": None},
            ],
            "rows": window,
            "selected": await self._detail(chosen) if chosen is not None else None,
            "total": total,
            "page": page,
            "pages": pages,
            "per_page": per_page,
            "sort": sort,
            "backlog_total": len(stories),
            "counts": {
                "total": total,
                "showing": len(window),
                "story_points": story_points,
                **{s.value: counted[s] for s in StoryWorkflowStatus},
            },
            "filters": {
                "epics": [{"key": e.key, "title": e.title} for e in epics],
                "sprints": [_sprint(s) for s in sprint_rows],
                "assignees": [
                    {**_person(m.student), "responsibility": m.responsibility}
                    for m in self._roster(batch)
                ],
                "statuses": [{"value": s.value, "label": WORKFLOW_LABELS[s]}
                             for s in StoryWorkflowStatus],
                "priorities": [{"value": p.value, "label": p.value.title()}
                               for p in StoryPriority],
                "types": [{"value": t.value, "label": TYPE_LABELS[t]} for t in StoryType],
                "points": sorted({r["story_points"] for r in rows}),
                "creators": sorted({r["created_by"] for r in rows if r["created_by"]}),
                "sorts": [{"value": k, "label": v[0]} for k, v in SORTS.items()],
            },
            "students": self._student_summaries(batch, rows),
            # Nothing has been approved and moved across yet, so the screen says
            # what is holding the backlog up rather than showing a bare zero.
            "planning": await self._planning_state(batch.id) if not stories else None,
        }

    def _student_summaries(self, batch: ProjectBatch, rows: List[dict]) -> List[dict]:
        """
        One card per member of the roster - including members holding nothing.

        A student with no stories is the useful signal on this row of cards, so
        the cards come from the roster rather than from the stories that happen
        to have an assignee.
        """
        held: Dict[str, dict] = defaultdict(lambda: {"stories": 0, "points": 0, "done": 0})
        for r in rows:
            who = r["assignee"]
            if who is None:
                continue
            entry = held[who["id"]]
            entry["stories"] += 1
            entry["points"] += r["story_points"]
            if r["status"] == StoryWorkflowStatus.DONE.value:
                entry["done"] += 1

        cards = []
        for member in self._roster(batch):
            person = _person(member.student)
            stats = held.get(person["id"], {"stories": 0, "points": 0, "done": 0})
            cards.append({
                **person,
                "responsibility": member.responsibility,
                "stories": stats["stories"],
                "points": stats["points"],
                "done": stats["done"],
                "percent": int(round(_pct(stats["done"], stats["stories"]))),
            })
        cards.sort(key=lambda c: (-c["stories"], c["name"]))
        return cards

    # -------------------------------------------------------------- writing

    async def _story(self, batch: ProjectBatch, story_id: str) -> ProjectUserStory:
        story = (await self.db.execute(
            select(ProjectUserStory)
            .where(ProjectUserStory.id == story_id)
            .where(ProjectUserStory.batch_id == batch.id)
            .options(
                selectinload(ProjectUserStory.epic),
                selectinload(ProjectUserStory.sprint),
                selectinload(ProjectUserStory.assignee),
                selectinload(ProjectUserStory.created_by),
            )
        )).scalars().first()
        if story is None:
            raise StoryError("That story is not on this batch.")
        return story

    def _log(self, story: ProjectUserStory, user: Optional[User], kind: StoryEventKind,
             summary: str, *, before: Optional[str] = None, after: Optional[str] = None) -> None:
        self.db.add(StoryEvent(
            story_id=story.id,
            actor_id=user.id if user else None,
            actor_name=(user.full_name or user.email) if user else None,
            kind=kind,
            summary=summary,
            from_value=before,
            to_value=after,
        ))

    async def _next_key(self, batch_id) -> str:
        """
        Continue the batch's own numbering, drafted stories included.

        `key` is unique per batch, so a new story must not reuse US-104 just
        because the AI-drafted story holding that key was rejected.
        """
        keys = (await self.db.execute(
            select(ProjectUserStory.key).where(ProjectUserStory.batch_id == batch_id)
        )).scalars().all()
        highest = 100
        for key in keys:
            tail = (key or "").rsplit("-", 1)[-1]
            if tail.isdigit():
                highest = max(highest, int(tail))
        return f"US-{highest + 1}"

    def _member_or_none(self, batch: ProjectBatch, assignee_id: Optional[str]) -> Optional[User]:
        if not assignee_id:
            return None
        for member in self._roster(batch):
            if str(member.student_id) == str(assignee_id):
                return member.student
        raise StoryError("That student is not an active member of this batch.")

    async def _sprint_or_none(self, batch: ProjectBatch,
                              sprint_id: Optional[str]) -> Optional[ProjectSprint]:
        if not sprint_id:
            return None
        sprint = (await self.db.execute(
            select(ProjectSprint)
            .where(ProjectSprint.id == sprint_id)
            .where(ProjectSprint.batch_id == batch.id)
        )).scalars().first()
        if sprint is None:
            raise StoryError("That sprint belongs to a different batch.")
        return sprint

    async def create(self, batch: ProjectBatch, payload: dict, user: User) -> dict:
        title = (payload.get("title") or "").strip()
        if not title:
            raise StoryError("A story needs a title.")

        epic = None
        if payload.get("epic_key"):
            epic = next((e for e in await self._epics(batch.id)
                         if e.key == payload["epic_key"]), None)
            if epic is None:
                raise StoryError("That epic is not on this batch.")

        assignee = self._member_or_none(batch, payload.get("assignee_id"))
        sprint = await self._sprint_or_none(batch, payload.get("sprint_id"))
        now = datetime.utcnow()

        story = ProjectUserStory(
            batch_id=batch.id,
            epic_id=epic.id if epic else None,
            key=await self._next_key(batch.id),
            title=title[:240],
            narrative=(payload.get("narrative") or "").strip() or None,
            dependencies=(payload.get("dependencies") or "").strip() or None,
            story_points=int(payload.get("story_points") or 0),
            priority=StoryPriority(payload.get("priority") or "medium"),
            story_type=StoryType(payload.get("story_type") or "story"),
            workflow_status=StoryWorkflowStatus(payload.get("status") or "to_do"),
            due_date=payload.get("due_date"),
            assignee_id=assignee.id if assignee else None,
            sprint_id=sprint.id if sprint else None,
            created_by_id=user.id,
            # A trainer writing a story is the approval. The review gate exists
            # to stop the model's drafts reaching the backlog unread, and there
            # is no draft here to read.
            review_status=StoryReviewStatus.APPROVED,
            reviewed_by_id=user.id,
            reviewed_at=now,
            moved_to_backlog_at=now,
        )
        self.db.add(story)
        await self.db.flush()

        for index, text in enumerate(payload.get("acceptance_criteria") or []):
            text = (text or "").strip()
            if text:
                self.db.add(StoryCriterion(
                    story_id=story.id, kind=CriterionKind.ACCEPTANCE,
                    text=text, met=True, position=index,
                ))

        self._log(story, user, StoryEventKind.CREATED, f"{story.key} created")
        if assignee:
            self._log(story, user, StoryEventKind.ASSIGNED,
                      f"Assigned to {assignee.full_name}", after=assignee.full_name)
        await self.db.commit()
        return {"id": str(story.id), "key": story.key}

    async def update(self, batch: ProjectBatch, story_id: str,
                     payload: dict, user: User) -> dict:
        story = await self._story(batch, story_id)
        changes: List[str] = []

        if payload.get("title"):
            title = payload["title"].strip()[:240]
            if title != story.title:
                self._log(story, user, StoryEventKind.EDITED, "Title edited",
                          before=story.title[:120], after=title[:120])
                story.title = title
                changes.append("title")

        if "narrative" in payload:
            narrative = (payload["narrative"] or "").strip() or None
            if narrative != story.narrative:
                story.narrative = narrative
                self._log(story, user, StoryEventKind.EDITED, "Description edited")
                changes.append("narrative")

        if "dependencies" in payload:
            story.dependencies = (payload["dependencies"] or "").strip() or None
            changes.append("dependencies")

        if payload.get("story_points") is not None:
            new_points = int(payload["story_points"])
            if not 0 <= new_points <= 100:
                raise StoryError("Story points must be between 0 and 100.")
            if new_points != (story.story_points or 0):
                self._log(story, user, StoryEventKind.POINTS_CHANGED, "Story points changed",
                          before=str(story.story_points or 0), after=str(new_points))
                story.story_points = new_points
                changes.append("story_points")

        if payload.get("priority"):
            new_priority = StoryPriority(payload["priority"])
            if new_priority != story.priority:
                self._log(story, user, StoryEventKind.PRIORITY_CHANGED, "Priority changed",
                          before=story.priority.value.title(),
                          after=new_priority.value.title())
                story.priority = new_priority
                changes.append("priority")

        if payload.get("story_type"):
            story.story_type = StoryType(payload["story_type"])
            changes.append("story_type")

        if payload.get("status"):
            new_status = StoryWorkflowStatus(payload["status"])
            if new_status != story.workflow_status:
                self._log(story, user, StoryEventKind.STATUS_CHANGED, "Status changed",
                          before=WORKFLOW_LABELS[story.workflow_status],
                          after=WORKFLOW_LABELS[new_status])
                # Both timestamps follow from the transition rather than from a
                # separate field a trainer would have to remember to set.
                if new_status != StoryWorkflowStatus.TO_DO and story.started_at is None:
                    story.started_at = datetime.utcnow()
                story.completed_at = (datetime.utcnow()
                                      if new_status == StoryWorkflowStatus.DONE else None)
                story.workflow_status = new_status
                changes.append("status")

        if "assignee_id" in payload:
            assignee = self._member_or_none(batch, payload["assignee_id"])
            new_id = assignee.id if assignee else None
            if str(new_id) != str(story.assignee_id):
                before = story.assignee.full_name if story.assignee else None
                self._log(story, user, StoryEventKind.ASSIGNED,
                          f"Assigned to {assignee.full_name}" if assignee else "Unassigned",
                          before=before, after=assignee.full_name if assignee else None)
                story.assignee_id = new_id
                changes.append("assignee")

        if "sprint_id" in payload:
            sprint = await self._sprint_or_none(batch, payload["sprint_id"])
            new_id = sprint.id if sprint else None
            if str(new_id) != str(story.sprint_id):
                before = story.sprint.name if story.sprint else None
                self._log(story, user, StoryEventKind.SPRINT_CHANGED,
                          f"Moved to {sprint.name}" if sprint else "Removed from sprint",
                          before=before, after=sprint.name if sprint else None)
                story.sprint_id = new_id
                changes.append("sprint")

        if "due_date" in payload:
            new_due = payload["due_date"]
            if new_due != story.due_date:
                self._log(story, user, StoryEventKind.EDITED, "Due date changed",
                          before=story.due_date.isoformat() if story.due_date else None,
                          after=new_due.isoformat() if new_due else None)
                story.due_date = new_due
                changes.append("due_date")

        if "trainer_comment" in payload:
            story.trainer_comment = (payload["trainer_comment"] or "").strip() or None
            changes.append("trainer_comment")

        story.updated_at = datetime.utcnow()
        await self.db.commit()
        return {"id": str(story.id), "key": story.key, "changed": changes}

    async def reorder(self, batch: ProjectBatch, story_ids: List[str]) -> dict:
        """
        Set the manual order of a column from the ids it now contains.

        The client sends the column top to bottom after the drop, so the whole
        column is rewritten rather than the moved card patched: two people
        dragging at once then converge on one order instead of trading
        increments that never settle.

        Ids that are not on this batch are ignored rather than rejected - a
        stale board should not be able to fail the whole write.
        """
        if not story_ids:
            raise StoryError("No stories to reorder.")

        rows = (await self.db.execute(
            select(ProjectUserStory)
            .where(ProjectUserStory.batch_id == batch.id)
            .where(ProjectUserStory.id.in_(story_ids))
        )).scalars().all()
        by_id = {str(r.id): r for r in rows}

        moved = 0
        for index, story_id in enumerate(story_ids):
            story = by_id.get(str(story_id))
            if story is None:
                continue
            if story.position != index:
                story.position = index
                moved += 1

        await self.db.commit()
        return {"reordered": len(by_id), "changed": moved}

    async def comment(self, batch: ProjectBatch, story_id: str, body: str, user: User) -> dict:
        text = (body or "").strip()
        if not text:
            raise StoryError("A comment needs something in it.")
        story = await self._story(batch, story_id)
        author = user.full_name or user.email
        self.db.add(StoryComment(
            story_id=story.id, author_id=user.id, author_name=author, body=text,
        ))
        self._log(story, user, StoryEventKind.COMMENTED, "Comment added")
        await self.db.commit()
        return {"story": story.key, "author": author}

    # ------------------------------------------------------- acceptance criteria

    async def _criterion(self, story_id, criterion_id) -> StoryCriterion:
        row = (await self.db.execute(
            select(StoryCriterion)
            .where(StoryCriterion.id == criterion_id)
            .where(StoryCriterion.story_id == story_id)
        )).scalars().first()
        if row is None:
            raise StoryError("That criterion is not on this story.")
        return row

    async def add_criterion(self, batch: ProjectBatch, story_id: str, text: str,
                            kind: str, user: User) -> dict:
        """Append a criterion. New ones start unmet - nothing is true by default."""
        story = await self._story(batch, story_id)
        text = (text or "").strip()
        if not text:
            raise StoryError("A criterion needs some text.")

        wanted = CriterionKind(kind)
        siblings = (await self.db.execute(
            select(StoryCriterion)
            .where(StoryCriterion.story_id == story.id)
            .where(StoryCriterion.kind == wanted)
        )).scalars().all()

        criterion = StoryCriterion(
            story_id=story.id, kind=wanted, text=text, met=False,
            position=len(siblings),
        )
        self.db.add(criterion)
        self._log(story, user, StoryEventKind.EDITED,
                  "Acceptance criterion added" if wanted == CriterionKind.ACCEPTANCE
                  else "Definition of done item added", after=text[:120])
        await self.db.commit()
        await self.db.refresh(criterion)
        return {"id": str(criterion.id), "text": criterion.text, "met": criterion.met}

    async def update_criterion(self, batch: ProjectBatch, story_id: str,
                               criterion_id: str, payload: dict, user: User) -> dict:
        story = await self._story(batch, story_id)
        criterion = await self._criterion(story.id, criterion_id)

        if payload.get("text") is not None:
            text = payload["text"].strip()
            if not text:
                raise StoryError("A criterion needs some text.")
            if text != criterion.text:
                self._log(story, user, StoryEventKind.EDITED, "Criterion edited",
                          before=criterion.text[:120], after=text[:120])
                criterion.text = text

        if payload.get("met") is not None and payload["met"] != criterion.met:
            # Worth recording: the met count is what the approval checklist
            # reads, so a tick here moves a figure on another screen.
            criterion.met = bool(payload["met"])
            self._log(story, user, StoryEventKind.EDITED,
                      "Criterion met" if criterion.met else "Criterion marked unmet",
                      after=criterion.text[:120])

        await self.db.commit()
        return {"id": str(criterion.id), "text": criterion.text, "met": criterion.met}

    async def delete_criterion(self, batch: ProjectBatch, story_id: str,
                               criterion_id: str, user: User) -> dict:
        story = await self._story(batch, story_id)
        criterion = await self._criterion(story.id, criterion_id)
        text = criterion.text
        await self.db.delete(criterion)
        self._log(story, user, StoryEventKind.EDITED, "Criterion removed",
                  before=text[:120])
        await self.db.commit()
        return {"removed": text}


    async def delete_story(self, batch: ProjectBatch, story_id: str,
                           user: User) -> dict:
        """
        Remove a story from the backlog.

        What hangs off it is decided by the schema, not here: its criteria,
        comments and history go with it, while sub-tasks and any commits that
        named it are detached rather than destroyed. A commit is evidence that
        somebody did work - deleting the ticket must not delete the record of
        the work.
        """
        story = await self._story(batch, story_id)
        key, title = story.key, story.title

        commits = (await self.db.execute(
            select(func.count(StoryCommit.id))
            .where(StoryCommit.story_id == story.id)
        )).scalar() or 0
        tasks = (await self.db.execute(
            select(func.count(ProjectTask.id))
            .where(ProjectTask.story_id == story.id)
        )).scalar() or 0

        await self.db.delete(story)
        await self.db.commit()
        logger.info(f"[Stories] {user.email} deleted {key} from {batch.batch_code}")
        return {
            "deleted": key,
            "title": title,
            # Said plainly, because the trainer is about to wonder where they went.
            "detached_commits": commits,
            "detached_tasks": tasks,
            "message": f"{key} deleted."
                       + (f" {commits} commit(s) and {tasks} sub-task(s) were kept,"
                          " no longer attached to any story." if commits or tasks else ""),
        }

    # ------------------------------------------------------------ files

    async def attach(self, batch: ProjectBatch, story_id: str, upload, user: User) -> dict:
        """
        Hang a file on a story.

        The bytes go to the same content-addressed store the batch documents
        use, which is what enforces the size and type rules - this method only
        records that this story points at them.
        """
        story = await self._story(batch, story_id)
        try:
            # The batch's college, not the uploader's: a student's profile may
            # carry none, and the file belongs to the institution that owns the
            # project either way.
            stored = await file_store.save(self.db, upload, user,
                                           college_id=batch.college_id)
        except file_store.FileStoreError as exc:
            raise StoryError(str(exc))

        name = file_store.safe_name(upload.filename or file_store.upload_name(stored))
        attachment = StoryAttachment(
            story_id=story.id,
            file_id=stored.id,
            name=name[:255],
            uploaded_by_id=user.id,
        )
        self.db.add(attachment)
        self._log(story, user, StoryEventKind.ATTACHED, f"Attached {name}", after=name)
        await self.db.commit()
        await self.db.refresh(attachment)
        return {
            "id": str(attachment.id),
            "name": attachment.name,
            "size_label": file_store.human_size(stored.byte_size),
        }

    async def _attachment(self, story_id, attachment_id) -> StoryAttachment:
        attachment = (await self.db.execute(
            select(StoryAttachment)
            .where(StoryAttachment.id == attachment_id)
            .where(StoryAttachment.story_id == story_id)
            .options(selectinload(StoryAttachment.file))
        )).scalars().first()
        if attachment is None:
            raise StoryError("That attachment is not on this story.")
        return attachment

    async def read_attachment(self, batch: ProjectBatch, story_id: str,
                              attachment_id: str):
        """The bytes and the headers to send them with."""
        story = await self._story(batch, story_id)
        attachment = await self._attachment(story.id, attachment_id)
        if attachment.file is None:
            raise StoryError("The file behind that attachment is missing.")
        try:
            content = await file_store.read(attachment.file)
        except file_store.FileStoreError as exc:
            raise StoryError(str(exc))
        return content, file_store.download_headers(attachment.file, attachment.name)

    async def detach(self, batch: ProjectBatch, story_id: str,
                     attachment_id: str, user: User) -> dict:
        """
        Remove the attachment, not the blob.

        Another story - or another team - may point at the same content, and
        the file store is deliberately append-only about bytes.
        """
        story = await self._story(batch, story_id)
        attachment = await self._attachment(story.id, attachment_id)
        name = attachment.name
        await self.db.delete(attachment)
        self._log(story, user, StoryEventKind.DETACHED, f"Removed {name}", before=name)
        await self.db.commit()
        return {"removed": name}

    # ---------------------------------------------------- import and export

    @staticmethod
    def export_columns() -> List[str]:
        return ["key", "title", "type", "epic", "assignee", "roll_number", "sprint",
                "priority", "story_points", "status", "labels", "created_by", "created_at"]

    @staticmethod
    def export_row(row: dict) -> List:
        who = row["assignee"]
        return [
            row["key"], row["title"], row["type_label"], row["epic_key"] or "",
            who["name"] if who else "", (who["roll"] if who else "") or "",
            row["sprint"]["name"] if row["sprint"] else "",
            row["priority_label"], row["story_points"], row["status_label"],
            row.get("labels") or "", row["created_by"] or "",
            row["created_at"].strftime("%Y-%m-%d %H:%M") if row["created_at"] else "",
        ]

    @staticmethod
    def import_template() -> bytes:
        """
        The trainer template as a workbook: the import sheet plus its guide.

        Handing back the same shape that is accepted means a trainer who
        downloads it, fills it in and uploads it cannot be told their columns
        are wrong.
        """
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "User Stories Import"

        head = Font(bold=True, color="FFFFFF")
        fill = PatternFill("solid", fgColor="1B2A6B")
        for column, heading in enumerate(IMPORT_COLUMNS, start=1):
            cell = sheet.cell(row=1, column=column)
            # Mandatory fields carry the star the guide explains, exactly as
            # the issued template marks them.
            mandatory = heading in {"Work Type", "Summary", "Description",
                                    "Acceptance Criteria", "Priority", "Story Points",
                                    "Assignee Name", "Status"}
            cell.value = f"{heading} *" if mandatory else heading
            cell.font = head
            cell.fill = fill
        sheet.append([
            "US-101", "Story", "Authentication", "Student Login and Authentication",
            "As a student, I want to log in using my credentials so that I can "
            "access my project workspace.",
            "1. Valid credentials allow login\n2. Invalid credentials show an error\n"
            "3. Successful login redirects to dashboard",
            "High", 5, "23K91A05L5", "Royyala Sindhuja", "Sprint 1", "To Do",
            "authentication,frontend",
        ])
        for column, width in zip("ABCDEFGHIJKLM",
                                 (10, 12, 18, 34, 52, 46, 10, 12, 18, 22, 12, 12, 24)):
            sheet.column_dimensions[column].width = width
        sheet.row_dimensions[2].height = 60
        for cell in sheet[2]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        guide = wb.create_sheet("Field Guide")
        guide.append(["BharatBuild - User Stories Excel Import Guide"])
        guide.append(["Fields marked with * are mandatory. An assignee is matched on "
                      "roll number first, then on name, and must be an active member "
                      "of the selected batch."])
        guide.append([])
        guide.append(["Field", "Mandatory", "Example", "Validation / Usage"])
        for cell in guide[4]:
            cell.font = Font(bold=True)
        for line in (
            ("Story ID", "No", "US-101",
             "Kept when it is free on this batch; a new one is generated otherwise."),
            ("Work Type", "Yes", "Story", "Story, Task, Spike or Bug."),
            ("Epic", "No", "Authentication",
             "Matched on key or title; created on this batch if it does not exist."),
            ("Summary", "Yes", "Student Login and Authentication", "The story title."),
            ("Description", "Yes", "As a student, I want to...",
             "The narrative shown on the story."),
            ("Acceptance Criteria", "Yes", "1. Valid login...",
             "Numbered lines in one cell; pipes and semicolons also split."),
            ("Priority", "Yes", "High",
             "Highest, High, Medium, Low, Lowest. Highest and Lowest are "
             "recorded as High and Low."),
            ("Story Points", "Yes", "5", "A whole number from 0 to 100."),
            ("Assignee Roll No", "No", "23K91A05L5",
             "Matched against this batch. Anyone else imports unassigned, with a note."),
            ("Assignee Name", "No", "Royyala Sindhuja",
             "Used when no roll number is given. Blank leaves the story unassigned."),
            ("Sprint", "No", "Sprint 1",
             "Matched on name; created on this batch if it does not exist."),
            ("Status", "Yes", "To Do", "To Do, In Progress, In Review or Done."),
            ("Labels", "No", "authentication,frontend", "Comma-separated tags."),
        ):
            guide.append(list(line))
        for column, width in zip("ABCD", (22, 12, 34, 66)):
            guide.column_dimensions[column].width = width

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    # ----------------------------------------------------------- reading in

    @staticmethod
    def _sheet_rows(raw: bytes, filename: str) -> Tuple[List[str], List[List]]:
        """
        The header row and the data rows, from .xlsx or .csv alike.

        The workbook is searched for the sheet that actually holds the stories
        rather than assuming the first one: the issued template ships a Field
        Guide and an Import Process sheet beside it.
        """
        name = (filename or "").lower()
        if name.endswith((".xlsx", ".xlsm")):
            import openpyxl
            try:
                wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
            except Exception:
                raise StoryError("That file could not be opened as an Excel workbook.")

            wanted = {alias for _, aliases in IMPORT_FIELDS.values() for alias in aliases}
            best: Optional[Tuple[List[str], List[List]]] = None
            for ws in wb.worksheets:
                rows = [list(r) for r in ws.iter_rows(values_only=True)]
                for index, row in enumerate(rows[:5]):
                    headings = [str(c).strip().lower().rstrip("*").strip()
                                for c in row if c is not None]
                    if "summary" in headings or "title" in headings or (
                            len(set(headings) & wanted) >= 3):
                        found = ([str(c or "") for c in row], rows[index + 1:])
                        if best is None or len(found[1]) > len(best[1]):
                            best = found
                        break
            if best is None:
                raise StoryError(
                    "No sheet in that workbook has the template's columns. "
                    f"Expected a header row with: {', '.join(IMPORT_COLUMNS)}.")
            return best

        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            raise StoryError("That file is not UTF-8 text. Save it as .xlsx or CSV.")
        reader = list(csv.reader(io.StringIO(text)))
        if not reader:
            raise StoryError("That file is empty.")
        return reader[0], reader[1:]

    @classmethod
    def _map_columns(cls, headings: List[str]) -> dict:
        """Which column holds which field, by heading. Order is not assumed."""
        cleaned = [str(h or "").strip().lower().rstrip("*").strip() for h in headings]
        mapping = {}
        for field, (_, aliases) in IMPORT_FIELDS.items():
            for index, heading in enumerate(cleaned):
                if heading in aliases:
                    mapping[field] = index
                    break
        if "summary" not in mapping:
            raise StoryError(
                "The sheet needs a Summary column (the story title). "
                f"Expected columns: {', '.join(IMPORT_COLUMNS)}.")
        return mapping

    async def import_rows(self, batch: ProjectBatch, raw: bytes, filename: str,
                          user: User, *, dry_run: bool = False) -> dict:
        """
        Read the trainer's sheet, and either report what it would do or do it.

        `dry_run` is the Validate step of the template's own workflow: every
        row is parsed and checked, nothing is written, and the caller gets back
        the same list it will get after confirming. A row with a problem is
        reported against its row number and skipped; the rest still land.
        """
        headings, data = self._sheet_rows(raw, filename)
        mapping = self._map_columns(headings)

        epics = {}
        for epic in await self._epics(batch.id):
            epics[epic.key.lower()] = epic
            epics[epic.title.lower()] = epic
        sprints = {s.name.lower(): s for s in await self.sprints(batch.id)}
        roster = self._roster(batch)
        taken = set((await self.db.execute(
            select(ProjectUserStory.key).where(ProjectUserStory.batch_id == batch.id)
        )).scalars().all())

        def cell(row: List, field: str) -> str:
            index = mapping.get(field)
            if index is None or index >= len(row):
                return ""
            value = row[index]
            return "" if value is None else str(value).strip()

        def find_member(roll: str, name: str) -> Tuple[Optional[User], Optional[str]]:
            """
            Roll number first - it is the identifier the college actually uses.

            A name that resolves to nobody is a note, not a refusal: the story
            itself is fine, an assignee is optional, and a sheet written before
            the roster was finalised should still import. The row lands
            unassigned and says so, which the trainer can fix on the board in
            one click.
            """
            if roll:
                for member in roster:
                    if (member.student.roll_number or "").lower() == roll.lower():
                        return member.student, None
                return None, (f"No student with roll number {roll} is on this batch "
                              "- imported unassigned")
            if name:
                matches = [m.student for m in roster
                           if (m.student.full_name or "").lower() == name.lower()]
                if len(matches) == 1:
                    return matches[0], None
                if not matches:
                    return None, f"{name} is not on this batch - imported unassigned"
                return None, (f"{name} matches more than one member - imported "
                              "unassigned; use a roll number to be specific")
            return None, None

        preview: List[dict] = []
        issues: List[dict] = []
        # Anything worth telling the trainer that does not stop the row.
        notes: List[dict] = []
        planned_keys = set()
        now = datetime.utcnow()

        for line, row in enumerate(data, start=2):
            if all((c is None or str(c).strip() == "") for c in row):
                continue

            summary = cell(row, "summary")
            problems: List[str] = []
            warnings: List[str] = []
            if not summary:
                issues.append({"row": line, "issue": "No Summary"})
                continue

            student, who_note = find_member(cell(row, "roll"), cell(row, "assignee"))
            if who_note:
                warnings.append(who_note)

            priority = PRIORITY_WORDS.get(cell(row, "priority").lower() or "medium")
            if priority is None:
                problems.append(f"Priority '{cell(row, 'priority')}' is not one of "
                                "Highest, High, Medium, Low, Lowest")
            work_type = TYPE_WORDS.get(cell(row, "work_type").lower() or "story")
            if work_type is None:
                problems.append(f"Work Type '{cell(row, 'work_type')}' is not one of "
                                "Story, Task, Spike, Bug")
            status = STATUS_WORDS.get(cell(row, "status").lower() or "to do")
            if status is None:
                problems.append(f"Status '{cell(row, 'status')}' is not one of "
                                "To Do, In Progress, In Review, Done")

            raw_points = cell(row, "points") or "0"
            points = None
            try:
                points = int(float(raw_points))
            except ValueError:
                problems.append(f"Story Points '{raw_points}' is not a number")
            if points is not None and not 0 <= points <= 100:
                problems.append("Story Points must be between 0 and 100")

            # The sheet's own id is kept when it is free, so a trainer's
            # numbering survives the import.
            wanted_key = cell(row, "story_id")
            key = wanted_key if wanted_key and wanted_key not in taken | planned_keys else None
            note = None
            if wanted_key and key is None:
                note = f"{wanted_key} is already used on this batch; a new id will be given"

            epic_name = cell(row, "epic")
            epic = epics.get(epic_name.lower()) if epic_name else None
            new_epic = bool(epic_name) and epic is None

            sprint_name = cell(row, "sprint")
            sprint = sprints.get(sprint_name.lower()) if sprint_name else None
            new_sprint = bool(sprint_name) and sprint is None

            labels = ",".join(
                part.strip() for part in cell(row, "labels").split(",") if part.strip()
            ) or None

            entry = {
                "row": line,
                # Only an id that is actually free is shown; a clashing one
                # would read as "this row keeps US-101" when it will not.
                "key": key or "(new)",
                "summary": summary[:240],
                "work_type": work_type.value if work_type else None,
                "epic": epic_name or None,
                "new_epic": new_epic,
                "priority": priority.value if priority else None,
                "story_points": points,
                "assignee": _person(student),
                "sprint": sprint_name or None,
                "new_sprint": new_sprint,
                "status": status.value if status else None,
                "labels": labels,
                "criteria": len(_criteria_lines(cell(row, "acceptance"))),
                "note": note,
                "issues": problems,
                "warnings": warnings,
            }
            preview.append(entry)
            for problem in problems:
                issues.append({"row": line, "issue": problem})
            for warning in warnings:
                notes.append({"row": line, "issue": warning})

            if problems or dry_run:
                if key:
                    planned_keys.add(key)
                continue

            # ---- write ------------------------------------------------------
            if new_epic:
                epic = ProjectEpic(
                    batch_id=batch.id,
                    key=f"EP-{len([e for e in set(epics.values())]) + 1:02d}",
                    title=epic_name[:200],
                    position=len(set(epics.values())),
                )
                self.db.add(epic)
                await self.db.flush()
                epics[epic.key.lower()] = epic
                epics[epic.title.lower()] = epic

            if new_sprint:
                sprint = ProjectSprint(
                    batch_id=batch.id,
                    key=f"SP-{len(sprints) + 1:02d}",
                    name=sprint_name[:80],
                    position=len(sprints),
                )
                self.db.add(sprint)
                await self.db.flush()
                sprints[sprint.name.lower()] = sprint

            story_key = key or await self._next_key(batch.id)
            story = ProjectUserStory(
                batch_id=batch.id,
                epic_id=epic.id if epic else None,
                key=story_key,
                title=summary[:240],
                narrative=cell(row, "description") or None,
                story_points=points or 0,
                priority=priority,
                story_type=work_type,
                workflow_status=status,
                assignee_id=student.id if student else None,
                sprint_id=sprint.id if sprint else None,
                labels=labels,
                created_by_id=user.id,
                review_status=StoryReviewStatus.APPROVED,
                reviewed_by_id=user.id,
                reviewed_at=now,
                moved_to_backlog_at=now,
            )
            self.db.add(story)
            await self.db.flush()
            taken.add(story_key)
            entry["key"] = story_key

            for index, text_item in enumerate(_criteria_lines(cell(row, "acceptance"))):
                self.db.add(StoryCriterion(
                    story_id=story.id, kind=CriterionKind.ACCEPTANCE,
                    text=text_item, met=True, position=index,
                ))

            self._log(story, user, StoryEventKind.IMPORTED,
                      f"{story.key} imported from {filename or 'a sheet'}")

        ready = [p for p in preview if not p["issues"]]
        if dry_run:
            # Nothing was written, but epics and sprints were looked up on this
            # session; rolling back keeps that explicit.
            await self.db.rollback()
        else:
            await self.db.commit()

        return {
            "dry_run": dry_run,
            "columns": [IMPORT_FIELDS[f][0] for f in mapping],
            "rows": len(preview),
            "ready": len(ready),
            "preview": preview,
            "created": [] if dry_run else [p["key"] for p in ready],
            "count": 0 if dry_run else len(ready),
            "issues": issues,
            "notes": notes,
        }
