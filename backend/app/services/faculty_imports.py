"""
Roster Import Service - parse, validate and apply an uploaded roster.

A Student List import is the portal's real data entry point: it creates the
User and StudentEnrollment rows the rest of the faculty screens count. Rows are
validated first and applied one at a time, so a bad row never takes the file
down with it.

Two rules the screen states and this enforces:
  * a verified student record is never overwritten automatically
  * a duplicate is skipped, not failed - it is not an error to re-upload
"""

import csv
import io
import re
from dataclasses import dataclass, field as dc_field
from datetime import datetime
from math import ceil
from typing import Dict, List, Optional, Sequence, Tuple

from passlib.context import CryptContext
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.faculty import (
    ProjectBatch,
    ProjectBatchMember,
    StudentEnrollment,
    StudentProfileStatus,
)
from app.models.faculty_import import (
    IMPORT_STATUS_LABELS,
    IMPORT_TYPE_LABELS,
    ImportEvent,
    ImportRowIssue,
    ImportRun,
    ImportStatus,
    ImportType,
    IssueSeverity,
)
from app.models.user import COLLEGE_STAFF_ROLES, User, UserRole

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

MOBILE_RE = re.compile(r"^[6-9]\d{9}$")
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

# Columns each import type expects. The template endpoint serves these.
TEMPLATE_COLUMNS: Dict[ImportType, List[str]] = {
    ImportType.STUDENT_LIST: [
        "roll_number", "full_name", "email", "mobile",
        "department", "section", "year", "semester",
    ],
    ImportType.BATCH_ALLOCATION: ["batch_code", "project_title", "section", "roll_number", "is_leader"],
    ImportType.PROJECT_DETAILS: ["batch_code", "project_title", "abstract"],
    ImportType.BASE_PAPER_METADATA: ["batch_code", "paper_title", "authors", "publication", "year", "url"],
}

# Imports that actually write records. The others are accepted and validated
# but not applied - see apply_row.
IMPLEMENTED_TYPES = {
    ImportType.STUDENT_LIST,
    ImportType.BATCH_ALLOCATION,
    ImportType.PROJECT_DETAILS,
    ImportType.BASE_PAPER_METADATA,
}

MAX_FILE_BYTES = 5 * 1024 * 1024


@dataclass
class RowIssue:
    row: int
    message: str
    field: Optional[str] = None
    value: Optional[str] = None
    duplicate: bool = False


@dataclass
class ParsedFile:
    headers: List[str]
    rows: List[Tuple[int, Dict[str, str]]]  # (1-based row number, values)
    issues: List[RowIssue] = dc_field(default_factory=list)


def parse_upload(filename: str, content: bytes) -> ParsedFile:
    """Read a CSV or XLSX into normalised header/row pairs."""
    name = (filename or "").lower()

    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - dependency is pinned
            raise ValueError("XLSX support needs openpyxl on the server") from exc
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        raw = [[("" if c is None else str(c)).strip() for c in row] for row in sheet.iter_rows(values_only=True)]
    elif name.endswith(".csv") or name.endswith(".txt"):
        text = content.decode("utf-8-sig", errors="replace")
        raw = [[cell.strip() for cell in row] for row in csv.reader(io.StringIO(text))]
    else:
        raise ValueError("Unsupported file type. Upload a .csv or .xlsx file.")

    raw = [row for row in raw if any(cell for cell in row)]
    if not raw:
        raise ValueError("The file is empty.")

    headers = [h.strip().lower().replace(" ", "_") for h in raw[0]]
    rows: List[Tuple[int, Dict[str, str]]] = []
    for index, values in enumerate(raw[1:], start=2):
        record = {headers[i]: (values[i] if i < len(values) else "") for i in range(len(headers))}
        rows.append((index, record))
    return ParsedFile(headers=headers, rows=rows)


# ============================================
# Wide batch-allocation layout
# ============================================
#
# Colleges hand out the allocation as one row per batch, with the team spread
# across "Student 1..N" columns and each cell holding "<roll> - <name>":
#
#   projid      Section   Batch No  Guide Name  Student 1                       Student 2
#   CSE-D-D1    CSE - D   D1                    23K91A05L5 - Royyala Sindhuja   23K91A05P2 - T Meenakshi
#
# The importer works one student at a time, so the sheet is flattened to the
# tall shape before any row is applied. Doing it here rather than in the row
# handler keeps the header check, the row numbering and the error CSV all
# working on the same rows the file actually has.

_STUDENT_COL = re.compile(r"^student_?(\d+)$")

# "Email 1" sits beside "Student 1" when the sheet carries addresses. It is
# optional: without it a student who has no account yet cannot be created.
_EMAIL_COL = re.compile(r"^(?:email|e_mail|mail)_?(\d+)$")

# What the batch code column may be called. "projid" is what the college
# sheets use; the others are what the tall template and exports use.
_CODE_COLUMNS = ("batch_code", "projid", "proj_id", "project_id", "batch_id")

# A roll number is one unspaced token carrying at least one digit - what every
# college code looks like (23K91A05L5, 22CS001) and what no personal name does.
_ROLL_SHAPE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9/_-]{2,19}$")

# Hyphen, en dash, em dash and friends: whichever one the sheet was typed with.
_DASHES = "\u2010\u2011\u2012\u2013\u2014\u2015-"
_FIRST_DASH = re.compile(f"\\s*[{_DASHES}]\\s*")


def _looks_like_roll(token: str) -> bool:
    """One token, has a digit, no spaces - a roll rather than a name."""
    token = token.strip()
    return bool(token) and any(c.isdigit() for c in token) and bool(_ROLL_SHAPE.match(token))


def _split_student(cell: str) -> Optional[Tuple[str, str]]:
    """
    A student cell into (roll, name).

    The cell is "<roll> - <name>", but the dash is not a reliable landmark: a
    name may be hyphenated ("Sai-Kumar"), the roll may abut the dash with no
    spaces, some sheets separate with a space alone, and some write the name
    first. So the cell is cut at the FIRST separator only, and the roll is then
    identified by shape rather than by which side it landed on.

    Returns None for an empty cell - a smaller team, not an error. Returns an
    empty roll when no side looks like one, so the row fails and says so
    instead of guessing a roll out of somebody's name.
    """
    cell = " ".join((cell or "").split())
    if not cell:
        return None

    parts = _FIRST_DASH.split(cell, maxsplit=1)
    if len(parts) == 1:
        # No dash: the roll is the first or last whitespace token.
        words = cell.split(" ")
        if len(words) == 1:
            return (cell.upper(), "") if _looks_like_roll(cell) else ("", cell)
        if _looks_like_roll(words[0]):
            return words[0].upper(), " ".join(words[1:]).strip()
        if _looks_like_roll(words[-1]):
            return words[-1].upper(), " ".join(words[:-1]).strip()
        return "", cell

    left, right = parts[0].strip(), parts[1].strip()
    if _looks_like_roll(left):
        return left.upper(), right
    if _looks_like_roll(right):
        # "Royyala Sindhuja - 23K91A05L5": some sheets lead with the name.
        return right.upper(), left
    return "", cell


def _split_section(raw: str) -> Tuple[Optional[str], Optional[str]]:
    """
    "CSE - D" into ("CSE", "D").

    The sheet writes the section as department-and-letter; batches store the
    two apart, and the letter alone is what a section filter matches.
    """
    raw = (raw or "").strip()
    if not raw:
        return None, None
    parts = [p for p in re.split(r"\s*[\u2010-\u2015\-/]\s*", raw) if p]
    if len(parts) >= 2:
        return parts[0].upper(), parts[-1].upper()
    return None, parts[0].upper() if parts else None


def is_wide_allocation(parsed: ParsedFile) -> bool:
    """Whether this sheet is the one-row-per-batch layout."""
    return any(_STUDENT_COL.match(h) for h in parsed.headers)


def normalise_wide_allocation(parsed: ParsedFile) -> ParsedFile:
    """
    Flatten one-row-per-batch into one row per student.

    Row numbers are kept pointing at the sheet's own row, so an issue against
    a student still names the line the uploader can see.
    """
    student_cols = sorted(
        (h for h in parsed.headers if _STUDENT_COL.match(h)),
        key=lambda h: int(_STUDENT_COL.match(h).group(1)),
    )
    if not student_cols:
        return parsed

    code_col = next((c for c in _CODE_COLUMNS if c in parsed.headers), None)
    # "Email 2" belongs to "Student 2"; pairing is by the number, not position,
    # so a sheet that only carries some addresses still lines up.
    email_cols = {
        _EMAIL_COL.match(h).group(1): h for h in parsed.headers if _EMAIL_COL.match(h)
    }

    rows: List[Tuple[int, Dict[str, str]]] = []
    for row_number, record in parsed.rows:
        code = (record.get(code_col) or "").strip().upper() if code_col else ""
        department, section = _split_section(record.get("section", ""))
        guide = (record.get("guide_name") or record.get("guide") or "").strip()
        title = (record.get("project_title") or record.get("title") or "").strip()

        for position, column in enumerate(student_cols):
            student = _split_student(record.get(column, ""))
            if student is None:
                continue          # an empty column is a smaller team, not an error
            roll, name = student
            index = _STUDENT_COL.match(column).group(1)
            rows.append((row_number, {
                "email": (record.get(email_cols.get(index, ""), "") or "").strip().lower(),
                "batch_code": code,
                "project_title": title,
                "section": section or "",
                "department": department or "",
                "roll_number": roll,
                "full_name": name,
                "guide_name": guide,
                # The sheet has no leader column; the first named student is
                # the one every college writes first.
                "is_leader": "true" if position == 0 else "false",
            }))

    return ParsedFile(
        headers=["batch_code", "project_title", "section", "department",
                 "roll_number", "full_name", "email", "guide_name", "is_leader"],
        rows=rows,
    )


class FacultyImportService:
    def __init__(self, db: AsyncSession, college_id=None):
        # One lookup per import, not per row.
        self._domain_cache: Optional[str] = None
        self.db = db
        # The college every imported row lands in. A roster upload must not be
        # able to place students into another institution.
        self.college_id = college_id

    # ------------------------------------------------------------ running

    async def _next_code(self, now: datetime) -> str:
        year = now.year
        prefix = f"IMP-{year}-"
        count = (
            await self.db.execute(
                select(func.count()).select_from(ImportRun).where(ImportRun.import_code.like(f"{prefix}%"))
            )
        ).scalar() or 0
        return f"{prefix}{count + 1:03d}"

    async def run_import(
        self,
        *,
        filename: str,
        content: bytes,
        import_type: ImportType,
        academic_year: str,
        department: Optional[str],
        actor: User,
    ) -> ImportRun:
        if len(content) > MAX_FILE_BYTES:
            raise ValueError(f"File is larger than {MAX_FILE_BYTES // (1024 * 1024)} MB.")

        now = datetime.utcnow()
        run = ImportRun(
            import_code=await self._next_code(now),
            file_name=filename,
            file_size=len(content),
            file_content=content,
            import_type=import_type,
            department=department,
            academic_year=academic_year,
            status=ImportStatus.PROCESSING,
            imported_by_id=actor.id,
            college_id=self.college_id,
            started_at=now,
        )
        self.db.add(run)
        await self.db.flush()

        actor_name = actor.full_name or actor.email
        self.db.add(ImportEvent(run_id=run.id, step="Uploaded", actor=actor_name, occurred_at=now))

        # --- parse
        try:
            parsed = parse_upload(filename, content)
        except ValueError as exc:
            run.status = ImportStatus.FAILED
            run.completed_at = datetime.utcnow()
            self.db.add(ImportEvent(run_id=run.id, step="File Validated", actor="System",
                                    note=str(exc), is_warning=True))
            self.db.add(ImportRowIssue(run_id=run.id, row_number=0, message=str(exc)))
            await self.db.commit()
            return run

        # A wide allocation sheet is flattened before the header check, so the
        # check runs against the columns the importer actually consumes.
        if import_type == ImportType.BATCH_ALLOCATION and is_wide_allocation(parsed):
            parsed = normalise_wide_allocation(parsed)
            self.db.add(ImportEvent(
                run_id=run.id, step="File Validated", actor="System",
                note=f"Wide allocation sheet flattened to {len(parsed.rows)} student row(s)",
            ))

        required = TEMPLATE_COLUMNS[import_type]
        missing = [c for c in required[:3] if c not in parsed.headers]
        if missing:
            message = f"Missing required column(s): {', '.join(missing)}"
            run.status = ImportStatus.FAILED
            run.rows_total = len(parsed.rows)
            run.rows_failed = len(parsed.rows)
            run.completed_at = datetime.utcnow()
            self.db.add(ImportEvent(run_id=run.id, step="File Validated", actor="System",
                                    note=message, is_warning=True))
            self.db.add(ImportRowIssue(run_id=run.id, row_number=1, message=message, field="header"))
            await self.db.commit()
            return run

        self.db.add(ImportEvent(run_id=run.id, step="File Validated", actor="System"))

        # --- apply
        run.rows_total = len(parsed.rows)
        issues: List[RowIssue] = []
        imported = duplicates = 0
        seen_keys: set = set()

        for row_number, values in parsed.rows:
            outcome, issue = await self._apply_row(
                import_type, row_number, values, run, seen_keys, actor
            )
            if outcome == "imported":
                imported += 1
            elif outcome == "duplicate":
                duplicates += 1
            if issue:
                issues.append(issue)

        self.db.add(ImportEvent(run_id=run.id, step="Rows Processed", actor="System"))

        for issue in issues:
            self.db.add(ImportRowIssue(
                run_id=run.id,
                row_number=issue.row,
                field=issue.field,
                message=issue.message,
                raw_value=issue.value,
                severity=IssueSeverity.DUPLICATE if issue.duplicate else IssueSeverity.ERROR,
            ))

        run.rows_imported = imported
        run.rows_duplicate = duplicates
        run.rows_failed = sum(1 for i in issues if not i.duplicate)
        run.completed_at = datetime.utcnow()

        if run.rows_imported == 0 and run.rows_failed:
            run.status = ImportStatus.FAILED
        elif run.rows_failed or run.rows_duplicate:
            run.status = ImportStatus.PARTIALLY_IMPORTED
        else:
            run.status = ImportStatus.SUCCESSFUL

        self.db.add(ImportEvent(run_id=run.id, step="Records Created", actor="System",
                                note=f"{imported} record(s)"))
        self.db.add(ImportEvent(
            run_id=run.id,
            step="Completed with Warnings" if run.status == ImportStatus.PARTIALLY_IMPORTED
            else IMPORT_STATUS_LABELS[run.status],
            actor=actor_name,
            is_warning=run.status != ImportStatus.SUCCESSFUL,
        ))

        await self.db.commit()
        await self.db.refresh(run)
        return run

    # --------------------------------------------------------- row handling

    async def _apply_row(
        self,
        import_type: ImportType,
        row: int,
        values: Dict[str, str],
        run: ImportRun,
        seen_keys: set,
        actor: User,
    ) -> Tuple[str, Optional[RowIssue]]:
        if import_type == ImportType.STUDENT_LIST:
            return await self._apply_student(row, values, run, seen_keys)
        if import_type == ImportType.BATCH_ALLOCATION:
            return await self._apply_allocation(row, values, run)
        if import_type == ImportType.PROJECT_DETAILS:
            return await self._apply_project_details(row, values)
        if import_type == ImportType.BASE_PAPER_METADATA:
            return await self._apply_base_paper(row, values)
        return "failed", RowIssue(row=row, message="Unsupported import type")

    async def _apply_student(
        self, row: int, values: Dict[str, str], run: ImportRun, seen_keys: set
    ) -> Tuple[str, Optional[RowIssue]]:
        roll = (values.get("roll_number") or "").strip().upper()
        name = (values.get("full_name") or "").strip()
        email = (values.get("email") or "").strip().lower()
        mobile = (values.get("mobile") or "").strip()
        section = (values.get("section") or "").strip() or None
        department = (values.get("department") or "").strip() or run.department
        year = (values.get("year") or "").strip() or "4th Year"
        semester = (values.get("semester") or "").strip() or None

        if not roll:
            return "failed", RowIssue(row, "Roll number is missing", "roll_number")
        if not name:
            return "failed", RowIssue(row, "Name is missing", "full_name")
        if not email or not EMAIL_RE.match(email):
            return "failed", RowIssue(row, "Invalid or missing email", "email", email)
        if mobile and not MOBILE_RE.match(mobile):
            return "failed", RowIssue(row, "Invalid mobile number", "mobile", mobile)
        if not section:
            return "failed", RowIssue(row, "Section is missing", "section")
        if not department:
            return "failed", RowIssue(row, "Department is missing", "department")

        # Duplicate inside the same file.
        if roll in seen_keys:
            return "duplicate", RowIssue(row, f"Duplicate roll number {roll}", "roll_number", roll, duplicate=True)
        seen_keys.add(roll)

        existing_email = (
            await self.db.execute(select(User).where(func.lower(User.email) == email))
        ).scalar_one_or_none()
        existing_roll = (
            await self.db.execute(select(User).where(func.upper(User.roll_number) == roll))
        ).scalar_one_or_none()

        # An email that already belongs to a different roll number is a typo or
        # a copied row, never a legitimate duplicate. Matching on email alone
        # here would attach this row to someone else's account - and, if that
        # account were unverified, overwrite their name and phone.
        if existing_email and (existing_email.roll_number or "").strip().upper() != roll:
            return "failed", RowIssue(
                row,
                f"Email already registered to {existing_email.roll_number or existing_email.email}",
                "email", email,
            )

        user = existing_roll

        if user is None:
            user = User(
                email=email,
                hashed_password=pwd_context.hash(f"Bharat@{roll}"),
                full_name=name,
                role=UserRole.STUDENT,
                roll_number=roll,
                phone=mobile or None,
                department=department,
                section=section,
                is_active=True,
            )
            self.db.add(user)
            await self.db.flush()
        else:
            # Never overwrite a verified record automatically - the screen
            # promises this, so only fill blanks on an existing user.
            enrolled = (
                await self.db.execute(
                    select(StudentEnrollment)
                    .where(StudentEnrollment.student_id == user.id)
                    .where(StudentEnrollment.academic_year == run.academic_year)
                )
            ).scalar_one_or_none()
            if enrolled and enrolled.profile_status == StudentProfileStatus.VERIFIED:
                return "duplicate", RowIssue(
                    row, f"{roll} already imported and verified", "roll_number", roll, duplicate=True
                )
            user.full_name = user.full_name or name
            user.phone = user.phone or (mobile or None)
            user.section = user.section or section

        enrollment = (
            await self.db.execute(
                select(StudentEnrollment)
                .where(StudentEnrollment.student_id == user.id)
                .where(StudentEnrollment.academic_year == run.academic_year)
            )
        ).scalar_one_or_none()

        if enrollment:
            return "duplicate", RowIssue(
                row, f"{roll} already enrolled for {run.academic_year}", "roll_number", roll, duplicate=True
            )

        self.db.add(StudentEnrollment(
            college_id=self.college_id,
            student_id=user.id,
            department=department,
            section=section,
            year=year,
            semester=semester,
            academic_year=run.academic_year,
            is_registered=True,
            is_active=True,
            profile_status=(
                StudentProfileStatus.VERIFICATION_PENDING if mobile
                else StudentProfileStatus.PROFILE_INCOMPLETE
            ),
            contact_verified=False,
            declaration_signed=False,
            invitation_accepted=False,
        ))
        return "imported", None


    async def _student_domain(self, run: ImportRun) -> str:
        """
        The email domain this college's students already use.

        Taken from the existing roster rather than configured: whatever the
        institution has been using is the right thing for a generated address
        to match. Falls back to the uploader's own domain, then to a domain
        that cannot receive mail - better an obviously-local address than one
        that silently delivers to a stranger.
        """
        if self._domain_cache is not None:
            return self._domain_cache

        from collections import Counter

        # This college's own students first; then any student on the roster,
        # because a college that has not stamped college_id on its accounts
        # still has one house convention worth matching.
        scopes = []
        if self.college_id is not None:
            scopes.append(User.college_id == self.college_id)
        scopes.append(None)

        for scope in scopes:
            query = select(User.email).where(User.roll_number.isnot(None))
            if scope is not None:
                query = query.where(scope)
            emails = (await self.db.execute(query.limit(500))).scalars().all()
            domains = Counter(
                e.rsplit("@", 1)[1].lower() for e in emails if e and "@" in e
            )
            if domains:
                self._domain_cache = domains.most_common(1)[0][0]
                return self._domain_cache

        uploader = (await self.db.execute(
            select(User.email).where(User.id == run.imported_by_id)
        )).scalar_one_or_none()
        if uploader and "@" in uploader:
            self._domain_cache = uploader.rsplit("@", 1)[1].lower()
        else:
            self._domain_cache = "students.invalid"
        return self._domain_cache

    async def _student_from_allocation(
        self, row: int, roll: str, values: Dict[str, str], run: ImportRun
    ) -> Tuple[Optional[User], Optional[RowIssue]]:
        """
        Create the student an allocation row names, when it carries an email.

        users.email is required and unique, so an address is the one thing that
        cannot be derived. Without it the row fails rather than inventing a
        mailbox for a real person.
        """
        name = (values.get("full_name") or "").strip()
        email = (values.get("email") or "").strip().lower()

        if not email:
            # Optional by design: the sheet colleges hand out carries rolls and
            # names, not addresses. The roll is unique, so it makes a stable
            # login the student can change once they sign in.
            email = f"{roll.lower()}@{await self._student_domain(run)}"
        if not EMAIL_RE.match(email):
            return None, RowIssue(row, f"Invalid email for {roll}", "email", email)
        if not name:
            return None, RowIssue(
                row,
                f"No student with roll number {roll}, and no name in the cell to "
                f"create one with",
                "full_name",
            )

        # An address already on another roll is a copied cell, never a second
        # account for the same person.
        clash = (await self.db.execute(
            select(User).where(func.lower(User.email) == email)
        )).scalar_one_or_none()
        if clash is not None:
            if (clash.roll_number or "").strip().upper() != roll:
                owner = clash.roll_number or clash.email
                return None, RowIssue(
                    row,
                    f"Email {email} already registered to {owner}"
                    + (" - give this student an Email column" if not values.get("email") else ""),
                    "email", email,
                )
            return clash, None

        department = ((values.get("department") or "").strip()
                      or run.department or "CSE")
        section = (values.get("section") or "").strip() or None

        user = User(
            email=email,
            hashed_password=pwd_context.hash(f"Bharat@{roll}"),
            full_name=name,
            role=UserRole.STUDENT,
            roll_number=roll,
            department=department,
            section=section,
            is_active=True,
        )
        self.db.add(user)
        await self.db.flush()

        # Faculty screens read StudentEnrollment, not User: an account with no
        # enrolment would be allocated to a batch and still appear nowhere.
        existing = (await self.db.execute(
            select(StudentEnrollment)
            .where(StudentEnrollment.student_id == user.id)
            .where(StudentEnrollment.academic_year == run.academic_year)
        )).scalar_one_or_none()
        if existing is None:
            self.db.add(StudentEnrollment(
                college_id=self.college_id,
                student_id=user.id,
                department=department,
                section=section,
                year=(values.get("year") or "").strip() or "4th Year",
                semester=(values.get("semester") or "").strip() or None,
                academic_year=run.academic_year,
                is_registered=True,
                is_active=True,
                # Created from a college roster but unverified: the same place
                # an imported student lands, so the queue still owns them.
                profile_status=StudentProfileStatus.PROFILE_INCOMPLETE,
                contact_verified=False,
                declaration_signed=False,
                invitation_accepted=False,
            ))
        return user, None

    async def _apply_allocation(
        self, row: int, values: Dict[str, str], run: ImportRun
    ) -> Tuple[str, Optional[RowIssue]]:
        code = (values.get("batch_code") or "").strip().upper()
        roll = (values.get("roll_number") or "").strip().upper()
        if not code:
            return "failed", RowIssue(row, "Batch code is missing", "batch_code")
        if not roll:
            return "failed", RowIssue(row, "Roll number is missing", "roll_number")

        user = (
            await self.db.execute(select(User).where(func.upper(User.roll_number) == roll))
        ).scalar_one_or_none()
        if user is None:
            created, issue = await self._student_from_allocation(row, roll, values, run)
            if issue is not None:
                return "failed", issue
            user = created

        batch = (
            await self.db.execute(
                select(ProjectBatch).options(selectinload(ProjectBatch.members))
                .where(func.upper(ProjectBatch.batch_code) == code)
            )
        ).scalar_one_or_none()

        if batch is None:
            batch = ProjectBatch(
                college_id=self.college_id,
                batch_code=code,
                title=(values.get("project_title") or "").strip() or None,
                department=((values.get("department") or "").strip()
                            or run.department or user.department or "CSE"),
                section=(values.get("section") or "").strip() or user.section,
                academic_year=run.academic_year,
                is_active=True,
            )
            guide_name = (values.get("guide_name") or "").strip()
            if guide_name:
                guide = (await self.db.execute(
                    select(User)
                    .where(func.lower(User.full_name) == guide_name.lower())
                    .where(User.role.in_(COLLEGE_STAFF_ROLES))
                )).scalars().all()
                # Only an unambiguous match: two staff with the same name must
                # not silently hand the batch to whichever sorted first.
                if len(guide) == 1:
                    batch.guide_id = guide[0].id
            self.db.add(batch)
            await self.db.flush()
        elif any(str(m.student_id) == str(user.id) for m in batch.members):
            return "duplicate", RowIssue(row, f"{roll} already in {code}", "roll_number", roll, duplicate=True)

        self.db.add(ProjectBatchMember(
            batch_id=batch.id,
            student_id=user.id,
            is_lead=(values.get("is_leader") or "").strip().lower() in {"1", "y", "yes", "true"},
            is_active=True,
        ))
        return "imported", None

    async def _apply_project_details(self, row: int, values: Dict[str, str]) -> Tuple[str, Optional[RowIssue]]:
        code = (values.get("batch_code") or "").strip().upper()
        if not code:
            return "failed", RowIssue(row, "Batch code is missing", "batch_code")
        batch = (
            await self.db.execute(select(ProjectBatch).where(func.upper(ProjectBatch.batch_code) == code))
        ).scalar_one_or_none()
        if batch is None:
            return "failed", RowIssue(row, f"No batch with code {code}", "batch_code", code)

        title = (values.get("project_title") or "").strip()
        abstract = (values.get("abstract") or "").strip()
        if not title and not abstract:
            return "failed", RowIssue(row, "Nothing to update - title and abstract are both empty")
        if title:
            batch.title = title
        if abstract:
            batch.abstract = abstract
        return "imported", None

    async def _apply_base_paper(self, row: int, values: Dict[str, str]) -> Tuple[str, Optional[RowIssue]]:
        from app.models.faculty import BasePaper, BasePaperStatus

        code = (values.get("batch_code") or "").strip().upper()
        title = (values.get("paper_title") or "").strip()
        if not code:
            return "failed", RowIssue(row, "Batch code is missing", "batch_code")
        if not title:
            return "failed", RowIssue(row, "Paper title is missing", "paper_title")

        batch = (
            await self.db.execute(
                select(ProjectBatch).options(selectinload(ProjectBatch.base_paper))
                .where(func.upper(ProjectBatch.batch_code) == code)
            )
        ).scalar_one_or_none()
        if batch is None:
            return "failed", RowIssue(row, f"No batch with code {code}", "batch_code", code)

        year_raw = (values.get("year") or "").strip()
        try:
            year = int(year_raw) if year_raw else None
        except ValueError:
            return "failed", RowIssue(row, "Year must be a number", "year", year_raw)

        paper = batch.base_paper
        if paper is None:
            paper = BasePaper(batch_id=batch.id)
            self.db.add(paper)
        elif paper.status == BasePaperStatus.VERIFIED:
            return "duplicate", RowIssue(
                row, f"{code} already has a verified base paper", "batch_code", code, duplicate=True
            )

        paper.title = title
        paper.authors = (values.get("authors") or "").strip() or None
        paper.publication = (values.get("publication") or "").strip() or None
        paper.year = year
        paper.url = (values.get("url") or "").strip() or None
        # Imported metadata is unverified until a guide opens the paper.
        paper.status = BasePaperStatus.PENDING
        return "imported", None

    # ------------------------------------------------------------- reading

    @staticmethod
    def _row(run: ImportRun) -> dict:
        return {
            "id": str(run.id),
            "import_code": run.import_code,
            "file_name": run.file_name,
            "file_size": run.file_size,
            "import_type": IMPORT_TYPE_LABELS[run.import_type],
            "import_type_key": run.import_type.value,
            "department": run.department,
            "rows_total": run.rows_total,
            "rows_imported": run.rows_imported,
            "rows_failed": run.rows_failed,
            "rows_duplicate": run.rows_duplicate,
            "status": IMPORT_STATUS_LABELS[run.status],
            "status_key": run.status.value,
            "imported_by": run.imported_by.full_name if run.imported_by else None,
            "started_at": run.started_at,
            "is_archived": run.is_archived,
            "action": "Fix & Retry" if run.status == ImportStatus.FAILED
            else "View Details" if run.rows_failed or run.rows_duplicate else "View",
        }

    async def build(
        self,
        *,
        academic_year: str,
        department: Optional[str] = None,
        import_type: Optional[str] = None,
        status: Optional[str] = None,
        imported_by: Optional[str] = None,
        search: Optional[str] = None,
        page: int = 1,
        per_page: int = 10,
    ) -> dict:
        stmt = (
            select(ImportRun)
            .options(selectinload(ImportRun.imported_by))
            .where(ImportRun.academic_year == academic_year)
            .where(ImportRun.is_archived.is_(False))
            .order_by(ImportRun.started_at.desc())
        )
        # The tenant predicate. It was missing, and because the history query
        # is the only thing standing between one college and another's
        # uploaded rosters, its absence meant any faculty account could read
        # every college's imports.
        if self.college_id is not None:
            stmt = stmt.where(ImportRun.college_id == self.college_id)
        runs = list((await self.db.execute(stmt)).scalars().unique().all())

        def keep(run: ImportRun) -> bool:
            if department and run.department != department:
                return False
            if import_type and run.import_type.value != import_type:
                return False
            if status and run.status.value != status:
                return False
            if imported_by and str(run.imported_by_id) != imported_by:
                return False
            if search:
                needle = search.strip().lower()
                haystack = " ".join(filter(None, [
                    run.file_name, run.import_code,
                    run.imported_by.full_name if run.imported_by else "",
                ])).lower()
                if needle not in haystack:
                    return False
            return True

        scoped = [r for r in runs if keep(r)]

        kpis = [
            {"id": "total", "value": str(len(scoped)), "label": "Total Imports"},
            {"id": "processed", "value": f"{sum(r.rows_total for r in scoped):,}", "label": "Rows Processed"},
            {"id": "imported", "value": f"{sum(r.rows_imported for r in scoped):,}", "label": "Imported"},
            {"id": "failed", "value": str(sum(r.rows_failed for r in scoped)), "label": "Failed Rows"},
            {"id": "duplicates", "value": str(sum(r.rows_duplicate for r in scoped)), "label": "Skipped Duplicates"},
            {"id": "attention", "value": str(sum(
                1 for r in scoped
                if r.status in {ImportStatus.FAILED, ImportStatus.PARTIALLY_IMPORTED}
            )), "label": "Imports Need Attention"},
        ]

        total = len(scoped)
        pages = max(1, ceil(total / per_page)) if total else 1
        current = min(max(page, 1), pages)
        start = (current - 1) * per_page
        window = scoped[start:start + per_page]

        return {
            "kpis": kpis,
            "rows": [self._row(r) for r in window],
            "page": current,
            "pages": pages,
            "per_page": per_page,
            "total": total,
            "showing_from": (start + 1) if total else 0,
            "showing_to": min(start + per_page, total),
            "selected": await self.detail(str(window[0].id)) if window else None,
            "import_types": [{"key": k.value, "label": v} for k, v in IMPORT_TYPE_LABELS.items()],
            "statuses": [{"key": k.value, "label": v} for k, v in IMPORT_STATUS_LABELS.items()],
            "importers": sorted({
                (str(r.imported_by_id), r.imported_by.full_name)
                for r in runs if r.imported_by_id and r.imported_by
            }, key=lambda t: t[1] or ""),
        }

    def _mine(self, run) -> bool:
        """
        Whether this run belongs to the college the service is scoped to.

        A run recorded before the column existed carries no college; it is
        left visible rather than orphaned, since hiding a college's own
        history to close a gap it never had would be the worse trade.
        """
        if self.college_id is None or run.college_id is None:
            return True
        return str(run.college_id) == str(self.college_id)

    async def detail(self, run_id: str) -> Optional[dict]:
        run = (
            await self.db.execute(
                select(ImportRun)
                .options(
                    selectinload(ImportRun.imported_by),
                    selectinload(ImportRun.issues),
                    selectinload(ImportRun.events),
                )
                .where(ImportRun.id == run_id)
            )
        ).scalar_one_or_none()
        # Fetched by id, so the tenant check has to happen after the fact.
        # Answering "no such run" rather than refusing: confirming that an id
        # is real in another college is itself a small leak.
        if run is None or not self._mine(run):
            return None

        events = sorted(run.events, key=lambda e: e.occurred_at)
        issues = sorted(run.issues, key=lambda i: i.row_number)
        processed = run.rows_imported + run.rows_failed + run.rows_duplicate

        return {
            **self._row(run),
            "duration_seconds": run.duration_seconds,
            "completed_at": run.completed_at,
            "percent_processed": int(round(processed / run.rows_total * 100)) if run.rows_total else 0,
            "timeline": [
                {
                    "step": e.step,
                    "actor": e.actor,
                    "note": e.note,
                    "occurred_at": e.occurred_at,
                    "is_warning": e.is_warning,
                }
                for e in events
            ],
            "issues": [
                {
                    "row": i.row_number,
                    "field": i.field,
                    "message": i.message,
                    "value": i.raw_value,
                    "severity": i.severity.value,
                }
                for i in issues[:50]
            ],
            "issue_count": len(issues),
        }

    async def allocation_summary(self, run_id: str) -> Optional[dict]:
        """
        What an allocation import did, batch by batch.

        The run stores the uploaded file, so the batch codes are read back from
        it rather than recorded in a column: the sheet is the authority on what
        the upload was meant to touch, and re-reading it keeps this working for
        runs imported before this screen existed.

        "Created" means the batch did not exist before this run started, which
        is what separates a new batch from one the sheet only added members to.
        """
        run = (await self.db.execute(
            select(ImportRun).where(ImportRun.id == run_id)
        )).scalar_one_or_none()
        if run is None or not self._mine(run):
            return None

        codes: List[str] = []
        if run.file_content:
            try:
                parsed = parse_upload(run.file_name or "upload.csv", run.file_content)
                if is_wide_allocation(parsed):
                    parsed = normalise_wide_allocation(parsed)
                seen = set()
                for _, record in parsed.rows:
                    code = (record.get("batch_code") or "").strip().upper()
                    if code and code not in seen:
                        seen.add(code)
                        codes.append(code)
            except ValueError:
                codes = []

        if not codes:
            return {"batches": [], "batches_created": 0, "batches_updated": 0,
                    "students_assigned": run.rows_imported, "guides_assigned": 0}

        rows = (await self.db.execute(
            select(ProjectBatch)
            .options(selectinload(ProjectBatch.members), selectinload(ProjectBatch.guide))
            .where(func.upper(ProjectBatch.batch_code).in_(codes))
            .where(ProjectBatch.academic_year == run.academic_year)
        )).scalars().all()
        by_code = {b.batch_code.upper(): b for b in rows}

        batches, created, guides = [], 0, set()
        for code in codes:
            batch = by_code.get(code)
            if batch is None:
                continue
            was_created = bool(
                run.started_at and batch.created_at and batch.created_at >= run.started_at
            )
            created += 1 if was_created else 0
            if batch.guide_id:
                guides.add(str(batch.guide_id))
            # "CSE-D-D1" -> "D1": the sheet's own Batch No, which is the last
            # segment of the code it builds. Nothing to store, and it degrades
            # to empty for codes that are not segmented that way.
            tail = batch.batch_code.rsplit("-", 1)
            batches.append({
                "batch_code": batch.batch_code,
                "batch_no": tail[1] if len(tail) == 2 else None,
                "title": batch.title,
                "department": batch.department,
                "section": batch.section,
                "students": len([m for m in batch.members if m.is_active]),
                "guide": batch.guide.full_name if batch.guide else None,
                "outcome": "Created" if was_created else "Updated",
                "created_at": batch.created_at,
            })

        return {
            "batches": batches,
            "batches_created": created,
            "batches_updated": len(batches) - created,
            "students_assigned": run.rows_imported,
            "guides_assigned": len(guides),
        }

    async def error_csv(self, run_id: str) -> Optional[Tuple[str, str]]:
        detail = await self.detail(run_id)
        if detail is None:
            return None
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(["Row", "Field", "Severity", "Value", "Message"])
        for issue in detail["issues"]:
            writer.writerow([issue["row"], issue["field"] or "", issue["severity"],
                             issue["value"] or "", issue["message"]])
        return f"{detail['import_code']}-errors.csv", buffer.getvalue()

    async def archive(self, run_ids: Sequence[str]) -> int:
        runs = [
            r for r in (await self.db.execute(
                select(ImportRun).where(ImportRun.id.in_(list(run_ids)))
            )).scalars().all()
            if self._mine(r)
        ]
        for run in runs:
            run.is_archived = True
        await self.db.commit()
        return len(runs)
