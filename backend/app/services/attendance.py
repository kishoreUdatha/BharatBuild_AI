"""
Taking attendance.

The faculty portal has always been able to report attendance - a rate per
student, a roll-up for today, a floor to fall below - but nothing could record
it. Every number on that screen came from the seeder, so a real deployment
would have shown an empty page and a coordinator no way to fill it.

The unit is a section in a session - forenoon 09:30-12:30, afternoon
13:30-16:30 - which is what `attendance_records` enforces: one row per student
per date per session. Marking the same session twice corrects the earlier
marks rather than adding to them, because a register is a statement about a
session, not a log of who said what about it. A student who attends the
morning and goes home after lunch is present for one and absent for the other,
which a single row per day could not say.

Authority is the caller's job.
"""

from datetime import date, datetime, timedelta
from typing import Dict, List, Optional

from sqlalchemy import case, func, select
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.institution_time import (
    current_session, local_today, nearest_session, session_window,
)
from app.core.logging_config import logger
from app.models.faculty import (
    AttendanceRecord,
    AttendanceSession,
    AttendanceSessionLog,
    AttendanceStatus,
    ProjectBatch,
    ProjectBatchMember,
    StudentEnrollment,
)
from app.core.types import generate_uuid
from app.models.user import User

# Below this, a student is flagged. Mirrors the dashboard's floor.
FLOOR = 75.0

# LATE counts as attended - the student was there, just not on time - so only
# ABSENT lowers a rate. The read endpoint already defines it this way and a
# second definition here would eventually disagree with it.
# EXCUSED joins them: leave the college approved should not read as a student
# who did not turn up.
ATTENDED = {AttendanceStatus.PRESENT, AttendanceStatus.LATE, AttendanceStatus.EXCUSED}

STATUS_LABELS = {
    AttendanceStatus.PRESENT: "Present",
    AttendanceStatus.ABSENT: "Absent",
    AttendanceStatus.LATE: "Late",
    AttendanceStatus.EXCUSED: "Excused",
}

# The single letter the register shows, and what it means. Kept beside the
# labels so a new status cannot be added without deciding both.
STATUS_CODES = {
    AttendanceStatus.PRESENT: ("P", "Student is present"),
    AttendanceStatus.ABSENT: ("A", "Student is absent"),
    AttendanceStatus.LATE: ("L", "Student arrived late"),
    AttendanceStatus.EXCUSED: ("E", "Student is excused"),
}

MAX_BACKDATE_DAYS = 60

SESSION_LABELS = {
    AttendanceSession.FORENOON: "Forenoon (9:30 AM - 12:30 PM)",
    AttendanceSession.AFTERNOON: "Afternoon (1:30 PM - 4:30 PM)",
}


def parse_session(raw) -> AttendanceSession:
    """The session named, defaulting to the one the clock is nearest."""
    if raw is None or str(raw).strip() == "":
        return AttendanceSession(nearest_session())
    try:
        return AttendanceSession(str(raw).strip().lower())
    except ValueError:
        raise AttendanceError("Session must be forenoon or afternoon.")


class AttendanceError(Exception):
    """A refusal the caller can show the user as-is."""


def parse_day(raw: Optional[str]) -> date:
    """A YYYY-MM-DD day, defaulting to today, never in the future."""
    now = local_today()
    if not raw or not str(raw).strip():
        return now
    try:
        day = datetime.strptime(str(raw)[:10], "%Y-%m-%d").date()
    except ValueError:
        raise AttendanceError("Give a date like %s." % now.isoformat())
    if day > now:
        raise AttendanceError("Attendance cannot be taken for a day that has not happened.")
    if (now - day).days > MAX_BACKDATE_DAYS:
        raise AttendanceError(
            f"That day is more than {MAX_BACKDATE_DAYS} days ago. "
            "Ask an administrator to correct a register that old.")
    return day


def parse_status(raw: str) -> AttendanceStatus:
    try:
        return AttendanceStatus(str(raw).strip().lower())
    except ValueError:
        raise AttendanceError(
            "Status must be one of "
            + ", ".join(s.value for s in AttendanceStatus) + ".")


class AttendanceService:
    def __init__(self, db: AsyncSession):
        self.db = db

    # ------------------------------------------------------------- roster

    async def _cohort(self, department: str, year: Optional[str],
                      section: Optional[str], academic_year: str) -> List[tuple]:
        stmt = (
            select(StudentEnrollment, User)
            .join(User, StudentEnrollment.student_id == User.id)
            .where(StudentEnrollment.academic_year == academic_year)
            .where(StudentEnrollment.department == department)
            .where(StudentEnrollment.is_active.is_(True))
        )
        if year:
            stmt = stmt.where(StudentEnrollment.year == year)
        if section:
            stmt = stmt.where(StudentEnrollment.section == section)
        rows = (await self.db.execute(stmt)).all()
        return sorted(rows, key=lambda r: (r[1].roll_number or "", r[1].full_name or ""))

    async def roster(
        self,
        *,
        department: str,
        year: Optional[str],
        section: Optional[str],
        academic_year: str,
        on: date,
        session: AttendanceSession,
    ) -> dict:
        """
        The register for one section in one session.

        Every student carries the mark already recorded for that session, if
        any, so reopening it shows what was taken rather than a blank sheet
        someone might save over.
        """
        cohort = await self._cohort(department, year, section, academic_year)
        student_ids = [str(e.student_id) for e, _ in cohort]

        existing: Dict[str, AttendanceRecord] = {}
        if student_ids:
            for record in (await self.db.execute(
                select(AttendanceRecord)
                .where(AttendanceRecord.student_id.in_(student_ids))
                .where(AttendanceRecord.attendance_date == on)
                .where(AttendanceRecord.session == session)
            )).scalars().all():
                existing[str(record.student_id)] = record

        rates = await self._rates(student_ids, academic_year)

        students = []
        for enrollment, user in cohort:
            key = str(enrollment.student_id)
            record = existing.get(key)
            rate = rates.get(key)
            students.append({
                "student_id": key,
                "full_name": user.full_name,
                "roll_number": user.roll_number,
                "section": enrollment.section,
                "status": record.status.value if record else None,
                "status_label": STATUS_LABELS[record.status] if record else "Not marked",
                "attendance_rate": rate,
                "below_floor": rate is not None and rate < FLOOR,
            })

        marked = sum(1 for s in students if s["status"])
        return {
            "date": on,
            "session": session.value,
            "session_label": SESSION_LABELS[session],
            "sessions": self.session_options(on),
            "department": department,
            "year": year,
            "section": section,
            "academic_year": academic_year,
            "students": students,
            "total": len(students),
            "marked": marked,
            # A day nobody has touched reads differently from one taken and
            # saved, so the screen can say which it is looking at.
            "already_taken": marked > 0,
            "counts": self._counts(students),
            "floor": FLOOR,
            "statuses": [
                {"value": s.value, "label": STATUS_LABELS[s]} for s in AttendanceStatus
            ],
        }


    # --------------------------------------------------------- the two halves

    @staticmethod
    def session_options(on: Optional[date] = None) -> List[dict]:
        """
        The day's two sessions, and which one the clock is inside.

        `open` is advisory. Marking is never refused for being outside the
        window: catching up on a register at the end of the day is the normal
        case, and a system that forbade it would just get a wrong day marked
        instead.
        """
        live = current_session()
        today = on is None or on == local_today()
        out = []
        for member in AttendanceSession:
            start, end = session_window(member.value)
            out.append({
                "value": member.value,
                "label": SESSION_LABELS[member],
                "starts": start.strftime("%H:%M"),
                "ends": end.strftime("%H:%M"),
                # Only meaningful for today; a past day has no live session.
                "open": today and live == member.value,
            })
        return out

    async def trainer_roster(self, trainer: User, *, batch_code: Optional[str],
                             on: date, session: AttendanceSession) -> dict:
        """
        The register for a trainer's own batch.

        A trainer marks the students they guide, not a whole department
        section - their authority comes from the batch, so the register is
        built from its members rather than from an enrolment cohort.
        """
        batches = await self._my_batches(trainer, on)
        if not batches:
            raise AttendanceError("You have no batches assigned to you yet.")

        chosen = next((b for b in batches if b.batch_code == batch_code), batches[0])

        rows = (await self.db.execute(
            select(User, ProjectBatchMember)
            .join(ProjectBatchMember, ProjectBatchMember.student_id == User.id)
            .where(ProjectBatchMember.batch_id == chosen.id)
            .order_by(User.roll_number, User.full_name)
        )).all()
        student_ids = [str(u.id) for u, _ in rows]

        existing: Dict[str, AttendanceRecord] = {}
        if student_ids:
            for record in (await self.db.execute(
                select(AttendanceRecord)
                .where(AttendanceRecord.student_id.in_(student_ids))
                .where(AttendanceRecord.attendance_date == on)
                .where(AttendanceRecord.session == session)
            )).scalars().all():
                existing[str(record.student_id)] = record

        year = chosen.academic_year
        rates = await self._rates(student_ids, year) if student_ids else {}

        students = []
        for user, _member in rows:
            key = str(user.id)
            record = existing.get(key)
            rate = rates.get(key)
            students.append({
                "student_id": key,
                "full_name": user.full_name,
                "roll_number": user.roll_number,
                "section": chosen.section,
                "status": record.status.value if record else None,
                "status_label": STATUS_LABELS[record.status] if record else "Not marked",
                "attendance_rate": rate,
                "below_floor": rate is not None and rate < FLOOR,
            })

        marked = sum(1 for row in students if row["status"])
        return {
            "date": on,
            "session": session.value,
            "session_label": SESSION_LABELS[session],
            "sessions": self.session_options(on),
            "batch_code": chosen.batch_code,
            "batches": [{"code": b.batch_code,
                         "title": b.title or b.batch_code} for b in batches],
            "academic_year": year,
            "students": students,
            "total": len(students),
            "marked": marked,
            "already_taken": marked > 0,
            "counts": self._counts(students),
            "floor": FLOOR,
            "statuses": [
                {"value": st.value, "label": STATUS_LABELS[st]} for st in AttendanceStatus
            ],
        }


    # ------------------------------------------------------- the whole day

    @staticmethod
    def _academic_year(on: date) -> str:
        """June starts a new academic year, matching the rest of the portal."""
        start = on.year if on.month >= 6 else on.year - 1
        return f"{start}-{str(start + 1)[-2:]}"


    @staticmethod
    def _filter_options(batches, department: Optional[str],
                        section: Optional[str], batch_code: Optional[str]) -> dict:
        """
        Options that narrow with the choices already made.

        Each list is built from the batches that survive the *other* filters,
        so picking CSE and section A leaves only that section's batches in the
        batch list. Offering all of them was how a section filter appeared not
        to work: the section was applied, but the dropdown beneath it still
        listed batches that no longer belonged.
        """
        def matches(batch, *, ignore: str) -> bool:
            if ignore != "department" and department and batch.department != department:
                return False
            if ignore != "section" and section and (batch.section or "") != section:
                return False
            return True

        return {
            # The top of the chain, so it is never narrowed by what sits below.
            "departments": sorted({b.department for b in batches if b.department}),
            "sections": sorted({b.section for b in batches
                                if b.section and matches(b, ignore="section")}),
            "batches": [
                {"code": b.batch_code, "title": b.title or b.batch_code,
                 "section": b.section, "department": b.department}
                for b in batches if matches(b, ignore="")
            ],
            "department": department,
            "section": section,
            "batch_code": batch_code,
        }

    async def _my_batches(self, trainer: User, on: Optional[date] = None):
        """
        Every batch this trainer answers for - not only the ones they guide.

        Authority comes from three places: batches they guide or review, the
        sections they coordinate, and the departments they run. Reading only
        `guide_id` is what made a trainer with three sections see one, and it
        disagreed with My Batches, which has always used this rule.
        """
        from app.services.faculty_authority import FacultyAuthority

        year = self._academic_year(on or local_today())
        ids = await FacultyAuthority(self.db).managed_batch_ids(trainer, year)
        if not ids:
            return []
        return list((await self.db.execute(
            select(ProjectBatch)
            .where(ProjectBatch.id.in_(list(ids)))
            .where(ProjectBatch.is_active.is_(True))
            .order_by(ProjectBatch.batch_code)
        )).scalars().all())

    async def day_register(
        self, trainer: User, *, on: date,
        department: Optional[str] = None,
        section: Optional[str] = None,
        batch_code: Optional[str] = None,
        page: int = 1, per_page: int = 10,
    ) -> dict:
        """
        Every student this trainer is responsible for, with both sessions.

        One screen rather than two: a trainer takes the morning and the
        afternoon for the same faces, and flipping between two registers to
        answer "who was here all day" is how half-days get mis-recorded.
        """
        batches = await self._my_batches(trainer, on)
        if not batches:
            raise AttendanceError("You have no batches assigned to you yet.")

        chosen = [b for b in batches
                  if (not department or b.department == department)
                  and (not section or (b.section or "") == section)
                  and (not batch_code or b.batch_code == batch_code)]

        rows = []
        if chosen:
            rows = (await self.db.execute(
                select(User, ProjectBatch)
                .join(ProjectBatchMember, ProjectBatchMember.student_id == User.id)
                .join(ProjectBatch, ProjectBatch.id == ProjectBatchMember.batch_id)
                .where(ProjectBatch.id.in_([b.id for b in chosen]))
                .order_by(User.roll_number, User.full_name)
            )).all()

        student_ids = [str(u.id) for u, _ in rows]
        marks: Dict[str, Dict[str, AttendanceRecord]] = {}
        if student_ids:
            for record in (await self.db.execute(
                select(AttendanceRecord)
                .where(AttendanceRecord.student_id.in_(student_ids))
                .where(AttendanceRecord.attendance_date == on)
            )).scalars().all():
                marks.setdefault(str(record.student_id), {})[record.session.value] = record

        def cell(student_id: str, session: AttendanceSession) -> dict:
            record = marks.get(student_id, {}).get(session.value)
            return {
                "status": record.status.value if record else None,
                "status_label": STATUS_LABELS[record.status] if record else "Not marked",
                "code": STATUS_CODES[record.status][0] if record else None,
                "remarks": record.remarks if record else None,
            }

        students = []
        for user, batch in rows:
            key = str(user.id)
            fn = cell(key, AttendanceSession.FORENOON)
            an = cell(key, AttendanceSession.AFTERNOON)
            # Half a day each. A student marked for only one session is scored
            # out of the one that was taken, not out of two.
            scored = [c for c in (fn, an) if c["status"]]
            attended = sum(1 for c in scored
                           if AttendanceStatus(c["status"]) in ATTENDED)
            students.append({
                "student_id": key,
                "roll_number": user.roll_number,
                "full_name": user.full_name,
                "batch_code": batch.batch_code,
                "section": batch.section,
                "forenoon": fn,
                "afternoon": an,
                "day_percent": round(attended / len(scored) * 100) if scored else None,
            })

        total = len(students)
        start = max(page - 1, 0) * per_page
        page_rows = students[start:start + per_page]

        def tally(field: str) -> dict:
            counts = {st.value: 0 for st in AttendanceStatus}
            counts["unmarked"] = 0
            for row in students:
                value = row[field]["status"]
                counts[value if value else "unmarked"] += 1
            marked = total - counts["unmarked"]
            counts["marked"] = marked
            # Over the students actually marked, so it answers "of the ones I
            # have recorded, how many were here". It is NOT the cohort's
            # attendance until the register is finished - ten present out of
            # ten marked is 100% of ten, not 100% of a hundred and eighty-one,
            # and a screen that prints the first as "100% today" is lying.
            counts["rate"] = round(
                sum(counts[st.value] for st in ATTENDED) / marked * 100, 2
            ) if marked else None
            counts["complete"] = total > 0 and marked == total
            return counts

        fn_counts, an_counts = tally("forenoon"), tally("afternoon")
        logs = {
            log.session.value: log
            for log in (await self.db.execute(
                select(AttendanceSessionLog)
                .where(AttendanceSessionLog.trainer_id == trainer.id)
                .where(AttendanceSessionLog.attendance_date == on)
            )).scalars().all()
        }

        live = current_session()
        today = on == local_today()
        sessions = []
        for member, counts in ((AttendanceSession.FORENOON, fn_counts),
                               (AttendanceSession.AFTERNOON, an_counts)):
            start_t, end_t = session_window(member.value)
            log = logs.get(member.value)
            sessions.append({
                "value": member.value,
                "label": SESSION_LABELS[member].split(" (")[0],
                "window": f"{start_t.strftime('%I:%M %p').lstrip('0')} - "
                          f"{end_t.strftime('%I:%M %p').lstrip('0')} IST",
                "open": log is None or log.submitted_at is None,
                "live": today and live == member.value,
                "taken_by": (log.trainer.full_name
                             if log and log.trainer else None),
                "started_at": log.started_at if log else None,
                "submitted_at": log.submitted_at if log else None,
                "counts": counts,
            })

        # The day's figure, averaged over the two sessions - but only once
        # both registers are actually complete. A half-taken day has no
        # attendance rate yet, and publishing one invites it to be read as the
        # cohort's.
        rates = [s["counts"]["rate"] for s in sessions if s["counts"]["rate"] is not None]
        complete = all(s["counts"].get("complete") for s in sessions) and total > 0
        overall = round(sum(rates) / len(rates), 2) if (rates and complete) else None

        return {
            "date": on,
            "students": page_rows,
            "total": total,
            "page": page,
            "per_page": per_page,
            "pages": max(1, -(-total // per_page)) if total else 1,
            "sessions": sessions,
            "overall_rate": overall,
            # How far through the day's two registers the trainer is, which is
            # what there is to report while one of them is unfinished.
            "marked_total": sum(int(s["counts"].get("marked") or 0) for s in sessions),
            "marked_of": total * len(sessions),
            "complete": complete,
            "kpis": {
                "total": total,
                "present": fn_counts[AttendanceStatus.PRESENT.value],
                "absent": fn_counts[AttendanceStatus.ABSENT.value],
                "late": fn_counts[AttendanceStatus.LATE.value],
                "excused": fn_counts[AttendanceStatus.EXCUSED.value],
            },
            "filters": self._filter_options(batches, department, section, batch_code),
            "statuses": [
                {"value": st.value, "label": STATUS_LABELS[st],
                 "code": STATUS_CODES[st][0], "hint": STATUS_CODES[st][1]}
                for st in AttendanceStatus
            ],
            "floor": FLOOR,
        }


    # ------------------------------------------------------------- importing

    IMPORT_COLUMNS = ["Roll Number", "Student Name", "Forenoon", "Afternoon",
                      "Forenoon Remarks", "Afternoon Remarks"]

    # What a trainer might actually type in a status cell. The single letters
    # are what the register itself shows, so they are the likeliest.
    STATUS_WORDS = {
        "p": AttendanceStatus.PRESENT, "present": AttendanceStatus.PRESENT,
        "a": AttendanceStatus.ABSENT, "absent": AttendanceStatus.ABSENT,
        "l": AttendanceStatus.LATE, "late": AttendanceStatus.LATE,
        "e": AttendanceStatus.EXCUSED, "excused": AttendanceStatus.EXCUSED,
    }

    async def import_template(self, trainer: User, *, on: date,
                              batch_code: Optional[str] = None) -> bytes:
        """
        The sheet to fill in, already carrying this trainer's students.

        Rolls are pre-printed rather than left blank: a register typed from
        memory is where wrong roll numbers come from, and a sheet that already
        knows who is in the room cannot introduce one.
        """
        import io as _io

        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        register = await self.day_register(
            trainer, on=on, batch_code=batch_code, page=1, per_page=10000)

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Attendance"

        head = Font(bold=True, color="FFFFFF")
        fill = PatternFill("solid", fgColor="1B2A6B")
        for column, heading in enumerate(self.IMPORT_COLUMNS, start=1):
            cell = sheet.cell(row=1, column=column)
            cell.value = heading
            cell.font = head
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

        for row in register["students"]:
            sheet.append([
                row["roll_number"] or "",
                row["full_name"],
                # Whatever is already recorded, so a part-taken day round-trips
                # instead of being wiped by an upload.
                (row["forenoon"]["code"] or ""),
                (row["afternoon"]["code"] or ""),
                row["forenoon"]["remarks"] or "",
                row["afternoon"]["remarks"] or "",
            ])

        for column, width in zip("ABCDEF", (18, 30, 12, 12, 30, 30)):
            sheet.column_dimensions[column].width = width
        sheet.freeze_panes = "A2"

        guide = wb.create_sheet("How to fill this in")
        for line in [
            [f"Attendance for {on:%d %B %Y}"],
            [],
            ["Put one of these in the Forenoon and Afternoon columns:"],
            ["P", "Present"],
            ["A", "Absent"],
            ["L", "Late - the student was there, just not on time"],
            ["E", "Excused - leave the college accepted; counts as attended"],
            [],
            ["Leave a cell blank to leave that session alone. It is not the same"],
            ["as marking somebody absent."],
            [],
            ["Do not add or rename columns, and do not change the roll numbers -"],
            ["they are how each row finds its student."],
            [],
            ["Forenoon runs 09:30-12:30 and Afternoon 13:30-16:30 IST."],
        ]:
            guide.append(line)
        guide.column_dimensions["A"].width = 68
        guide["A1"].font = Font(bold=True, size=13)

        buffer = _io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    async def import_register(self, trainer: User, *, raw: bytes, filename: str,
                              on: date, dry_run: bool = True) -> dict:
        """
        Read a filled-in sheet.

        `dry_run` is the Validate step: every row is checked and nothing is
        written, so the trainer sees what confirming would do. Either way a
        bad row is reported against its row number and skipped while the rest
        of the sheet still lands - one wrong roll number should not cost a
        trainer the other fifty-nine.
        """
        rows = self._read_sheet(raw, filename)
        if not rows:
            raise AttendanceError("That sheet has no rows under the headings.")

        register = await self.day_register(
            trainer, on=on, page=1, per_page=10000)
        by_roll = {(r["roll_number"] or "").strip().upper(): r
                   for r in register["students"] if r["roll_number"]}

        preview, seen = [], set()
        marks: Dict[str, List[dict]] = {"forenoon": [], "afternoon": []}
        for number, row in enumerate(rows, start=2):
            roll = str(row.get("roll number") or "").strip().upper()
            entry = {"row": number, "roll_number": roll,
                     "student": None, "forenoon": None, "afternoon": None,
                     "issues": [], "warnings": []}
            if not roll:
                entry["issues"].append("No roll number in this row.")
                preview.append(entry)
                continue
            student = by_roll.get(roll)
            if student is None:
                entry["issues"].append(
                    f"{roll} is not one of your students on this date.")
                preview.append(entry)
                continue
            if roll in seen:
                entry["issues"].append(f"{roll} appears more than once in the sheet.")
                preview.append(entry)
                continue
            seen.add(roll)
            entry["student"] = student["full_name"]

            # Parsed first, written only if the whole row is clean. A row
            # reported as rejected must not quietly land half of itself -
            # "3 rejected" and four marks written is the kind of arithmetic
            # nobody can reconcile against a register later.
            staged = []
            for session in ("forenoon", "afternoon"):
                cell = str(row.get(session) or "").strip().lower()
                if not cell:
                    continue
                status = self.STATUS_WORDS.get(cell)
                if status is None:
                    entry["issues"].append(
                        f"{session.title()} says {cell!r}. Use P, A, L or E.")
                    continue
                entry[session] = STATUS_LABELS[status]
                remarks = str(row.get(f"{session} remarks") or "").strip()[:300]
                staged.append((session, {
                    "student_id": student["student_id"],
                    "status": status.value,
                    "remarks": remarks or None,
                }))

            if entry["issues"]:
                # Keep what parsed on screen so the trainer sees what the row
                # was trying to say, but record none of it.
                entry["warnings"].append("Nothing from this row was recorded.")
            else:
                for session, mark in staged:
                    marks[session].append(mark)
                if not entry["forenoon"] and not entry["afternoon"]:
                    entry["warnings"].append(
                        "Both sessions were blank, so nothing changes.")
            preview.append(entry)

        ok = [p for p in preview if not p["issues"]]
        result = {
            "dry_run": dry_run,
            "date": on,
            "rows": preview,
            "total": len(preview),
            "valid": len(ok),
            "rejected": len(preview) - len(ok),
            "forenoon_marks": len(marks["forenoon"]),
            "afternoon_marks": len(marks["afternoon"]),
        }
        if dry_run:
            return result

        applied = {}
        for session, entries in marks.items():
            if entries:
                applied[session] = await self.trainer_mark(
                    trainer, batch_code=None, on=on,
                    session=AttendanceSession(session), marks=entries)
        result["applied"] = applied
        logger.info(f"[Attendance] {trainer.email} imported {len(ok)} rows for {on}")
        return result

    @staticmethod
    def _read_sheet(raw: bytes, filename: str) -> List[dict]:
        """Rows as dicts keyed by lower-cased heading, from .xlsx or .csv."""
        import csv as _csv
        import io as _io

        if filename.lower().endswith(".csv"):
            text = raw.decode("utf-8-sig", errors="replace")
            return [{(k or "").strip().lower(): v
                     for k, v in row.items()}
                    for row in _csv.DictReader(_io.StringIO(text))]
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise AttendanceError("XLSX support needs openpyxl on the server") from exc
        try:
            wb = load_workbook(_io.BytesIO(raw), read_only=True, data_only=True)
        except Exception as exc:
            raise AttendanceError("That file could not be read as a spreadsheet.") from exc
        sheet = wb[wb.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headings = [str(h or "").strip().lower() for h in rows[0]]
        out = []
        for values in rows[1:]:
            if all(v in (None, "") for v in values):
                continue
            out.append({h: v for h, v in zip(headings, values) if h})
        return out


    # ----------------------------------------------------------- the month

    async def month_register(
        self, trainer: User, *, year: int, month: int,
        department: Optional[str] = None,
        section: Optional[str] = None,
        batch_code: Optional[str] = None,
        search: Optional[str] = None,
        page: int = 1, per_page: int = 25,
    ) -> dict:
        """
        Every day of a month for every student the trainer answers for.

        The register a college actually files: a grid of students against
        days, two marks per day. It answers the question a daily screen
        cannot - whether a student has been drifting - and it is what a
        shortage notice has to be built from.

        Days with nothing recorded for anybody are dropped. A month of blank
        columns for Sundays and holidays makes the grid unreadable and says
        nothing; the ones that survive are the days the college actually met.
        """
        if not 1 <= month <= 12:
            raise AttendanceError("That is not a month.")
        first = date(year, month, 1)
        last = date(year + (month == 12), (month % 12) + 1, 1) - timedelta(days=1)

        batches = await self._my_batches(trainer, first)
        if not batches:
            raise AttendanceError("You have no batches assigned to you yet.")
        chosen = [b for b in batches
                  if (not department or b.department == department)
                  and (not section or (b.section or "") == section)
                  and (not batch_code or b.batch_code == batch_code)]

        rows = []
        if chosen:
            stmt = (
                select(User, ProjectBatch)
                .join(ProjectBatchMember, ProjectBatchMember.student_id == User.id)
                .join(ProjectBatch, ProjectBatch.id == ProjectBatchMember.batch_id)
                .where(ProjectBatch.id.in_([b.id for b in chosen]))
            )
            if search:
                term = f"%{search.strip().lower()}%"
                stmt = stmt.where(
                    func.lower(User.full_name).like(term)
                    | func.lower(User.roll_number).like(term))
            rows = (await self.db.execute(
                stmt.order_by(User.roll_number, User.full_name))).all()

        student_ids = [str(u.id) for u, _ in rows]
        marks: Dict[str, Dict[str, Dict[str, AttendanceRecord]]] = {}
        days_seen = set()
        if student_ids:
            for record in (await self.db.execute(
                select(AttendanceRecord)
                .where(AttendanceRecord.student_id.in_(student_ids))
                .where(AttendanceRecord.attendance_date >= first)
                .where(AttendanceRecord.attendance_date <= last)
            )).scalars().all():
                key = record.attendance_date.isoformat()
                days_seen.add(record.attendance_date)
                marks.setdefault(str(record.student_id), {}) \
                     .setdefault(key, {})[record.session.value] = record

        # Every day of the month, not only the ones with marks: a calendar
        # with holes in it stops being a calendar, and the gaps are exactly
        # what a trainer scans for.
        days = [first + timedelta(days=n) for n in range((last - first).days + 1)]
        # A class was held on a day somebody was marked. Weekends are shown but
        # never counted, so a Sunday cannot drag a percentage down.
        held = {d for d in days if d in days_seen}
        working = [d for d in days if d.weekday() < 5]

        held_count = len(held)
        students = []
        for user, batch in rows:
            mine = marks.get(str(user.id), {})
            cells, attended, recorded = [], 0, 0
            absent_days, late_days, present_days = set(), set(), set()
            # Counted per session as well as per day: a student who never
            # misses a morning but keeps missing the afternoon is a pattern one
            # combined figure hides completely.
            per_session = {
                member.value: {st.value: 0 for st in AttendanceStatus}
                for member in AttendanceSession
            }
            for day in days:
                key = day.isoformat()
                pair = mine.get(key, {})
                cell = {"date": key}
                for session in ("forenoon", "afternoon"):
                    record = pair.get(session)
                    if record is None:
                        cell[session] = None
                        continue
                    recorded += 1
                    per_session[session][record.status.value] += 1
                    if record.status in ATTENDED:
                        attended += 1
                    if record.status == AttendanceStatus.ABSENT:
                        absent_days.add(day)
                    elif record.status == AttendanceStatus.LATE:
                        late_days.add(day)
                    cell[session] = {
                        "status": record.status.value,
                        "code": STATUS_CODES[record.status][0],
                        "remarks": record.remarks,
                    }
                cells.append(cell)
                # A day counts as present only when nothing on it was an
                # absence: half a day attended is not a day attended.
                if day in held and day not in absent_days and mine.get(key):
                    present_days.add(day)

            # Over the sessions actually recorded for this student, so a day
            # the college did not meet cannot count against them.
            rate = round(attended / recorded * 100, 2) if recorded else None

            def breakdown_for(counts: dict) -> dict:
                # Each session is scored against its own register, not against
                # the day. A college that only ever takes the morning would
                # otherwise show every student at 0% for the afternoon - which
                # reads as absent, when in truth nobody was ever marked.
                taken = sum(counts.values())

                def share(count: int) -> Optional[float]:
                    return round(count / taken * 100, 2) if taken else None

                return {
                    "present": share(counts[AttendanceStatus.PRESENT.value]),
                    "absent": share(counts[AttendanceStatus.ABSENT.value]),
                    "late": share(counts[AttendanceStatus.LATE.value]),
                    "excused": share(counts[AttendanceStatus.EXCUSED.value]),
                    "recorded": taken,
                }

            breakdown = {name: breakdown_for(counts)
                         for name, counts in per_session.items()}

            students.append({
                "student_id": str(user.id),
                "roll_number": user.roll_number,
                "full_name": user.full_name,
                "batch_code": batch.batch_code,
                "days": cells,
                "sessions_recorded": recorded,
                "sessions_attended": attended,
                "absent": recorded - attended,
                "present_days": len(present_days),
                "absent_days": len(absent_days),
                "late_days": len(late_days),
                "classes_held": held_count,
                "breakdown": breakdown,
                # Over the sessions actually recorded, for the same reason: a
                # session nobody took is not an absence, and assuming two a day
                # halves the score of a college that only takes one.
                "overall": rate,
                "rate": rate,
                "below_floor": rate is not None and rate < FLOOR,
            })

        total = len(students)
        start = max(page - 1, 0) * per_page
        overall = self._month_overall(students_all=students)
        return {
            "month": f"{year:04d}-{month:02d}",
            "label": first.strftime("%B %Y"),
            "from": first,
            "to": last,
            "days": [
                {
                    "date": d.isoformat(),
                    "day": d.day,
                    "weekday": d.strftime("%a"),
                    "weekend": d.weekday() >= 5,
                    # No marks at all: a holiday, or a register nobody took.
                    "held": d in held,
                }
                for d in days
            ],
            "classes_held": len(held),
            "working_days": len(working),
            "students": students[start:start + per_page],
            "total": total,
            "page": page,
            "per_page": per_page,
            "pages": max(1, -(-total // per_page)) if total else 1,
            "filters": {**self._filter_options(batches, department, section, batch_code),
                        "search": search},
            "statuses": [
                {"value": st.value, "label": STATUS_LABELS[st],
                 "code": STATUS_CODES[st][0]} for st in AttendanceStatus
            ],
            # The cohort's month, over every session recorded for anybody in
            # it. Computed across all students, not the page being shown.
            "overall": overall,
            "sessions": [
                {"value": member.value,
                 "label": SESSION_LABELS[member].split(" (")[0],
                 "window": SESSION_LABELS[member].split("(")[1].rstrip(")")}
                for member in AttendanceSession
            ],
            "floor": FLOOR,
        }

    @staticmethod
    def _month_overall(*, students_all: List[dict]) -> dict:
        """Present, absent and late as shares of everything recorded."""
        recorded = sum(row["sessions_recorded"] for row in students_all)
        if not recorded:
            return {"present": None, "absent": None, "late": None, "recorded": 0}
        attended = sum(row["sessions_attended"] for row in students_all)
        absent = sum(row["absent"] for row in students_all)
        return {
            "present": round(attended / recorded * 100, 2),
            "absent": round(absent / recorded * 100, 2),
            # Late students are present; this is how often lateness happened,
            # which is a different question from whether they attended.
            "late": round(sum(row["late_days"] for row in students_all)
                          / recorded * 100, 2),
            "recorded": recorded,
        }


    # ----------------------------------------------------------- the report

    async def month_summary(
        self, trainer: User, *, year: int, month: int,
        department: Optional[str] = None,
        section: Optional[str] = None,
        batch_code: Optional[str] = None,
    ) -> dict:
        """
        The month rolled up, over every student - never a page of them.

        A summary computed from whichever twenty-five students happened to be
        on screen is worse than no summary: it looks authoritative and changes
        when you turn the page.
        """
        # The grid already resolves the cohort, the filters and the days met.
        # Asking for every student keeps one definition of who is counted.
        grid = await self.month_register(
            trainer, year=year, month=month, department=department,
            section=section, batch_code=batch_code, page=1, per_page=100000)
        students = grid["students"]

        counts = {st.value: 0 for st in AttendanceStatus}
        per_day: Dict[str, Dict[str, int]] = {}
        per_session = {member.value: {st.value: 0 for st in AttendanceStatus}
                       for member in AttendanceSession}
        for row in students:
            for cell in row["days"]:
                bucket = per_day.setdefault(cell["date"], {"attended": 0, "recorded": 0})
                for name in ("forenoon", "afternoon"):
                    mark = cell[name]
                    if not mark:
                        continue
                    counts[mark["status"]] += 1
                    per_session[name][mark["status"]] += 1
                    bucket["recorded"] += 1
                    if mark["status"] != AttendanceStatus.ABSENT.value:
                        bucket["attended"] += 1

        marked = sum(counts.values())
        attended = marked - counts[AttendanceStatus.ABSENT.value]

        def share(part: int, whole: int) -> Optional[float]:
            return round(part / whole * 100, 2) if whole else None

        def session_row(name: str) -> dict:
            row = per_session[name]
            taken = sum(row.values())
            member = AttendanceSession(name)
            start, end = session_window(name)
            return {
                "value": name,
                "label": SESSION_LABELS[member].split(" (")[0],
                "window": f"{start.strftime('%I:%M %p').lstrip('0')} - "
                          f"{end.strftime('%I:%M %p').lstrip('0')}",
                # Days this session was actually taken, not days in the month.
                "classes_held": len({d for d, v in per_day.items() if v["recorded"]}
                                    ) if taken else 0,
                "recorded": taken,
                "present": share(row[AttendanceStatus.PRESENT.value], taken),
                "absent": share(row[AttendanceStatus.ABSENT.value], taken),
                "late": share(row[AttendanceStatus.LATE.value], taken),
            }

        ranked = sorted(
            (r for r in students if r["overall"] is not None),
            key=lambda r: (-r["overall"], r["full_name"]))
        brief = lambda r: {
            "student_id": r["student_id"], "full_name": r["full_name"],
            "roll_number": r["roll_number"], "overall": r["overall"],
        }

        return {
            "month": grid["month"],
            "label": grid["label"],
            "students": len(students),
            "classes_held": grid["classes_held"],
            "working_days": grid["working_days"],
            "marked": marked,
            "counts": {
                "present": counts[AttendanceStatus.PRESENT.value],
                "absent": counts[AttendanceStatus.ABSENT.value],
                "late": counts[AttendanceStatus.LATE.value],
                "excused": counts[AttendanceStatus.EXCUSED.value],
            },
            "shares": {
                "present": share(counts[AttendanceStatus.PRESENT.value], marked),
                "absent": share(counts[AttendanceStatus.ABSENT.value], marked),
                "late": share(counts[AttendanceStatus.LATE.value], marked),
                "excused": share(counts[AttendanceStatus.EXCUSED.value], marked),
            },
            # Late counts as attended, so this is not the sum of "present".
            "overall": share(attended, marked),
            "sessions": [session_row(m.value) for m in AttendanceSession],
            "by_day": [
                {
                    "date": day["date"], "day": day["day"], "weekday": day["weekday"],
                    "weekend": day["weekend"], "held": day["held"],
                    "rate": share(per_day.get(day["date"], {}).get("attended", 0),
                                  per_day.get(day["date"], {}).get("recorded", 0)),
                }
                for day in grid["days"]
            ],
            "top": [brief(r) for r in ranked[:5]],
            "bottom": [brief(r) for r in reversed(ranked[-5:])] if ranked else [],
            "below_floor": sum(1 for r in students if r["below_floor"]),
            "trend": await self._trend(trainer, year=year, month=month,
                                       department=department, section=section,
                                       batch_code=batch_code),
            "floor": FLOOR,
            "filters": grid["filters"],
        }

    async def _trend(self, trainer: User, *, year: int, month: int,
                     department: Optional[str], section: Optional[str],
                     batch_code: Optional[str], months: int = 6) -> List[dict]:
        """
        This month and the five before it, as one figure each.

        Read in one query rather than six passes of the grid: a trend line is
        a shape, and it does not need every student's cells to draw it.
        """
        first = date(year, month, 1)
        start = first
        for _ in range(months - 1):
            start = date(start.year - (start.month == 1),
                         12 if start.month == 1 else start.month - 1, 1)
        last = date(year + (month == 12), (month % 12) + 1, 1) - timedelta(days=1)

        batches = await self._my_batches(trainer, first)
        chosen = [b for b in batches
                  if (not department or b.department == department)
                  and (not section or (b.section or "") == section)
                  and (not batch_code or b.batch_code == batch_code)]
        if not chosen:
            return []

        rows = (await self.db.execute(
            select(
                func.to_char(AttendanceRecord.attendance_date, "YYYY-MM").label("month"),
                AttendanceRecord.status,
                func.count(AttendanceRecord.id),
            )
            .join(ProjectBatchMember,
                  ProjectBatchMember.student_id == AttendanceRecord.student_id)
            .where(ProjectBatchMember.batch_id.in_([b.id for b in chosen]))
            .where(AttendanceRecord.attendance_date >= start)
            .where(AttendanceRecord.attendance_date <= last)
            .group_by("month", AttendanceRecord.status)
        )).all()

        buckets: Dict[str, Dict[str, int]] = {}
        for key, status, count in rows:
            bucket = buckets.setdefault(key, {"marked": 0, "absent": 0})
            bucket["marked"] += count
            if status == AttendanceStatus.ABSENT:
                bucket["absent"] += count

        out, cursor = [], start
        while cursor <= first:
            key = f"{cursor.year:04d}-{cursor.month:02d}"
            bucket = buckets.get(key)
            out.append({
                "month": key,
                "label": cursor.strftime("%b %Y"),
                "rate": (round((bucket["marked"] - bucket["absent"])
                               / bucket["marked"] * 100, 2)
                         if bucket and bucket["marked"] else None),
            })
            cursor = date(cursor.year + (cursor.month == 12),
                          (cursor.month % 12) + 1, 1)
        return out

    async def submit_session(self, trainer: User, *, on: date,
                             session: AttendanceSession) -> dict:
        """
        Mark a session finished.

        Corrections are still accepted afterwards. Submitting says the trainer
        is done, not that the record is frozen - a register nobody can fix
        just gets a wrong one filed instead.
        """
        log = await self._session_log(trainer, on, session, create=True)
        log.submitted_at = datetime.utcnow()
        await self.db.commit()
        logger.info(f"[Attendance] {trainer.email} submitted {session.value} for {on}")
        return {
            "date": on, "session": session.value,
            "submitted_at": log.submitted_at,
            "message": f"{SESSION_LABELS[session].split(' (')[0]} session submitted.",
        }

    async def _session_log(self, trainer: User, on: date,
                           session: AttendanceSession, create: bool = False):
        log = (await self.db.execute(
            select(AttendanceSessionLog)
            .where(AttendanceSessionLog.trainer_id == trainer.id)
            .where(AttendanceSessionLog.attendance_date == on)
            .where(AttendanceSessionLog.session == session)
        )).scalars().first()
        if log is None and create:
            log = AttendanceSessionLog(
                college_id=trainer.college_id, trainer_id=trainer.id,
                attendance_date=on, session=session, started_at=datetime.utcnow())
            self.db.add(log)
            await self.db.flush()
        return log

    async def trainer_mark(self, trainer: User, *, batch_code: Optional[str],
                           on: date, session: AttendanceSession,
                           marks: List[dict]) -> dict:
        """Record a trainer's register. Only their own batch's students count."""
        if not marks:
            raise AttendanceError("No attendance was marked.")

        batches = await self._my_batches(trainer, on)
        if batch_code:
            batches = [b for b in batches if b.batch_code == batch_code]
            if not batches:
                raise AttendanceError("That batch is not one you are assigned to.")
        if not batches:
            raise AttendanceError("You have no batches assigned to you yet.")

        # The register can span every batch this trainer has, so a mark is
        # attributed to whichever of them the student is actually in.
        members = {}
        for user, member_batch in (await self.db.execute(
            select(User, ProjectBatch)
            .join(ProjectBatchMember, ProjectBatchMember.student_id == User.id)
            .join(ProjectBatch, ProjectBatch.id == ProjectBatchMember.batch_id)
            .where(ProjectBatch.id.in_([b.id for b in batches]))
        )).all():
            members[str(user.id)] = member_batch
        batch = batches[0]

        cleaned: Dict[str, AttendanceStatus] = {}
        for entry in marks:
            if not isinstance(entry, dict):
                raise AttendanceError("Each mark needs a student and a status.")
            student_id = str(entry.get("student_id") or "").strip()
            if student_id not in members:
                raise AttendanceError(
                    "One of those students is not in this batch. Reload the register.")
            cleaned[student_id] = (parse_status(entry.get("status")),
                                   (entry.get("remarks") or "").strip()[:300] or None)

        existing = {
            str(r.student_id): r
            for r in (await self.db.execute(
                select(AttendanceRecord)
                .where(AttendanceRecord.student_id.in_(list(cleaned)))
                .where(AttendanceRecord.attendance_date == on)
                .where(AttendanceRecord.session == session)
            )).scalars().all()
        }

        created = updated = unchanged = 0
        rows = []
        for student_id, (status, remarks) in cleaned.items():
            record = existing.get(student_id)
            if record is None:
                created += 1
            elif record.status == status and record.remarks == remarks:
                unchanged += 1
                continue
            else:
                updated += 1
            batch_of = members[student_id]
            rows.append({
                "id": generate_uuid(),
                "college_id": batch_of.college_id,
                "student_id": student_id,
                "attendance_date": on,
                "session": session,
                "status": status,
                "remarks": remarks,
                "department": batch_of.department,
                "section": batch_of.section,
                "academic_year": batch_of.academic_year,
                "marked_by_id": trainer.id,
                "created_at": datetime.utcnow(),
            })

        if rows:
            # Upsert rather than read-then-insert. Marks are saved as they are
            # clicked, so two requests can be in flight for the same session at
            # once; both would find no row and both would insert, and the
            # second would die on uq_attendance_student_session. Letting the
            # database settle it means the later write simply wins.
            statement = pg_insert(AttendanceRecord).values(rows)
            await self.db.execute(statement.on_conflict_do_update(
                constraint="uq_attendance_student_session",
                set_={
                    "status": statement.excluded.status,
                    "remarks": statement.excluded.remarks,
                    "marked_by_id": statement.excluded.marked_by_id,
                },
            ))

        # First mark of the session starts its clock, which is what the screen
        # shows as "Started At".
        await self._session_log(trainer, on, session, create=True)

        await self.db.commit()
        logger.info(f"[Attendance] {trainer.email} {on} {session.value}: "
                    f"{created} new, {updated} changed, {unchanged} unchanged")
        return {
            "date": on, "session": session.value,
            "created": created, "updated": updated, "unchanged": unchanged,
            "message": (f"{SESSION_LABELS[session].split(' (')[0]} register saved for "
                        f"{created + updated + unchanged} students."),
        }

    @staticmethod
    def _counts(students: List[dict]) -> dict:
        counts = {s.value: 0 for s in AttendanceStatus}
        counts["unmarked"] = 0
        for student in students:
            if student["status"]:
                counts[student["status"]] += 1
            else:
                counts["unmarked"] += 1
        return counts

    async def _rates(self, student_ids: List[str], academic_year: str) -> Dict[str, float]:
        """Year-to-date rate per student. Absent is the only thing that lowers it."""
        if not student_ids:
            return {}
        rows = (await self.db.execute(
            select(
                AttendanceRecord.student_id,
                func.count().label("total"),
                func.sum(
                    case((AttendanceRecord.status == AttendanceStatus.ABSENT, 1), else_=0)
                ).label("absent"),
            )
            .where(AttendanceRecord.student_id.in_(student_ids))
            .where(AttendanceRecord.academic_year == academic_year)
            .group_by(AttendanceRecord.student_id)
        )).all()
        return {
            str(student_id): round((total - (absent or 0)) / total * 100, 1)
            for student_id, total, absent in rows if total
        }

    # --------------------------------------------------------------- write

    async def mark(
        self,
        user: User,
        *,
        department: str,
        year: Optional[str],
        section: Optional[str],
        academic_year: str,
        on: date,
        marks: List[dict],
        session: AttendanceSession,
    ) -> dict:
        """
        Record the register for a day.

        Only students actually in the cohort may be marked - a roll number in
        the payload that belongs to another section is refused rather than
        quietly written, because the denormalised department and section on
        each row are what every rate query trusts.
        """
        if not marks:
            raise AttendanceError("No attendance was marked.")

        cohort = await self._cohort(department, year, section, academic_year)
        by_id = {str(e.student_id): (e, u) for e, u in cohort}
        if not by_id:
            raise AttendanceError(
                f"No students are enrolled in {department} "
                f"{year or ''} section {section or '-'} for {academic_year}.".replace("  ", " "))

        cleaned: Dict[str, AttendanceStatus] = {}
        for entry in marks:
            if not isinstance(entry, dict):
                raise AttendanceError("Each mark needs a student and a status.")
            student_id = str(entry.get("student_id") or "").strip()
            if student_id not in by_id:
                raise AttendanceError(
                    "One of those students is not in this section. Reload the register.")
            cleaned[student_id] = parse_status(entry.get("status"))

        existing = {
            str(r.student_id): r
            for r in (await self.db.execute(
                select(AttendanceRecord)
                .where(AttendanceRecord.student_id.in_(list(cleaned)))
                .where(AttendanceRecord.attendance_date == on)
                .where(AttendanceRecord.session == session)
            )).scalars().all()
        }

        created = updated = unchanged = 0
        for student_id, status in cleaned.items():
            enrollment, _ = by_id[student_id]
            record = existing.get(student_id)
            if record is None:
                self.db.add(AttendanceRecord(
                    # Taken from the enrolment being marked, not the marker.
                    college_id=enrollment.college_id,
                    student_id=student_id,
                    attendance_date=on,
                    session=session,
                    status=status,
                    department=enrollment.department,
                    section=enrollment.section,
                    academic_year=academic_year,
                    marked_by_id=user.id if user else None,
                ))
                created += 1
                continue
            if record.status == status:
                unchanged += 1
                continue
            # Correcting a mark rewrites it and records who did: the register
            # states what happened that day, not the history of the argument.
            record.status = status
            record.marked_by_id = user.id if user else None
            record.department = enrollment.department
            record.section = enrollment.section
            record.academic_year = academic_year
            updated += 1

        await self.db.commit()
        logger.info(f"[Attendance] {getattr(user, 'email', 'system')} marked {on} for "
                    f"{department} {section or '-'}: {created} new, {updated} corrected")

        fresh = await self.roster(
            department=department, year=year, section=section,
            academic_year=academic_year, on=on)
        parts = []
        if created:
            parts.append(f"{created} marked")
        if updated:
            parts.append(f"{updated} corrected")
        if unchanged and not parts:
            parts.append("nothing changed")
        return {
            **fresh,
            "created": created,
            "updated": updated,
            "unchanged": unchanged,
            "message": f"Register saved for {on:%d %b %Y} — " + ", ".join(parts) + ".",
        }

    # ------------------------------------------------------- one student

    async def month_for_student(self, student: User, academic_year: str,
                                month: Optional[str] = None) -> dict:
        """
        One month of this student's own register, pivoted to a row per day.

        The trainer marks a session at a time, but a student reads a day at a
        time - "was I marked in on the 15th" is a question about a date, not
        about a forenoon. So the two session rows are folded into one row here
        rather than in the browser, where a missing row and a missing session
        would be indistinguishable.
        """
        rows = (await self.db.execute(
            select(AttendanceRecord)
            .where(AttendanceRecord.student_id == student.id)
            .where(AttendanceRecord.academic_year == academic_year)
            .order_by(AttendanceRecord.attendance_date.desc())
        )).scalars().all()

        # Months that actually have rows, newest first - so the picker cannot
        # offer a month that would come back empty.
        months = sorted({f"{r.attendance_date:%Y-%m}" for r in rows}, reverse=True)
        if month not in months:
            month = months[0] if months else None

        in_month = [r for r in rows if month
                    and f"{r.attendance_date:%Y-%m}" == month]

        # Who marked it. Looked up once per trainer rather than per row.
        marker_ids = {r.marked_by_id for r in in_month if r.marked_by_id}
        markers = {}
        if marker_ids:
            for user in (await self.db.execute(
                select(User).where(User.id.in_(marker_ids))
            )).scalars().all():
                markers[str(user.id)] = user.full_name

        by_day: dict = {}
        for r in in_month:
            day = by_day.setdefault(r.attendance_date, {
                "date": r.attendance_date,
                "day": r.attendance_date.strftime("%a"),
                "remarks": None,
                "trainer": None,
                "marked_at": None,
            })
            code, meaning = STATUS_CODES[r.status]
            day[r.session.value] = {
                "status": r.status.value,
                "code": code,
                "label": STATUS_LABELS[r.status],
                # The letter alone is not readable to somebody who has not
                # learnt the key, so it carries its meaning with it.
                "hint": meaning,
            }
            if r.remarks and not day["remarks"]:
                day["remarks"] = r.remarks
            if r.marked_by_id and not day["trainer"]:
                day["trainer"] = markers.get(str(r.marked_by_id))
            if r.created_at and (day["marked_at"] is None
                                 or r.created_at > day["marked_at"]):
                day["marked_at"] = r.created_at

        days = []
        for day in sorted(by_day.values(), key=lambda d: d["date"], reverse=True):
            # A session with no row is "no class", not an absence. Reading a
            # blank as absent is the one mistake that costs a student marks.
            for key in ("forenoon", "afternoon"):
                day.setdefault(key, None)
            days.append(day)

        classes = len(in_month)
        absent = sum(1 for r in in_month if r.status == AttendanceStatus.ABSENT)
        present = classes - absent
        pct = lambda n: round(n / classes * 100, 2) if classes else 0.0
        last = max((r.created_at for r in in_month if r.created_at), default=None)

        return {
            "month": month,
            "month_label": (datetime.strptime(month, "%Y-%m").strftime("%B %Y")
                            if month else None),
            "months": [{"value": m,
                        "label": datetime.strptime(m, "%Y-%m").strftime("%B %Y")}
                       for m in months],
            "sessions": [
                {"key": "forenoon", "name": "Session 1 (Morning)",
                 "time": "09:30 AM - 12:30 PM IST"},
                {"key": "afternoon", "name": "Session 2 (Afternoon)",
                 "time": "01:30 PM - 04:30 PM IST"},
            ],
            "totals": {
                # Classes counts sessions, not days: leaving after lunch costs
                # half a day, which is what it should cost.
                "classes": classes,
                "present": present,
                "present_pct": pct(present),
                "absent": absent,
                "absent_pct": pct(absent),
                "rate": pct(present),
                "floor": FLOOR,
                "below_floor": bool(classes) and pct(present) < FLOOR,
            },
            "last_updated": last,
            "days": days,
        }

    async def for_student(self, student: User, academic_year: str,
                          limit: int = 60) -> dict:
        """
        A student's own register.

        Absences are listed rather than only counted: a student challenging a
        rate needs to see which days it was built from.
        """
        records = (await self.db.execute(
            select(AttendanceRecord)
            .where(AttendanceRecord.student_id == student.id)
            .where(AttendanceRecord.academic_year == academic_year)
            .order_by(AttendanceRecord.attendance_date.desc(),
                      AttendanceRecord.session)
        )).scalars().all()

        total = len(records)
        absent = sum(1 for r in records if r.status == AttendanceStatus.ABSENT)
        late = sum(1 for r in records if r.status == AttendanceStatus.LATE)
        # A rate over sessions rather than days, now that a day has two: going
        # home after lunch costs half a day, which is what it should cost.
        rate = round((total - absent) / total * 100, 1) if total else None

        return {
            "academic_year": academic_year,
            "sessions_recorded": total,
            # The screen has always called this "Days recorded". It counts
            # sessions now, so the label there changes with it.
            "days_recorded": total,
            "days_covered": len({r.attendance_date for r in records}),
            "present": total - absent - late,
            "late": late,
            "absent": absent,
            "attendance_rate": rate,
            "floor": FLOOR,
            "below_floor": rate is not None and rate < FLOOR,
            "days": [
                {
                    "date": r.attendance_date,
                    # Two rows share a date, so each says which half it is.
                    "session": r.session.value,
                    "session_label": SESSION_LABELS[r.session].split(" (")[0],
                    "status": r.status.value,
                    "status_label": STATUS_LABELS[r.status],
                }
                for r in records[:limit]
            ],
            "absences": [
                {"date": r.attendance_date,
                 "session": SESSION_LABELS[r.session].split(" (")[0]}
                for r in records if r.status == AttendanceStatus.ABSENT
            ][:limit],
        }
