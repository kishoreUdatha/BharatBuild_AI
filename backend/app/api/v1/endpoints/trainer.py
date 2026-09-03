"""
Trainer Portal API - AI story approval.

The trainer is a faculty account acting in a different capacity, so these
routes use the faculty role check and the same per-batch authority model: an
identifier in the URL never grants access on its own.
"""

from datetime import date, datetime
from typing import List, Optional

from fastapi import (APIRouter, Depends, File, Form, HTTPException, Query, Request,
                     UploadFile, status)
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.core.logging_config import logger
from app.models.faculty import ProjectBatch
from app.models.user import User, UserRole
from app.modules.auth.dependencies import get_current_trainer
from app.services.ai_planning import AiPlanningService, PlanningError
from app.services.sprints import SprintService
from app.services.user_stories import StoryError, UserStoryService
from app.services.faculty_authority import FacultyAuthority
from app.services.git_commits import CommitError, GitCommitService
from app.models.faculty_import import ImportType
from app.services.faculty_imports import TEMPLATE_COLUMNS, FacultyImportService
from app.services.batch_creation import BatchCreationError, BatchCreationService
from app.services.tenancy import acting_college, tenant_of
from app.services.trainer_workspace import TrainerWorkspaceService
from app.services.batch_actions import (
    ActionError,
    cancel_review,
    complete_review,
    decide_document,
    reschedule_review,
)

router = APIRouter(prefix="/trainer", tags=["Trainer Portal"])


def _default_academic_year(today: Optional[datetime] = None) -> str:
    """June starts a new academic year, matching the faculty portal."""
    now = today or datetime.utcnow()
    start = now.year if now.month >= 6 else now.year - 1
    return f"{start}-{str(start + 1)[-2:]}"


class ReorderBody(BaseModel):
    """A column, top to bottom, after the drop."""
    story_ids: List[str] = Field(..., min_length=1)


class DecisionBody(BaseModel):
    decision: str = Field(..., description="approve | reject | reviewed | request_revision")
    note: Optional[str] = Field(None, max_length=2000)


class StoryUpdateBody(BaseModel):
    story_points: Optional[int] = None
    priority: Optional[str] = None
    trainer_comment: Optional[str] = Field(None, max_length=500)
    dependencies: Optional[str] = Field(None, max_length=300)


class MarkReviewedBody(BaseModel):
    story_ids: List[str] = Field(..., min_length=1)


class AttendanceBody(BaseModel):
    """One session's marks. The batch is optional - the register can span all."""
    batch_code: Optional[str] = None
    date: Optional[str] = None
    session: Optional[str] = None
    marks: List[dict] = Field(..., min_length=1, max_length=400)


class RegistrationDecisionBody(BaseModel):
    """What the trainer decided about a submitted registration."""
    decision: str = Field(..., description="request_changes | approve | reject")
    note: Optional[str] = Field(None, max_length=2000)


class SubmitSessionBody(BaseModel):
    date: Optional[str] = None
    session: str


class GitConnectBody(BaseModel):
    """Where the batch's code lives, and whether to mint a fresh secret."""
    repo_url: Optional[str] = Field(None, max_length=500)
    rotate_secret: bool = False


async def _require_batch(db: AsyncSession, identifier: str, user: User, *, manage: bool = False):
    """Load the batch and check this trainer is entitled to it."""
    service = AiPlanningService(db)
    batch = await service.load_batch(identifier)
    if batch is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Batch not found")

    authority = FacultyAuthority(db)
    allowed = (await authority.can_manage(user, batch) if manage
               else await authority.can_view(user, batch))
    if not allowed:
        logger.warning(f"[Trainer] {user.email} denied "
                       f"{'write' if manage else 'read'} access to {batch.batch_code}")
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=FacultyAuthority.denial("approve stories") if manage
            else "This batch belongs to a department you are not attached to.",
        )
    return service, batch


@router.get("/batches/{identifier}/stories")
async def get_story_board(
    identifier: str,
    search: Optional[str] = Query(None),
    status_filter: Optional[str] = Query(None, alias="status",
                                         description="all | needs_review | reviewed | rejected"),
    epic: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    confidence: Optional[str] = Query(None, description="high | medium | low"),
    selected: Optional[str] = Query(None, description="Story key to open"),
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """Everything the AI Story Approval screen shows."""
    service, batch = await _require_batch(db, identifier, current_user)
    try:
        return await service.board(
            batch, search=search, status=status_filter, epic=epic,
            priority=priority, confidence=confidence, selected_key=selected,
        )
    except Exception as exc:
        logger.error(f"[Trainer] Story board failed for {identifier}: {type(exc).__name__}: {exc}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                            detail="Could not load the story board.")


@router.post("/batches/{identifier}/stories/{story_id}/decision")
async def decide_story(
    identifier: str,
    story_id: str,
    payload: DecisionBody,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """Record the trainer's decision on one story."""
    service, batch = await _require_batch(db, identifier, current_user, manage=True)
    try:
        result = await service.decide(batch, story_id, payload.decision, current_user,
                                      note=payload.note)
    except PlanningError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))
    logger.info(f"[Trainer] {current_user.email} {payload.decision} {result['key']} "
                f"on {batch.batch_code}")
    return result


@router.patch("/batches/{identifier}/stories/{story_id}")
async def update_story(
    identifier: str,
    story_id: str,
    payload: StoryUpdateBody,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """Edit points, priority, dependencies or the trainer's comment."""
    service, batch = await _require_batch(db, identifier, current_user, manage=True)
    try:
        return await service.update(batch, story_id,
                                    payload.model_dump(exclude_unset=True), current_user)
    except PlanningError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))


@router.post("/batches/{identifier}/stories/mark-reviewed")
async def mark_reviewed(
    identifier: str,
    payload: MarkReviewedBody,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """Mark the selected stories reviewed in one action."""
    service, batch = await _require_batch(db, identifier, current_user, manage=True)
    return await service.mark_reviewed(batch, payload.story_ids, current_user)


@router.post("/batches/{identifier}/stories/move-to-backlog")
async def move_to_backlog(
    identifier: str,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """Move the approved set to the product backlog."""
    service, batch = await _require_batch(db, identifier, current_user, manage=True)
    try:
        result = await service.move_to_backlog(batch, current_user)
    except PlanningError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))
    logger.info(f"[Trainer] {current_user.email} moved {result['count']} stories "
                f"to the backlog on {batch.batch_code}")
    return result


class RegenerateBody(BaseModel):
    scope: str = Field("pending", description="pending | all")
    confirm: bool = Field(False, description="Required when replacing decided stories")
    model: Optional[str] = Field(None, description="Override the planning model")


@router.get("/batches/{identifier}/stories/regenerate-preview")
async def regenerate_preview(
    identifier: str,
    scope: str = Query("pending", description="pending | all"),
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """What a regeneration at this scope would replace, before running it."""
    service, batch = await _require_batch(db, identifier, current_user, manage=True)
    return await service.preview_regeneration(batch, scope)


@router.post("/batches/{identifier}/stories/regenerate")
async def regenerate_drafts(
    identifier: str,
    payload: RegenerateBody,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """
    Redraft stories from the batch's approved project details.

    Default scope replaces only stories nobody has ruled on. Replacing decided
    ones needs `confirm`, because it discards recorded trainer decisions.
    Whatever the model returns lands as Needs Review.
    """
    service, batch = await _require_batch(db, identifier, current_user, manage=True)
    try:
        return await service.regenerate(
            batch, current_user, scope=payload.scope,
            confirm=payload.confirm, model=payload.model,
        )
    except PlanningError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))


# ============================================
# User Stories - the product backlog
# ============================================
#
# AI planning hands a batch's approved stories over and stops. These routes are
# what happens next: assigning the work, scheduling it into sprints and
# tracking it to done. They read and write the same story rows, so a story does
# not exist twice and its review history stays attached to it.


class StoryCreateBody(BaseModel):
    title: str = Field(..., min_length=3, max_length=240)
    narrative: Optional[str] = Field(None, max_length=2000)
    epic_key: Optional[str] = Field(None, max_length=20)
    story_points: int = Field(0, ge=0, le=100)
    priority: str = Field("medium", description="high | medium | low")
    story_type: str = Field("story", description="story | task | bug | spike")
    status: str = Field("to_do", description="to_do | in_progress | in_review | done")
    assignee_id: Optional[str] = None
    sprint_id: Optional[str] = None
    dependencies: Optional[str] = Field(None, max_length=300)
    due_date: Optional[date] = None
    acceptance_criteria: List[str] = Field(default_factory=list)


class StoryPatchBody(BaseModel):
    """
    Every field optional, and absence means "leave it".

    The route sends `exclude_unset` through, so posting `assignee_id: null`
    unassigns a story while omitting the key entirely leaves it alone. Without
    that distinction there would be no way to clear a field at all.
    """
    title: Optional[str] = Field(None, min_length=3, max_length=240)
    narrative: Optional[str] = Field(None, max_length=2000)
    story_points: Optional[int] = Field(None, ge=0, le=100)
    priority: Optional[str] = None
    story_type: Optional[str] = None
    status: Optional[str] = None
    assignee_id: Optional[str] = None
    sprint_id: Optional[str] = None
    dependencies: Optional[str] = Field(None, max_length=300)
    trainer_comment: Optional[str] = Field(None, max_length=500)
    due_date: Optional[date] = None


class StoryCommentBody(BaseModel):
    body: str = Field(..., min_length=1, max_length=2000)


class SprintCreateBody(BaseModel):
    name: str = Field(..., min_length=2, max_length=80)
    goal: Optional[str] = Field(None, max_length=300)
    start_date: Optional[date] = None
    end_date: Optional[date] = None
    state: str = Field("planned", description="planned | active | completed")


async def _backlog_batch(db: AsyncSession, identifier: str, user: User, *, manage: bool = False):
    """Load the batch with its roster, and check this trainer is entitled to it."""
    service = UserStoryService(db)
    batch = await service.load_batch(identifier)
    if batch is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Batch not found")

    authority = FacultyAuthority(db)
    allowed = (await authority.can_manage(user, batch) if manage
               else await authority.can_view(user, batch))
    if not allowed:
        logger.warning(f"[Trainer] {user.email} denied "
                       f"{'write' if manage else 'read'} access to {batch.batch_code} backlog")
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=FacultyAuthority.denial("change stories") if manage
            else "This batch belongs to a department you are not attached to.",
        )
    return service, batch


@router.get("/batches/{identifier}/user-stories")
async def get_user_stories(
    identifier: str,
    search: Optional[str] = Query(None),
    status_filter: Optional[str] = Query(None, alias="status",
                                         description="to_do | in_progress | in_review | done"),
    epic: Optional[str] = Query(None),
    assignee: Optional[str] = Query(None, description="A student id, or 'unassigned'"),
    sprint: Optional[str] = Query(None, description="A sprint id, or 'unscheduled'"),
    priority: Optional[str] = Query(None),
    points: Optional[str] = Query(None),
    story_type: Optional[str] = Query(None, alias="type"),
    created_by: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None, description="YYYY-MM-DD, on created date"),
    date_to: Optional[str] = Query(None, description="YYYY-MM-DD, on created date"),
    sort: str = Query("created_desc"),
    page: int = Query(1, ge=1),
    per_page: int = Query(10, ge=1, le=200),
    selected: Optional[str] = Query(None, description="Story id to open in the panel"),
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """Everything the User Stories screen renders for one batch."""
    service, batch = await _backlog_batch(db, identifier, current_user)
    try:
        return await service.board(
            batch, search=search, status=status_filter, epic=epic, assignee=assignee,
            sprint=sprint, priority=priority, points=points, story_type=story_type,
            created_by=created_by, date_from=date_from, date_to=date_to,
            sort=sort, page=page, per_page=per_page, selected=selected,
        )
    except Exception as exc:
        logger.error(f"[Trainer] User stories failed for {identifier}: "
                     f"{type(exc).__name__}: {exc}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                            detail="Could not load the user stories.")


@router.post("/batches/{identifier}/user-stories", status_code=status.HTTP_201_CREATED)
async def add_user_story(
    identifier: str,
    payload: StoryCreateBody,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """
    Write a story straight onto the backlog.

    A trainer writing one is the approval: the review gate exists to stop the
    model's drafts arriving unread, and there is no draft here to read.
    """
    service, batch = await _backlog_batch(db, identifier, current_user, manage=True)
    try:
        result = await service.create(batch, payload.model_dump(), current_user)
    except StoryError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))
    logger.info(f"[Trainer] {current_user.email} added {result['key']} to {batch.batch_code}")
    return result


@router.patch("/batches/{identifier}/user-stories/{story_id}")
async def patch_user_story(
    identifier: str,
    story_id: str,
    payload: StoryPatchBody,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """Assign, schedule, re-point or move one story, recording what changed."""
    service, batch = await _backlog_batch(db, identifier, current_user, manage=True)
    try:
        return await service.update(batch, story_id,
                                    payload.model_dump(exclude_unset=True), current_user)
    except StoryError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))


@router.post("/batches/{identifier}/user-stories/{story_id}/comments",
             status_code=status.HTTP_201_CREATED)
async def comment_on_user_story(
    identifier: str,
    story_id: str,
    payload: StoryCommentBody,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """Leave a note on a story. Comments are kept, never edited."""
    service, batch = await _backlog_batch(db, identifier, current_user, manage=True)
    try:
        return await service.comment(batch, story_id, payload.body, current_user)
    except StoryError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))


@router.post("/batches/{identifier}/sprints", status_code=status.HTTP_201_CREATED)
async def add_sprint(
    identifier: str,
    payload: SprintCreateBody,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """Open a sprint on this batch so stories have somewhere to be scheduled."""
    _, batch = await _backlog_batch(db, identifier, current_user, manage=True)
    try:
        return await SprintService(db).create(batch, payload.model_dump())
    except StoryError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))


@router.post("/batches/{identifier}/user-stories/reorder")
async def reorder_user_stories(
    identifier: str,
    payload: ReorderBody,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """Set the manual order of one board column after a drag."""
    _, batch = await _require_batch(db, identifier, current_user, manage=True)
    try:
        return await UserStoryService(db).reorder(batch, payload.story_ids)
    except StoryError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))


@router.get("/batches/{identifier}/user-stories/template.xlsx")
async def user_stories_import_template(
    identifier: str,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """
    The import workbook: the story sheet and the field guide beside it.

    The same shape the importer reads, so a trainer who downloads this, fills
    it in and uploads it cannot be told their columns are wrong.
    """
    await _backlog_batch(db, identifier, current_user)
    return StreamingResponse(
        iter([UserStoryService.import_template()]),
        media_type=("application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"),
        headers={"Content-Disposition":
                 'attachment; filename="bharatbuild-user-stories-template.xlsx"'},
    )


@router.get("/batches/{identifier}/user-stories.csv")
async def export_user_stories(
    identifier: str,
    search: Optional[str] = Query(None),
    status_filter: Optional[str] = Query(None, alias="status"),
    epic: Optional[str] = Query(None),
    assignee: Optional[str] = Query(None),
    sprint: Optional[str] = Query(None),
    priority: Optional[str] = Query(None),
    story_type: Optional[str] = Query(None, alias="type"),
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """The stories as CSV - the same rows, and the same filters, as the screen."""
    import csv as _csv
    import io as _io

    service, batch = await _backlog_batch(db, identifier, current_user)
    try:
        # per_page is the service ceiling: an export of a filtered view should
        # be the whole view, not whichever page the trainer happened to be on.
        data = await service.board(
            batch, search=search, status=status_filter, epic=epic, assignee=assignee,
            sprint=sprint, priority=priority, story_type=story_type, per_page=200,
        )
    except Exception as exc:
        logger.error(f"[Trainer] Story export failed for {identifier}: "
                     f"{type(exc).__name__}: {exc}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                            detail="Could not export these stories.")

    buffer = _io.StringIO()
    writer = _csv.writer(buffer)
    writer.writerow(UserStoryService.export_columns())
    for row in data["rows"]:
        writer.writerow(UserStoryService.export_row(row))
    buffer.seek(0)

    logger.info(f"[Trainer] {current_user.email} exported {len(data['rows'])} stories "
                f"from {batch.batch_code}")
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition":
                 f'attachment; filename="{batch.batch_code}-user-stories.csv"'},
    )


@router.post("/batches/{identifier}/user-stories/import")
async def import_user_stories(
    identifier: str,
    file: UploadFile = File(..., description="The trainer template, .xlsx or .csv"),
    dry_run: bool = Query(False, description="Validate and preview without writing"),
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """
    Read the trainer's import sheet.

    `dry_run` is the Validate step of the template's workflow: every row is
    parsed and checked and nothing is written, so the trainer sees exactly
    what confirming would create. Either way a row with a problem is reported
    against its row number and skipped, and the rest of the sheet still lands.
    """
    service, batch = await _backlog_batch(db, identifier, current_user, manage=True)
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                            detail="That file is empty.")
    try:
        result = await service.import_rows(batch, raw, file.filename or "",
                                           current_user, dry_run=dry_run)
    except StoryError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))
    if not dry_run:
        logger.info(f"[Trainer] {current_user.email} imported {result['count']} stories "
                    f"into {batch.batch_code}")
    return result


# ============================================
# Sprints
# ============================================
#
# When a story is meant to happen. A sprint's progress is its stories' progress,
# never a number typed in beside them - so the board and the sprint cannot
# disagree about how far along the batch is.
#
# The work underneath a story is a sub-task, and lives on the story itself:
# see /stories/{id}/tasks in endpoints/stories.py.


class SprintPatchBody(BaseModel):
    name: Optional[str] = Field(None, min_length=2, max_length=80)
    goal: Optional[str] = Field(None, max_length=300)
    start_date: Optional[date] = None
    end_date: Optional[date] = None
    state: Optional[str] = Field(None, description="planned | active | completed")


class SprintScheduleBody(BaseModel):
    """`sprint_id: null` takes the stories out of every sprint."""
    sprint_id: Optional[str] = None
    story_ids: List[str] = Field(..., min_length=1)


@router.get("/batches/{identifier}/git")
async def git_connection(
    identifier: str,
    request: Request,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """The webhook URL and secret for this batch's repository, and what it has sent."""
    _, batch = await _require_batch(db, identifier, current_user, manage=True)
    return await GitCommitService(db).connection(batch, str(request.base_url))


@router.post("/batches/{identifier}/git")
async def git_connect(
    identifier: str,
    body: GitConnectBody,
    request: Request,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """Record the repository link and mint (or rotate) its push secret."""
    _, batch = await _require_batch(db, identifier, current_user, manage=True)
    service = GitCommitService(db)
    try:
        await service.connect(batch, repo_url=body.repo_url,
                              rotate=body.rotate_secret, actor=current_user)
    except CommitError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))
    return await service.connection(batch, str(request.base_url))


@router.get("/batches/{identifier}/sprints")
async def get_sprints(
    identifier: str,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """Every sprint on this batch, with what its stories say about it."""
    _, batch = await _backlog_batch(db, identifier, current_user)
    try:
        return await SprintService(db).board(batch)
    except Exception as exc:
        logger.error(f"[Trainer] Sprint board failed for {identifier}: "
                     f"{type(exc).__name__}: {exc}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                            detail="Could not load the sprints.")


@router.patch("/batches/{identifier}/sprints/{sprint_id}")
async def patch_sprint(
    identifier: str,
    sprint_id: str,
    payload: SprintPatchBody,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """Rename a sprint, move its dates, or start and close it."""
    _, batch = await _backlog_batch(db, identifier, current_user, manage=True)
    try:
        return await SprintService(db).update(batch, sprint_id,
                                              payload.model_dump(exclude_unset=True))
    except StoryError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))


@router.get("/batches/{identifier}/sprints/{sprint_id}/burndown")
async def get_sprint_burndown(
    identifier: str,
    sprint_id: str,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """Points remaining on each day of a sprint, against the ideal line."""
    _, batch = await _backlog_batch(db, identifier, current_user)
    try:
        return await SprintService(db).burndown(batch, sprint_id)
    except StoryError as exc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(exc))
    except Exception as exc:
        logger.error(f"[Trainer] Burndown failed for {identifier}/{sprint_id}: "
                     f"{type(exc).__name__}: {exc}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                            detail="Could not build the burndown.")


@router.post("/batches/{identifier}/sprints/schedule")
async def schedule_stories(
    identifier: str,
    payload: SprintScheduleBody,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """Move backlog stories into a sprint, or out of every sprint."""
    _, batch = await _backlog_batch(db, identifier, current_user, manage=True)
    try:
        return await SprintService(db).schedule(batch, payload.sprint_id, payload.story_ids)
    except StoryError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))


# ============================================
# Trainer workspace
# ============================================


def _workspace_route(name: str, builder: str, description: str, extra: str = ""):
    """
    Registers one workspace screen.

    Every one is scoped the same way - to the batches this trainer answers for -
    so they share a shape rather than repeating the scoping in six places.
    """
    async def endpoint(
        academic_year: Optional[str] = Query(None),
        search: Optional[str] = Query(None),
        status_filter: Optional[str] = Query(None, alias="status"),
        batch: Optional[str] = Query(None),
        current_user: User = Depends(get_current_trainer),
        db: AsyncSession = Depends(get_db),
    ):
        service = TrainerWorkspaceService(db)
        year = academic_year or _default_academic_year()
        kwargs = {}
        if "search" in extra:
            kwargs["search"] = search
        if "status" in extra:
            kwargs["status"] = status_filter
        if "batch" in extra:
            kwargs["batch"] = batch
        try:
            return await getattr(service, builder)(current_user, year, **kwargs)
        except Exception as exc:
            logger.error(f"[Trainer] {name} failed for {current_user.email}: "
                         f"{type(exc).__name__}: {exc}")
            raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                                detail=f"Could not load {name}.")

    endpoint.__name__ = f"get_trainer_{name.replace(' ', '_')}"
    endpoint.__doc__ = description
    router.get(f"/{name.replace(' ', '-')}")(endpoint)


for _name, _builder, _extra, _doc in [
    ("home", "home", "", "The trainer's worklist: what is waiting on them, most urgent first."),
    ("reviews", "reviews", "status", "Scheduled, overdue and completed reviews across those batches."),
    ("student-work", "student_work", "batch", "Team composition, stage progress and submissions."),
    ("evidence", "evidence", "status", "Documents, base papers and submissions in one list."),
    ("reports", "reports", "", "Section and stage roll-ups across the trainer's batches."),
    ("settings", "settings", "", "The trainer's own roles and the scope they grant."),
]:
    _workspace_route(_name, _builder, _doc, _extra)



@router.get("/batches/{identifier}/builder")
async def get_batch_builder(
    identifier: str,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """
    The build workspace a batch shares, if the team has opened one.

    Reached through the same authority check as the rest of the batch, so a
    trainer sees the workspaces of the batches they teach and no others.
    """
    from app.services.batch_projects import (describe, project_for_batch,
                                              repo_of)
    _, batch = await _backlog_batch(db, identifier, current_user)
    return {
        "batch_code": batch.batch_code,
        "workspace": describe(await project_for_batch(db, batch)),
        "repo": await repo_of(db, batch),
    }


@router.post("/batches/{identifier}/builder")
async def open_batch_builder(
    identifier: str,
    request: Request,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """
    Open a batch's workspace, creating it the first time.

    A trainer can start it rather than waiting for a student to, which is what
    makes "every active batch has a workspace" true in practice without
    creating one for every batch the moment it is imported.
    """
    from app.services.batch_projects import (describe, ensure_repo,
                                              open_for_batch)
    _, batch = await _backlog_batch(db, identifier, current_user, manage=True)
    project, created = await open_for_batch(db, current_user, batch)
    # And somewhere to push. Connected already: untouched. Not connected, and
    # the college has installed the GitHub App: created in their organisation,
    # with the team added and the push webhook set.
    repo = await ensure_repo(db, current_user, batch, str(request.base_url))
    return {
        "batch_code": batch.batch_code,
        "created": created,
        "workspace": describe(project),
        "repo": repo,
    }


@router.get("/colleges")
async def trainer_colleges(
    academic_year: Optional[str] = Query(None),
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """
    The colleges this trainer teaches at, and the sections in each.

    Drives the college picker. Read from assignments rather than from the
    account, because a trainer belongs to no college - these are the ones they
    have been given work in.
    """
    from app.models.college import College
    from app.models.trainer_assignment import TrainerAssignment
    from app.models.user import UserRole

    year = academic_year or _default_academic_year()

    # A manager has no assignments - they run every college - so the picker is
    # built from the colleges themselves. Without this their picker is empty
    # and every screen shows nothing.
    if current_user.role in (UserRole.MANAGER, UserRole.ADMIN):
        colleges = (await db.execute(
            select(College)
            .where(College.is_active.is_(True))
            .where(College.is_self_serve.is_(False))
            .order_by(College.name)
        )).scalars().all()
        return {
            "academic_year": year,
            "colleges": [{"id": str(c.id), "name": c.name, "code": c.code,
                          "sections": ["Every branch and section"]}
                         for c in colleges],
            "must_choose": len(colleges) > 1,
        }

    rows = (await db.execute(
        select(TrainerAssignment, College)
        .join(College, College.id == TrainerAssignment.college_id)
        .where(TrainerAssignment.trainer_id == current_user.id)
        .where(TrainerAssignment.is_active.is_(True))
        .where(TrainerAssignment.academic_year == year)
        .order_by(College.name, TrainerAssignment.department,
                  TrainerAssignment.section)
    )).all()

    grouped: dict = {}
    for assignment, college in rows:
        entry = grouped.setdefault(str(college.id), {
            "id": str(college.id),
            "name": college.name,
            "code": college.code,
            "sections": [],
        })
        entry["sections"].append(
            f"{assignment.department}-{assignment.section}"
            if assignment.section else assignment.department)

    colleges = list(grouped.values())
    return {
        "academic_year": year,
        "colleges": colleges,
        # A trainer at one college has nothing to choose, so the picker can
        # stay out of their way.
        "must_choose": len(colleges) > 1,
    }


@router.get("/trainers")
async def list_college_trainers(
    academic_year: Optional[str] = Query(None),
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """
    The trainers working in the college being looked at.

    A manager's screens otherwise merge every trainer's work into one pile,
    which answers "how is this college doing" but not "how is this trainer
    doing" - and the second is what a conversation with a trainer is about.
    Empty for anybody else: a trainer has no business enumerating colleagues.
    """
    if current_user.role != UserRole.MANAGER:
        return {"trainers": [], "can_filter": False}

    year = academic_year or _default_academic_year()
    from app.services.tenancy import tenants_of
    from app.models.trainer_assignment import TrainerAssignment
    colleges = tenants_of(current_user)

    rows = (await db.execute(
        select(TrainerAssignment, User)
        .join(User, User.id == TrainerAssignment.trainer_id)
        .where(TrainerAssignment.is_active.is_(True))
        .where(TrainerAssignment.academic_year == year)
        .where(TrainerAssignment.college_id.in_(colleges))
        .order_by(User.full_name)
    )).all()

    authority = FacultyAuthority(db)
    grouped: dict = {}
    for assignment, trainer in rows:
        entry = grouped.setdefault(str(trainer.id), {
            "id": str(trainer.id),
            "name": trainer.full_name or trainer.email,
            "email": trainer.email,
            "scope": [],
        })
        # What they cover here, in the words the assignment screen uses.
        entry["scope"].append(
            f"{assignment.department}-{assignment.section}"
            if assignment.section else assignment.department or "Whole college")

    for trainer_id, entry in grouped.items():
        entry["batches"] = len(await authority.batch_ids_for_trainer(
            trainer_id, colleges, year))

    trainers = sorted(grouped.values(), key=lambda t: t["name"].lower())
    return {
        "academic_year": year,
        "trainers": trainers,
        # Nothing to choose between when one trainer works here, so the picker
        # can stay out of the way.
        "can_filter": len(trainers) > 1,
    }


@router.get("/pending")
async def trainer_pending(
    academic_year: Optional[str] = Query(None),
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """
    Counts for the sidebar badges. Cheap on purpose - it runs on every page.
    """
    year = academic_year or _default_academic_year()
    return await TrainerWorkspaceService(db).pending(current_user, year)


@router.get("/batches")
async def list_trainer_batches(
    academic_year: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    department: Optional[str] = Query(None),
    section: Optional[str] = Query(None),
    batch_no: Optional[str] = Query(None),
    project_status: Optional[str] = Query(None, description="In Progress | Review | Completed"),
    semester: Optional[str] = Query(None),
    guide: Optional[str] = Query(None),
    batch_type: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None, description="YYYY-MM-DD, on created date"),
    date_to: Optional[str] = Query(None, description="YYYY-MM-DD, on created date"),
    sort: str = Query("latest", description="latest | oldest | code | progress | students"),
    page: int = Query(1, ge=1),
    per_page: int = Query(10, ge=1, le=100),
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """
    The batches this trainer guides, reviews or coordinates.

    Filtering, sorting and paging happen over that scope - never the whole
    department - so no filter can widen what the trainer is allowed to see.
    """
    try:
        return await TrainerWorkspaceService(db).batches(
            current_user,
            academic_year or _default_academic_year(),
            search=search, department=department, section=section, batch_no=batch_no,
            project_status=project_status, semester=semester, guide=guide,
            batch_type=batch_type, date_from=date_from, date_to=date_to,
            sort=sort, page=page, per_page=per_page,
        )
    except Exception as exc:
        logger.error(f"[Trainer] batches failed for {current_user.email}: "
                     f"{type(exc).__name__}: {exc}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                            detail="Could not load your batches.")


# ============================================
# Trainer decisions
# ============================================
#
# These write the same rows the Faculty Portal writes, through the same rules in
# `batch_actions`, behind the same per-batch authority check. A trainer acting
# here and a coordinator acting there produce an identical record.


class ReviewCompleteBody(BaseModel):
    score: Optional[float] = Field(None, ge=0, le=100)
    remarks: Optional[str] = Field(None, max_length=2000)


class ReviewRescheduleBody(BaseModel):
    scheduled_at: datetime


class ReviewCancelBody(BaseModel):
    reason: str = Field(..., min_length=4, max_length=500)


class DocumentDecisionBody(BaseModel):
    document_id: str
    decision: str = Field(..., description="verify | request_changes")
    note: Optional[str] = Field(None, max_length=2000)


async def _manageable_batch(db: AsyncSession, identifier: str, user: User, action: str):
    """
    Load a batch with its documents and reviews, and check the caller owns it.

    Its own query rather than the batch-detail loader: that one does not fetch
    reviews, and adding them there would make seven read-only tabs pay for rows
    they never use. Both collections are eager-loaded because lazy-loading
    either one under async raises MissingGreenlet.
    """
    batch = (await db.execute(
        select(ProjectBatch)
        .where((ProjectBatch.batch_code == identifier.strip())
               | (ProjectBatch.join_code == identifier.strip().upper()))
        .options(
            selectinload(ProjectBatch.documents),
            selectinload(ProjectBatch.reviews),
        )
    )).scalars().first()
    if batch is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Batch not found")
    if not await FacultyAuthority(db).can_manage(user, batch):
        logger.warning(f"[Trainer] {user.email} denied write access to {batch.batch_code}")
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN,
                            detail=FacultyAuthority.denial(action))
    return batch


@router.post("/batches/{identifier}/reviews/{review_id}/complete")
async def trainer_complete_review(
    identifier: str,
    review_id: str,
    payload: ReviewCompleteBody,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """Record that a review happened, with its score and remarks."""
    batch = await _manageable_batch(db, identifier, current_user, "complete reviews")
    try:
        result = await complete_review(db, batch, review_id, current_user,
                                       score=payload.score, remarks=payload.remarks)
    except ActionError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))
    logger.info(f"[Trainer] {current_user.email} completed {result['review_type']} "
                f"on {batch.batch_code}")
    return result


@router.post("/batches/{identifier}/reviews/{review_id}/reschedule")
async def trainer_reschedule_review(
    identifier: str,
    review_id: str,
    payload: ReviewRescheduleBody,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """Move a scheduled review to a new date."""
    batch = await _manageable_batch(db, identifier, current_user, "reschedule reviews")
    try:
        return await reschedule_review(db, batch, review_id, payload.scheduled_at)
    except ActionError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))


@router.post("/batches/{identifier}/reviews/{review_id}/cancel")
async def trainer_cancel_review(
    identifier: str,
    review_id: str,
    payload: ReviewCancelBody,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """Cancel a scheduled review, with the reason on the record."""
    batch = await _manageable_batch(db, identifier, current_user, "cancel reviews")
    try:
        return await cancel_review(db, batch, review_id, payload.reason)
    except ActionError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))


@router.post("/batches/{identifier}/documents/decide")
async def trainer_decide_document(
    identifier: str,
    payload: DocumentDecisionBody,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """
    Verify a document or send it back.

    Identical rules to the Faculty Portal, because it is the same function: a
    verified document stays locked here too.
    """
    batch = await _manageable_batch(db, identifier, current_user, "decide on documents")
    try:
        result = await decide_document(db, batch, payload.document_id, payload.decision,
                                       current_user, note=payload.note)
    except ActionError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))
    logger.info(f"[Trainer] {current_user.email} {payload.decision} "
                f"{result['name']} on {batch.batch_code}")
    return result


# ============================================
# Batch roster: export, import, create
# ============================================
#
# Export is available to every trainer - it only re-serves the batches the
# workspace already showed them. Import and creation are department-level acts,
# so they reuse `can_act_for_department` rather than trusting the trainer role:
# guiding a batch is not the same authority as forming one.


class CreateTrainerBatchBody(BaseModel):
    department: str
    year: str
    semester: str
    section: str
    project_type: str = "Major Project"
    guide_id: Optional[str] = None
    team_size: int = Field(4, ge=2, le=8)
    project_fee: int = Field(15000, ge=0, le=1000000)
    count: int = Field(1, ge=1, le=20)
    academic_year: Optional[str] = None


async def _require_department(db: AsyncSession, user: User, department: str, year: str) -> None:
    """Department-level authority, the same rule the faculty portal applies."""
    if not await FacultyAuthority(db).can_act_for_department(user, department, year):
        logger.warning(f"[Trainer] {user.email} denied department action on {department}")
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=f"You are not a coordinator for {department}, so you cannot do this.",
        )


async def _reachable_departments(db: AsyncSession, user: User, year: str) -> List[str]:
    """
    The branches this caller could be asked about.

    A college's own faculty carry their branch on the account, which is what
    this used to read. Platform staff do not - a trainer and a manager belong
    to no college, so their `department` is null and the list came back empty,
    which made every capability false and left Import Batches and Create Batch
    disabled for exactly the people whose job it is to run them.

    For them the branches are whatever the college they are working in
    actually runs. Authority is still decided one branch at a time by
    `can_act_for_department`; this only says which questions to ask.
    """
    if user.role in (UserRole.TRAINER, UserRole.MANAGER) or user.is_superuser:
        from app.services.tenancy import tenants_of
        rows = (await db.execute(
            select(ProjectBatch.department)
            .where(ProjectBatch.academic_year == year)
            .where(ProjectBatch.college_id.in_(tenants_of(user)))
            .distinct()
        )).scalars().all()
        return sorted({r for r in rows if r})
    return sorted({d for d in [user.department] if d})


@router.get("/capabilities")
async def get_trainer_capabilities(
    academic_year: Optional[str] = Query(None),
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """
    What this trainer may do beyond viewing, so the UI can disable an action
    rather than offer one that would only come back 403.
    """
    year = academic_year or _default_academic_year()
    authority = FacultyAuthority(db)
    departments = await _reachable_departments(db, current_user, year)
    allowed = []
    for code in departments:
        if await authority.can_act_for_department(current_user, code, year):
            allowed.append(code)
    return {
        "academic_year": year,
        "departments": departments,
        "manageable_departments": allowed,
        "can_manage_department": bool(allowed),
    }


@router.get("/batches.csv")
async def export_trainer_batches(
    academic_year: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """The trainer's own batches as CSV - the same rows and scope as the screen."""
    import csv as _csv
    import io as _io

    year = academic_year or _default_academic_year()
    try:
        data = await TrainerWorkspaceService(db).batches(current_user, year, search=search)
    except Exception as exc:
        logger.error(f"[Trainer] Batch export failed for {current_user.email}: "
                     f"{type(exc).__name__}: {exc}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                            detail="Could not export your batches.")

    columns = [
        "batch_code", "title", "department", "section", "year", "my_role",
        "members", "team_size", "progress", "registration_status",
        "reviews_pending", "reviews_overdue", "stories_total", "stories_needs_review",
    ]
    buffer = _io.StringIO()
    writer = _csv.writer(buffer)
    writer.writerow(columns)
    for row in data.get("rows", []):
        writer.writerow([row.get(c, "") for c in columns])
    buffer.seek(0)

    rows = len(data.get("rows", []))
    logger.info(f"[Trainer] {current_user.email} exported {rows} batches")
    filename = f"my-batches-{year}.csv"
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# The sample sheet mirrors the layout colleges actually hand out: one row per
# batch, the team across "Student N" columns as "<roll> - <name>", and an
# optional "Email N" beside each. It is filled in rather than header-only, so
# the shape of a cell is visible without reading any documentation.
SAMPLE_HEADERS = [
    "projid", "Section", "Batch No", "Guide Name",
    "Student 1", "Email 1", "Student 2", "Email 2",
    "Student 3", "Email 3", "Student 4", "Email 4",
]

SAMPLE_ROWS = [
    ["CSE-D-D1", "CSE - D", "D1", "Dr Kavitha",
     "23K91A05L5 - Royyala Sindhuja", "sindhuja@sgit.ac.in",
     "23K91A05P2 - T Meenakshi", "meenakshi@sgit.ac.in",
     "23K91A05Q2 - Thorlikonda Srihari", "srihari@sgit.ac.in",
     "23K91A05R1 - Vattela Akhil", "akhil@sgit.ac.in"],
    # A three-member team: trailing student columns may simply be blank.
    ["CSE-D-D2", "CSE - D", "D2", "",
     "23K91A05T1 - Sai-Kumar Reddy", "saikumar@sgit.ac.in",
     "23K91A05T2 - Anita Rao", "anita@sgit.ac.in",
     "23K91A05T3 - Bhavana Iyer", "bhavana@sgit.ac.in",
     "", ""],
]

SAMPLE_NOTES = [
    ("projid", "The batch code. Created if it does not exist yet."),
    ("Section", 'Department and section together, e.g. "CSE - D".'),
    ("Batch No", "Your own numbering. Recorded, not required."),
    ("Guide Name", "Optional. Matched to staff by full name, only when unambiguous."),
    ("Student N", 'Roll and name in one cell: "23K91A05L5 - Royyala Sindhuja". '
                  "Hyphen, en dash or a plain space all work, and the name may itself "
                  "be hyphenated. Student 1 is taken as the team leader."),
    ("Email N", "Optional, pairs with Student N by number. Leave it out and a login "
                "is generated from the roll number using your college's own email "
                "domain, which the student can change after signing in. Give a real "
                "address here when you have one."),
    ("Blank cells", "A blank Student column means a smaller team, not an error."),
    ("Re-uploading", "Safe. A student already in the batch is reported as a duplicate, "
                     "never added twice."),
]


@router.get("/imports/template")
async def download_trainer_import_template(
    file_format: str = Query("xlsx", alias="format", description="xlsx or csv"),
    current_user: User = Depends(get_current_trainer),
):
    """A filled-in sample allocation sheet, as .xlsx or .csv."""
    import csv as _csv
    import io as _io

    if file_format not in ("xlsx", "csv"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                            detail="Format must be xlsx or csv")

    if file_format == "csv":
        buffer = _io.StringIO()
        writer = _csv.writer(buffer)
        writer.writerow(SAMPLE_HEADERS)
        writer.writerows(SAMPLE_ROWS)
        buffer.seek(0)
        return StreamingResponse(
            iter([buffer.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition":
                     'attachment; filename="batch-allocation-sample.csv"'},
        )

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Batch Allocation"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    sheet.append(SAMPLE_HEADERS)
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    for row in SAMPLE_ROWS:
        sheet.append(row)

    for index, header in enumerate(SAMPLE_HEADERS, start=1):
        longest = max([len(header)] + [len(str(r[index - 1])) for r in SAMPLE_ROWS])
        sheet.column_dimensions[get_column_letter(index)].width = min(longest + 3, 34)
    sheet.freeze_panes = "A2"

    # Guidance lives on a second sheet: the importer reads the active sheet, so
    # notes beside the data would be parsed as columns.
    notes = workbook.create_sheet("How to fill this in")
    notes.append(["Column", "What to put in it"])
    for cell in notes[1]:
        cell.font = header_font
        cell.fill = header_fill
    for column, text in SAMPLE_NOTES:
        notes.append([column, text])
    notes.column_dimensions["A"].width = 16
    notes.column_dimensions["B"].width = 96
    for row in notes.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    workbook.active = 0   # the data sheet must be the one an upload reads

    stream = _io.BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return StreamingResponse(
        iter([stream.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 'attachment; filename="batch-allocation-sample.xlsx"'},
    )


@router.post("/imports", status_code=status.HTTP_201_CREATED)
async def create_trainer_import(
    file: UploadFile = File(..., description="CSV or XLSX batch allocation"),
    import_type: str = Form("batch_allocation"),
    academic_year: Optional[str] = Form(None),
    department: Optional[str] = Form(None),
    # Which college this roster is for. A trainer works across several, and
    # filing one against the wrong institution is not a mistake that shows up
    # until somebody's students are in another college's batches.
    college_id: Optional[str] = Form(None),
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """
    Apply a batch-allocation roster.

    Rows are validated one at a time; a bad row is recorded against the run
    rather than failing the whole upload.
    """
    try:
        kind = ImportType(import_type)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unknown import type.",
        )

    year = academic_year or _default_academic_year()
    target = department or current_user.department
    if not target:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                            detail="No department to import into.")
    await _require_department(db, current_user, target, year)

    content = await file.read()
    if not content:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                            detail="The uploaded file is empty.")

    service = FacultyImportService(db, acting_college(current_user, college_id))
    try:
        run = await service.run_import(
            filename=file.filename or "upload.csv",
            content=content,
            import_type=kind,
            academic_year=year,
            department=target,
            actor=current_user,
        )
    except (BatchCreationError, ValueError) as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))
    except Exception as exc:
        logger.error(f"[Trainer] Import failed for {current_user.email}: "
                     f"{type(exc).__name__}: {exc}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                            detail="The import could not be processed")

    logger.info(
        f"[Trainer] {current_user.email} imported {run.import_code}: "
        f"{run.rows_imported} in, {run.rows_failed} failed, {run.rows_duplicate} duplicate"
    )
    return await service.detail(str(run.id))


@router.post("/batches", status_code=status.HTTP_201_CREATED)
async def create_trainer_batches(
    payload: CreateTrainerBatchBody,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """Form one or more empty batches for a section of a department they run."""
    year = payload.academic_year or _default_academic_year()
    await _require_department(db, current_user, payload.department, year)
    try:
        return await BatchCreationService(db).create(
            current_user,
            department=payload.department,
            year=payload.year,
            semester=payload.semester,
            section=payload.section,
            project_type=payload.project_type,
            guide_id=payload.guide_id,
            team_size=payload.team_size,
            project_fee=payload.project_fee,
            count=payload.count,
            academic_year=year,
        )
    except (BatchCreationError, ValueError) as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))


# ============================================
# Import results
# ============================================


@router.get("/imports")
async def list_trainer_imports(
    limit: int = Query(20, ge=1, le=100),
    page: int = Query(1, ge=1),
    search: Optional[str] = Query(None, max_length=100),
    import_type: Optional[str] = Query(None),
    status_filter: Optional[str] = Query(None, alias="status"),
    academic_year: Optional[str] = Query(None),
    # Which college this import is for. A trainer works across several;
    # everyone else has one and may leave it out.
    college_id: Optional[str] = Query(None),
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """
    Import history for the college being worked in, newest first.

    A trainer sees their own runs. An import rewrites batch allocations, and
    whose upload did it is the first thing anyone asks when a roster looks
    wrong - so the list answers that before it is asked.

    A manager sees everybody's, because overseeing what the trainers did is
    the job; the uploader is on every row either way.
    """
    if current_user.role == UserRole.TRAINER:
        whose = str(current_user.id)
    else:
        # A manager sees everybody's, or one trainer's when they have focused
        # on one - the same choice the rest of their screens are following.
        focus = getattr(current_user, "_focus_trainer_id", None)
        whose = str(focus) if focus else None
    service = FacultyImportService(db, acting_college(current_user, college_id))
    return await service.build(
        academic_year=academic_year or _default_academic_year(),
        imported_by=whose,
        search=search,
        import_type=import_type,
        status=status_filter,
        page=page,
        per_page=limit,
    )


@router.get("/imports/{run_id}")
async def get_trainer_import(
    run_id: str,
    # Which college this import is for. A trainer works across several;
    # everyone else has one and may leave it out.
    college_id: Optional[str] = Query(None),
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """
    One import run, with what it did to the batches.

    The summary is what the results screen counts on: the row tallies come from
    the run, the batch tallies from re-reading the sheet it stored.
    """
    service = FacultyImportService(db, acting_college(current_user, college_id))
    detail = await service.detail(run_id)
    if detail is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Import not found")
    summary = await service.allocation_summary(run_id)
    return {**detail, "summary": summary or {}}


@router.get("/imports/{run_id}/report.csv")
async def download_trainer_import_report(
    run_id: str,
    # Which college this import is for. A trainer works across several;
    # everyone else has one and may leave it out.
    college_id: Optional[str] = Query(None),
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """The run as a CSV: the tallies, every batch it touched, and every issue."""
    import csv as _csv
    import io as _io

    service = FacultyImportService(db, acting_college(current_user, college_id))
    detail = await service.detail(run_id)
    if detail is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Import not found")
    summary = await service.allocation_summary(run_id) or {}

    buffer = _io.StringIO()
    writer = _csv.writer(buffer)

    writer.writerow(["Import", detail["import_code"]])
    writer.writerow(["File", detail["file_name"]])
    writer.writerow(["Imported by", detail.get("imported_by") or ""])
    writer.writerow(["Imported on", detail.get("started_at") or ""])
    writer.writerow(["Status", detail["status"]])
    writer.writerow([])
    writer.writerow(["Total rows", detail["rows_total"]])
    writer.writerow(["Imported", detail["rows_imported"]])
    writer.writerow(["Duplicate", detail["rows_duplicate"]])
    writer.writerow(["Failed", detail["rows_failed"]])
    writer.writerow(["Batches created", summary.get("batches_created", 0)])
    writer.writerow(["Batches updated", summary.get("batches_updated", 0)])
    writer.writerow(["Guides assigned", summary.get("guides_assigned", 0)])

    if summary.get("batches"):
        writer.writerow([])
        writer.writerow(["Batch", "Title", "Department", "Section", "Students", "Guide", "Outcome"])
        for b in summary["batches"]:
            writer.writerow([b["batch_code"], b.get("title") or "", b.get("department") or "",
                             b.get("section") or "", b["students"], b.get("guide") or "",
                             b["outcome"]])

    if detail.get("issues"):
        writer.writerow([])
        writer.writerow(["Sheet row", "Field", "Value", "Severity", "Message"])
        for i in detail["issues"]:
            writer.writerow([i.get("row"), i.get("field") or "", i.get("value") or "",
                             i.get("severity") or "", i.get("message") or ""])

    buffer.seek(0)
    filename = f"{detail['import_code']}-report.csv"
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ============================================
# Attendance - twice a day, for the batches this trainer guides
# ============================================


@router.get("/attendance")
async def attendance_register(
    batch_code: Optional[str] = Query(None),
    date: Optional[str] = Query(None, description="YYYY-MM-DD, default today"),
    session: Optional[str] = Query(None, description="forenoon | afternoon"),
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """
    The register for one of this trainer's batches, in one session.

    No batch given means the first one they guide, and no session means the
    one the clock is nearest - so opening the screen during a class shows the
    right sheet without anyone choosing anything.
    """
    from app.services import attendance as attendance_service

    try:
        return await attendance_service.AttendanceService(db).trainer_roster(
            current_user, batch_code=batch_code,
            on=attendance_service.parse_day(date),
            session=attendance_service.parse_session(session))
    except attendance_service.AttendanceError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))


@router.post("/attendance")
async def attendance_save(
    body: AttendanceBody,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """Save a session's register. Saving it again corrects it."""
    from app.services import attendance as attendance_service

    try:
        return await attendance_service.AttendanceService(db).trainer_mark(
            current_user, batch_code=body.batch_code,
            on=attendance_service.parse_day(body.date),
            session=attendance_service.parse_session(body.session),
            marks=body.marks)
    except attendance_service.AttendanceError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))


@router.get("/attendance/day")
async def attendance_day(
    date: Optional[str] = Query(None, description="YYYY-MM-DD, default today"),
    department: Optional[str] = Query(None, description="Branch, e.g. CSE"),
    section: Optional[str] = Query(None),
    batch_code: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(10, ge=1, le=100),
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """
    Both sessions for every student this trainer is responsible for.

    One screen rather than two: the same faces are marked morning and
    afternoon, and flipping between separate registers is how half-days get
    mis-recorded.
    """
    from app.services import attendance as attendance_service

    try:
        return await attendance_service.AttendanceService(db).day_register(
            current_user, on=attendance_service.parse_day(date),
            department=department, section=section, batch_code=batch_code,
            page=page, per_page=per_page)
    except attendance_service.AttendanceError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))


@router.post("/attendance/submit")
async def attendance_submit(
    body: SubmitSessionBody,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """Mark a session finished. Corrections are still accepted afterwards."""
    from app.services import attendance as attendance_service

    try:
        return await attendance_service.AttendanceService(db).submit_session(
            current_user, on=attendance_service.parse_day(body.date),
            session=attendance_service.parse_session(body.session))
    except attendance_service.AttendanceError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))


@router.get("/attendance/template.xlsx")
async def attendance_template(
    date: Optional[str] = Query(None),
    batch_code: Optional[str] = Query(None),
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """The import sheet, already carrying this trainer's students."""
    from app.services import attendance as attendance_service

    try:
        day = attendance_service.parse_day(date)
        blob = await attendance_service.AttendanceService(db).import_template(
            current_user, on=day, batch_code=batch_code)
    except attendance_service.AttendanceError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))
    return StreamingResponse(
        iter([blob]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="attendance-{day}.xlsx"'},
    )


@router.post("/attendance/import")
async def attendance_import(
    file: UploadFile = File(..., description="The attendance sheet, .xlsx or .csv"),
    date: Optional[str] = Query(None),
    dry_run: bool = Query(True, description="Validate and preview without writing"),
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """
    Read a filled-in register.

    Validate first, write second - the same two steps as the story importer,
    so a sheet with a bad row is seen before anything is recorded rather than
    after.
    """
    from app.services import attendance as attendance_service

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                            detail="That file is empty.")
    try:
        return await attendance_service.AttendanceService(db).import_register(
            current_user, raw=raw, filename=file.filename or "sheet.xlsx",
            on=attendance_service.parse_day(date), dry_run=dry_run)
    except attendance_service.AttendanceError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))


@router.get("/attendance/month")
async def attendance_month(
    month: Optional[str] = Query(None, description="YYYY-MM, default this month"),
    department: Optional[str] = Query(None),
    section: Optional[str] = Query(None),
    batch_code: Optional[str] = Query(None),
    search: Optional[str] = Query(None, description="Name or roll number"),
    page: int = Query(1, ge=1),
    per_page: int = Query(25, ge=1, le=100),
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """
    A month of attendance, one row per student.

    The view that answers "how has this student been over the month", which a
    daily register cannot.
    """
    from app.core.institution_time import local_today
    from app.services import attendance as attendance_service

    today = local_today()
    if month:
        try:
            year, month_number = (int(part) for part in month.split("-", 1))
        except (ValueError, TypeError):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                                detail="Month should look like 2026-09.")
    else:
        year, month_number = today.year, today.month

    try:
        return await attendance_service.AttendanceService(db).month_register(
            current_user, year=year, month=month_number, department=department,
            section=section, batch_code=batch_code, search=search,
            page=page, per_page=per_page)
    except attendance_service.AttendanceError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))


@router.get("/attendance/summary")
async def attendance_summary(
    month: Optional[str] = Query(None, description="YYYY-MM, default this month"),
    department: Optional[str] = Query(None),
    section: Optional[str] = Query(None),
    batch_code: Optional[str] = Query(None),
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """The month rolled up over every student the filters select."""
    from app.core.institution_time import local_today
    from app.services import attendance as attendance_service

    today = local_today()
    if month:
        try:
            year, month_number = (int(part) for part in month.split("-", 1))
        except (ValueError, TypeError):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                                detail="Month should look like 2026-09.")
    else:
        year, month_number = today.year, today.month

    try:
        return await attendance_service.AttendanceService(db).month_summary(
            current_user, year=year, month=month_number, department=department,
            section=section, batch_code=batch_code)
    except attendance_service.AttendanceError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))


@router.delete("/batches/{identifier}/user-stories/{story_id}")
async def delete_user_story(
    identifier: str,
    story_id: str,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """
    Remove a story from the backlog.

    Its criteria, comments and history go with it. Sub-tasks and any commits
    that named it are detached and kept - a commit is evidence that somebody
    did the work, and deleting the ticket must not delete that.
    """
    service, batch = await _require_batch(db, identifier, current_user, manage=True)
    try:
        return await UserStoryService(db).delete_story(batch, story_id, current_user)
    except StoryError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))


@router.post("/batches/{identifier}/registration/decide")
async def decide_registration(
    identifier: str,
    body: RegistrationDecisionBody,
    current_user: User = Depends(get_current_trainer),
    db: AsyncSession = Depends(get_db),
):
    """
    Approve a registration, reject it, or send it back for changes.

    The student's locked Project Setup says "ask them to send it back", and
    until now only a faculty coordinator could: the decision lived on the
    faculty queue behind `get_current_faculty`, which a trainer is not. A team
    whose guide is a trainer had nobody able to unlock their own proposal.

    Approval runs the same checklist the faculty queue does, through the same
    service - a trainer approving must not be an easier gate than a
    coordinator approving.
    """
    from app.models.faculty import BatchRegistrationStatus
    from app.services.faculty_workflow import FacultyWorkflowService

    _, batch = await _require_batch(db, identifier, current_user, manage=True)

    if body.decision not in {"approve", "reject", "request_changes"}:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Decision must be approve, reject or request_changes.")

    # Only something actually with the guide can be decided on. Sending back a
    # draft the team is still writing would lock nothing and confuse everyone.
    if batch.registration_status not in (BatchRegistrationStatus.SUBMITTED,
                                         BatchRegistrationStatus.PENDING_APPROVAL):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"{batch.batch_code} is not waiting on you - it is "
                   f"{batch.registration_status.value.replace('_', ' ')}.")

    service = FacultyWorkflowService(db, batch.college_id)
    try:
        result = await service.decide([str(batch.id)], body.decision, body.note)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))

    if result["skipped"]:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST,
                            detail=result["skipped"][0]["reason"])

    await db.refresh(batch)
    logger.info(f"[Trainer] {current_user.email} {body.decision} on {batch.batch_code}")
    wording = {
        "request_changes": "sent back to the team to edit",
        "approve": "approved",
        "reject": "rejected",
    }[body.decision]
    return {
        "batch_code": batch.batch_code,
        "registration_status": batch.registration_status.value,
        "message": f"{batch.batch_code} {wording}.",
    }
